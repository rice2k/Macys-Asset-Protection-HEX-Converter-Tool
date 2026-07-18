from __future__ import annotations

import csv
import json
import os
import random
import re
import subprocess
import sys
import webbrowser
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from PIL import Image, ImageTk

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
except Exception:  # pragma: no cover - optional dependency in dev
    colors = ParagraphStyle = None
    letter = landscape = getSampleStyleSheet = None
    SimpleDocTemplate = Paragraph = Spacer = Table = TableStyle = None

try:
    import xlrd
except Exception:  # pragma: no cover - optional dependency in dev
    xlrd = None

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except Exception:  # pragma: no cover - drag/drop is optional
    DND_FILES = None
    TkinterDnD = None


BLUEWAVE_URL = "http://ma000xsblw1001/"
PROJECT_URL = "https://github.com/rice2k/Macys-Asset-Protection-HEX-Converter-Tool"
CONTACT_EMAIL = "christopher.schumacher@macys.com"
APP_DISPLAY_NAME = "Macy's Asset Protection - China Grove Hex Converter Utility"
APP_SHORT_NAME = "Macy's AP China Grove Hex Utility"
APP_VERSION = "1.1.6"
APP_STATE_DIR = Path(os.environ.get("APPDATA", str(Path.home()))) / "AP_Access_Control_Converter"
SETTINGS_FILE = APP_STATE_DIR / "settings.json"
EXPORT_TYPE_CHOICES = ["Excel Workbook (.xlsx)", "CSV Report (.csv)", "TXT Report (.txt)", "PDF Report (.pdf)"]
UI_BG = "#f3f5f8"
UI_HEADER = "#ffffff"
UI_SURFACE = "#ffffff"
UI_SURFACE_ALT = "#f8fafc"
UI_INPUT = "#fbfcff"
UI_BORDER = "#d8dee8"
UI_BORDER_DARK = "#c5ccd8"
UI_TEXT = "#1f2937"
UI_MUTED = "#667085"
UI_RED = "#e51b2d"
UI_RED_DARK = "#b91525"
UI_RED_SOFT = "#fff1f2"
UI_BLUE = "#0b66c3"
UI_BLUE_SOFT = "#eef6ff"
UI_GREEN_SOFT = "#ecfdf3"
UI_GREEN_TEXT = "#166534"
UI_WARN_SOFT = "#fff7df"
UI_WARN_TEXT = "#7a4b00"
UI_BAD_SOFT = "#fff1f0"
UI_BAD_TEXT = "#991b1b"
UI_CARD_SHADOW = "#e6ebf2"
TkRoot = TkinterDnD.Tk if TkinterDnD else tk.Tk


class ToolTip:
    def __init__(self, widget: tk.Widget, text: str, delay: int = 650) -> None:
        self.widget = widget
        self.text = text
        self.delay = delay
        self.after_id: str | None = None
        self.window: tk.Toplevel | None = None
        widget.bind("<Enter>", self.schedule, add="+")
        widget.bind("<Leave>", self.hide, add="+")
        widget.bind("<ButtonPress>", self.hide, add="+")

    def schedule(self, _event: Any | None = None) -> None:
        self.cancel()
        self.after_id = self.widget.after(self.delay, self.show)

    def cancel(self) -> None:
        if self.after_id:
            self.widget.after_cancel(self.after_id)
            self.after_id = None

    def show(self) -> None:
        if self.window or not self.text:
            return
        x = self.widget.winfo_rootx() + 14
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.window = tk.Toplevel(self.widget)
        self.window.wm_overrideredirect(True)
        self.window.wm_geometry(f"+{x}+{y}")
        frame = tk.Frame(self.window, bg=UI_TEXT, highlightthickness=1, highlightbackground=UI_BORDER_DARK)
        frame.pack()
        tk.Label(
            frame,
            text=self.text,
            bg=UI_TEXT,
            fg="#ffffff",
            justify="left",
            wraplength=280,
            padx=10,
            pady=7,
            font=("Segoe UI", 9),
        ).pack()

    def hide(self, _event: Any | None = None) -> None:
        self.cancel()
        if self.window:
            self.window.destroy()
            self.window = None


@dataclass
class ConvertedRow:
    line: int
    raw: str
    hex_value: str
    facility: int
    card: int
    converted_at: str
    suggestions: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class InvalidRow:
    line: int
    raw: str
    reason: str
    converted_at: str


@dataclass
class UnconvertRow:
    line: int
    raw: str
    facility: int
    card: int
    hex_value: str
    converted_at: str
    warnings: list[str] = field(default_factory=list)


@dataclass
class HistoryEntry:
    timestamp: str
    action: str
    total: int
    valid: int
    invalid: int
    warnings: int


def asset_path(name: str) -> Path:
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / "assets" / name
    return Path(__file__).resolve().parent / "src" / "assets" / name


def valid_hex8(value: str) -> bool:
    return bool(re.fullmatch(r"[0-9A-Fa-f]{8}", value or ""))


def hex_to_fc_cn(hex_value: str) -> tuple[int, int]:
    clean = hex_value.strip().upper()
    if not valid_hex8(clean):
        raise ValueError("Hex value must be exactly 8 characters.")
    number = int(clean, 16)
    return (number >> 16) & 0xFFFF, number & 0xFFFF


def strict_int(value: str) -> int:
    text = str(value).strip()
    if not re.fullmatch(r"\d+", text):
        raise ValueError("Facility Code and Card Number must be whole numbers.")
    return int(text)


def fc_cn_to_hex(fc: str | int, cn: str | int) -> str:
    facility = strict_int(str(fc))
    card = strict_int(str(cn))
    if facility < 0 or facility > 65535 or card < 0 or card > 65535:
        raise ValueError("FC and CN must be between 0 and 65535.")
    return f"{((facility << 16) | card):08X}"


def clean_candidate_line(line: str) -> dict[str, Any]:
    original = str(line or "")
    cleaned = original.strip()
    suggestions: list[str] = []
    if not cleaned:
        return {"original": original, "cleaned": "", "extracted": "", "suggestions": suggestions}

    excel_number = re.search(r"\b(\d{8})\.0+\b", cleaned)
    numeric_token = re.search(r"\b(\d{8})\b", cleaned)
    numeric_split = re.search(r"\b(\d{4})[\s-]+(\d{4})\b", cleaned)
    token = re.search(r"\b([0-9A-Fa-f]{8})\b", cleaned)
    split_token = re.search(r"\b([0-9A-Fa-f]{4})[\s-]+([0-9A-Fa-f]{4})\b", cleaned)
    extracted = ""
    if excel_number:
        extracted = excel_number.group(1)
        suggestions.append("Cleaned Excel numeric ID.")
    elif numeric_token:
        extracted = numeric_token.group(1)
        if cleaned != numeric_token.group(1):
            suggestions.append("Extracted 8-digit ID from full text.")
    elif numeric_split:
        extracted = f"{numeric_split.group(1)}{numeric_split.group(2)}"
        suggestions.append("Joined a split 8-digit ID.")
    elif token:
        extracted = token.group(1).upper()
        if cleaned != token.group(1):
            suggestions.append("Extracted 8-character ID from full text.")
    elif split_token:
        extracted = f"{split_token.group(1)}{split_token.group(2)}".upper()
        suggestions.append("Joined a split 8-character ID.")

    fallback = re.sub(r"\s+", " ", re.sub(r"[\t,;\"']", " ", cleaned)).strip()
    compact = re.sub(r"[^0-9A-Fa-f]", "", fallback).upper()
    if not extracted and len(compact) == 8:
        extracted = compact
        if compact != cleaned.upper():
            suggestions.append("Removed spaces or punctuation around the ID.")

    return {"original": original, "cleaned": fallback, "extracted": extracted, "suggestions": suggestions}


def invalid_reason(original: str) -> str:
    raw = str(original or "").strip()
    if not raw:
        return "Blank line"
    hex_like = re.sub(r"[^0-9A-Fa-f]", "", raw)
    if len(hex_like) == 0:
        return "No hex characters found"
    if len(hex_like) < 8:
        return "Too short; needs exactly 8 characters"
    if len(hex_like) > 8 and not re.search(r"\b[0-9A-Fa-f]{8}\b", raw):
        return "Too long or mixed with extra characters"
    if re.search(r"[^0-9A-Fa-f\s,;\"'\-]", raw):
        return "Contains invalid characters"
    return "Must be exactly 8 characters using 0-9 and A-F"


def unusual_warnings(hex_value: str, facility: int, card: int) -> list[str]:
    warnings: list[str] = []
    if re.fullmatch(r"0{8}", hex_value):
        warnings.append("All zeros is unusual.")
    if re.fullmatch(r"F{8}", hex_value, re.I):
        warnings.append("All Fs is unusual.")
    if re.fullmatch(r"\d{8}", hex_value) and not re.fullmatch(r"88\d{6}", hex_value):
        warnings.append("Numeric ID does not start with the common 88 prefix.")
    if facility == 0:
        warnings.append("Facility Code is 0.")
    if card == 0:
        warnings.append("Card Number is 0.")
    if facility > 60000 or card > 60000:
        warnings.append("Very high FC or CN value.")
    return warnings


def convert_lines(text: str, converted_at: str) -> tuple[list[ConvertedRow], list[InvalidRow]]:
    converted: list[ConvertedRow] = []
    invalid: list[InvalidRow] = []
    seen_hex: dict[str, int] = {}
    for idx, line in enumerate(str(text or "").splitlines(), start=1):
        prepared = clean_candidate_line(line)
        if not prepared["original"].strip():
            continue
        hex_value = prepared["extracted"].strip().upper()
        if not valid_hex8(hex_value):
            invalid.append(InvalidRow(idx, prepared["original"], invalid_reason(prepared["original"]), converted_at))
            continue
        facility, card = hex_to_fc_cn(hex_value)
        warnings = unusual_warnings(hex_value, facility, card)
        if hex_value in seen_hex:
            warnings.append(f"Duplicate of line {seen_hex[hex_value]}.")
        else:
            seen_hex[hex_value] = idx
        converted.append(
            ConvertedRow(
                line=idx,
                raw=prepared["original"],
                hex_value=hex_value,
                facility=facility,
                card=card,
                converted_at=converted_at,
                suggestions=list(prepared["suggestions"]),
                warnings=warnings,
            )
        )
    return converted, invalid


def parse_fc_cn_line(line: str) -> tuple[int, int]:
    raw = str(line or "")
    labeled = re.search(r"\bfc\b\D*(\d{1,5})\D+\bcn\b\D*(\d{1,5})", raw, re.I)
    if labeled:
        return strict_int(labeled.group(1)), strict_int(labeled.group(2))
    numbers = re.findall(r"\b\d{1,5}\b", raw)
    if len(numbers) < 2:
        raise ValueError("Needs Facility Code and Card Number.")
    return strict_int(numbers[0]), strict_int(numbers[1])


def unconvert_lines(text: str, converted_at: str) -> tuple[list[UnconvertRow], list[InvalidRow]]:
    converted: list[UnconvertRow] = []
    invalid: list[InvalidRow] = []
    seen_hex: dict[str, int] = {}
    for idx, line in enumerate(str(text or "").splitlines(), start=1):
        raw = str(line or "").strip()
        if not raw:
            continue
        try:
            facility, card = parse_fc_cn_line(raw)
            hex_value = fc_cn_to_hex(facility, card)
        except ValueError as exc:
            invalid.append(InvalidRow(idx, raw, str(exc), converted_at))
            continue
        warnings = unusual_warnings(hex_value, facility, card)
        if hex_value in seen_hex:
            warnings.append(f"Duplicate of line {seen_hex[hex_value]}.")
        else:
            seen_hex[hex_value] = idx
        converted.append(UnconvertRow(idx, raw, facility, card, hex_value, converted_at, warnings))
    return converted, invalid


def cell_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def normalize_header(value: Any) -> str:
    text = cell_text(value).lower().replace("&nbsp;", " ")
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[^a-z0-9#]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


NAME_RULES = [
    ("candidate name", 100),
    ("colleague name", 92),
    ("employee name", 90),
    ("associate name", 88),
    ("full name", 80),
    ("name", 60),
]

ID_RULES = [
    ("colleague #", 100),
    ("colleague number", 98),
    ("colleague no", 96),
    ("colleague id", 94),
    ("employee id", 88),
    ("employee number", 84),
    ("associate id", 80),
    ("badge id", 72),
]

BAD_NAME_PATTERNS = [
    re.compile(r"\bpl name\b"),
    re.compile(r"\bmanager\b"),
    re.compile(r"\blead\b"),
    re.compile(r"\bsupervisor\b"),
    re.compile(r"\brecruiter\b"),
    re.compile(r"\btrainer\b"),
    re.compile(r"\binterviewer\b"),
]

BAD_ID_PATTERNS = [
    re.compile(r"\bcandidate id\b"),
    re.compile(r"\breq id\b"),
    re.compile(r"\bjob id\b"),
    re.compile(r"\bposition id\b"),
]


def header_score(normalized: str, rules: list[tuple[str, int]], bad_patterns: list[re.Pattern[str]]) -> int:
    if not normalized:
        return 0
    score = 0
    for label, points in rules:
        if normalized == label:
            score = max(score, points + 20)
        elif label in normalized:
            score = max(score, points)
    for pattern in bad_patterns:
        if pattern.search(normalized):
            score -= 120
    return score


def find_header_pair(rows: list[list[str]]) -> dict[str, Any] | None:
    candidates: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[:8]):
        for col_index, cell in enumerate(row):
            normalized = normalize_header(cell)
            name_score = header_score(normalized, NAME_RULES, BAD_NAME_PATTERNS)
            id_score = header_score(normalized, ID_RULES, BAD_ID_PATTERNS)
            if name_score > 0 or id_score > 0:
                candidates.append(
                    {
                        "row": row_index,
                        "col": col_index,
                        "normalized": normalized,
                        "name_score": name_score,
                        "id_score": id_score,
                    }
                )
    best = None
    for name in [c for c in candidates if c["name_score"] > 0]:
        for emp_id in [c for c in candidates if c["id_score"] > 0]:
            score = name["name_score"] + emp_id["id_score"]
            score -= abs(name["row"] - emp_id["row"]) * 12
            score -= max(0, abs(name["col"] - emp_id["col"]) - 8) * 2
            if name["normalized"] == "candidate name":
                score += 60
            if emp_id["normalized"] == "colleague #":
                score += 60
            if name["normalized"] == "candidate name" and emp_id["normalized"] == "colleague #":
                score += 200
            if not best or score > best["score"]:
                best = {
                    "score": score,
                    "name_index": name["col"],
                    "id_index": emp_id["col"],
                    "header_row": max(name["row"], emp_id["row"]),
                    "name_header": name["normalized"],
                    "id_header": emp_id["normalized"],
                    "exact": name["normalized"] == "candidate name" and emp_id["normalized"] == "colleague #",
                }
    return best


def extract_eight_digit_id(value: Any) -> str:
    raw = cell_text(value)
    excel_number = re.search(r"\b(\d{8})\.0+\b", raw)
    if excel_number:
        return excel_number.group(1)
    preferred = re.search(r"\b(88\d{6})\b", raw)
    if preferred:
        return preferred.group(1)
    split = re.search(r"\b(\d{4})[\s-]+(\d{4})\b", raw)
    if split:
        return f"{split.group(1)}{split.group(2)}"
    generic = re.search(r"\b(\d{8})\b", raw)
    return generic.group(1) if generic else ""


def is_noise_name(value: Any) -> bool:
    text = cell_text(value)
    normalized = normalize_header(text)
    if not normalized:
        return True
    if any(pattern.search(normalized) for pattern in BAD_NAME_PATTERNS):
        return True
    if re.search(r"\b(candidate id|colleague #|employee id|employee number|requisition|job title|position|department|location|status|email|phone)\b", normalized):
        return True
    return bool(re.fullmatch(r"\d+", text) or "@" in text)


def is_likely_person_name(value: Any) -> bool:
    text = cell_text(value)
    if not text or is_noise_name(text):
        return False
    if len(text) < 3 or len(text) > 80 or re.match(r"^\d", text):
        return False
    if not re.search(r"[A-Za-z]", text):
        return False
    if re.search(r"\b(warehouse|associate|operator|shift|full time|part time|china grove|salisbury|north carolina)\b", text, re.I):
        return False
    return bool(re.fullmatch(r"[A-Za-z][A-Za-z\s.\-'\u2019]+", text))


def clean_imported_pair(name: Any, emp_id: Any) -> tuple[str, str] | None:
    clean_name = cell_text(name)
    clean_id = extract_eight_digit_id(emp_id)
    if not is_likely_person_name(clean_name) or not clean_id:
        return None
    return clean_name, clean_id


def dedupe_pairs(pairs: list[tuple[str, str]]) -> list[tuple[str, str]]:
    seen = set()
    out = []
    for name, emp_id in pairs:
        key = emp_id.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append((name, emp_id))
    return out


def fallback_extract_pairs(rows: list[list[str]]) -> list[tuple[str, str]]:
    pairs = []
    for row in rows:
        ids = [(idx, extract_eight_digit_id(cell)) for idx, cell in enumerate(row)]
        ids = [(idx, emp_id) for idx, emp_id in ids if emp_id]
        if not ids:
            continue
        names = [(idx, cell_text(cell)) for idx, cell in enumerate(row) if is_likely_person_name(cell)]
        if not names:
            continue
        best_id = ids[0]
        names.sort(key=lambda item: abs(item[0] - best_id[0]))
        pair = clean_imported_pair(names[0][1], best_id[1])
        if pair:
            pairs.append(pair)
    return dedupe_pairs(pairs)


def extract_name_id_lines_from_tables(tables: list[list[list[str]]], source_label: str) -> dict[str, Any]:
    pairs: list[tuple[str, str]] = []
    strategy = "fallback"
    match_label = ""
    for rows in tables:
        header_pair = find_header_pair(rows)
        if header_pair:
            extracted = []
            for row in rows[header_pair["header_row"] + 1 :]:
                pair = clean_imported_pair(
                    row[header_pair["name_index"]] if header_pair["name_index"] < len(row) else "",
                    row[header_pair["id_index"]] if header_pair["id_index"] < len(row) else "",
                )
                if pair:
                    extracted.append(pair)
            if extracted:
                pairs.extend(dedupe_pairs(extracted))
                if header_pair["exact"]:
                    strategy = "exact headers"
                    match_label = "Candidate Name + Colleague #"
                elif strategy != "exact headers":
                    strategy = "alternate headers"
                    match_label = f"{header_pair['name_header']} + {header_pair['id_header']}"
                continue
        pairs.extend(fallback_extract_pairs(rows))

    unique = dedupe_pairs(pairs)
    mode = f"{strategy} ({match_label})" if match_label else strategy
    return {
        "lines": [f"{name}, {emp_id}" for name, emp_id in unique],
        "found_rows": len(unique),
        "strategy": strategy,
        "message": f"Imported {len(unique)} row(s) from {source_label} using {mode}." if unique else f"No matching name and employee ID pairs could be pulled from {source_label}.",
    }


def parse_delimited_text(text: str) -> list[list[str]]:
    lines = [line for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if line.strip()]
    sample = "\n".join(lines[:10])
    delimiter = max(["\t", ",", ";"], key=lambda item: sample.count(item))
    if sample.count(delimiter) == 0:
        delimiter = ","
    return [[cell_text(cell) for cell in row] for row in csv.reader(lines, delimiter=delimiter)]


def clean_id_lines_from_text(text: str, numeric_only: bool = False) -> dict[str, Any]:
    normalized = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    raw_lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    structured_lines: list[str] = []
    structured = False

    def keep_id(hex_value: str) -> bool:
        clean = hex_value.strip().upper()
        return valid_hex8(clean) and (not numeric_only or clean.isdigit())

    try:
        rows = parse_delimited_text(normalized)
        if any(len(row) > 1 for row in rows):
            result = extract_name_id_lines_from_tables([rows], "clipboard")
            if result.get("found_rows"):
                structured_lines = [str(line) for line in result.get("lines", [])]
                structured = True
    except Exception:
        structured_lines = []

    clean_ids: list[str] = []
    for line in structured_lines or raw_lines:
        prepared = clean_candidate_line(line)
        hex_value = prepared["extracted"].strip().upper()
        if keep_id(hex_value):
            clean_ids.append(hex_value)
    if structured_lines:
        seen_clean = set(clean_ids)
        for line in raw_lines:
            prepared = clean_candidate_line(line)
            hex_value = prepared["extracted"].strip().upper()
            if keep_id(hex_value) and hex_value not in seen_clean:
                clean_ids.append(hex_value)
                seen_clean.add(hex_value)

    if not clean_ids:
        pattern = r"\b(?:\d{8}\.0+|\d{4}[\s-]+\d{4}|\d{8})\b" if numeric_only else r"\b(?:\d{8}\.0+|\d{4}[\s-]+\d{4}|[0-9A-Fa-f]{8})\b"
        for match in re.finditer(pattern, normalized):
            prepared = clean_candidate_line(match.group(0))
            hex_value = prepared["extracted"].strip().upper()
            if keep_id(hex_value):
                clean_ids.append(hex_value)

    raw_normalized = "\n".join(raw_lines)
    cleaned_normalized = "\n".join(clean_ids)
    table_like = structured or "\t" in normalized or any("," in line or ";" in line for line in raw_lines)
    changed = bool(clean_ids) and (table_like or cleaned_normalized != raw_normalized)
    return {
        "lines": clean_ids,
        "raw_line_count": len(raw_lines),
        "cleaned_count": len(clean_ids),
        "structured": structured,
        "changed": changed,
        "message": f"Found {len(clean_ids)} clean numeric ID row(s) from {len(raw_lines)} pasted line(s)." if numeric_only else f"Found {len(clean_ids)} clean ID row(s) from {len(raw_lines)} pasted line(s).",
    }


def extract_id_lines_from_tables(tables: list[list[list[str]]], source_label: str) -> dict[str, Any]:
    lines: list[str] = []
    for rows in tables:
        for row in rows:
            text = " ".join(cell_text(cell) for cell in row if cell_text(cell))
            if text:
                lines.append(text)
    cleaned = clean_id_lines_from_text("\n".join(lines), numeric_only=True)
    count = len(cleaned["lines"])
    return {
        "lines": cleaned["lines"],
        "found_rows": count,
        "strategy": "detected IDs",
        "message": f"Imported {count} detected numeric ID row(s) from {source_label}." if count else f"No clean numeric 8-character IDs were detected in {source_label}.",
    }


def import_result_queue_lines(result: dict[str, Any]) -> list[str]:
    lines = [str(line) for line in result.get("lines", [])]
    cleaned = clean_id_lines_from_text("\n".join(lines), numeric_only=True)
    return cleaned["lines"]


def clean_fc_cn_lines_from_text(text: str) -> dict[str, Any]:
    normalized = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    raw_lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    pairs: list[str] = []
    for line in raw_lines:
        try:
            facility, card = parse_fc_cn_line(line)
            fc_cn_to_hex(facility, card)
        except ValueError:
            continue
        pairs.append(f"{facility},{card}")
    raw_normalized = "\n".join(raw_lines)
    cleaned_normalized = "\n".join(pairs)
    return {
        "lines": pairs,
        "raw_line_count": len(raw_lines),
        "cleaned_count": len(pairs),
        "changed": bool(pairs) and cleaned_normalized != raw_normalized,
        "message": f"Found {len(pairs)} FC/CN pair(s) from {len(raw_lines)} line(s).",
    }


def extract_fc_cn_lines_from_tables(tables: list[list[list[str]]], source_label: str) -> dict[str, Any]:
    lines: list[str] = []
    for rows in tables:
        for row in rows:
            text = " ".join(cell_text(cell) for cell in row if cell_text(cell))
            if text:
                lines.append(text)
    cleaned = clean_fc_cn_lines_from_text("\n".join(lines))
    count = len(cleaned["lines"])
    return {
        "lines": cleaned["lines"],
        "found_rows": count,
        "strategy": "detected FC/CN pairs",
        "message": f"Imported {count} FC/CN pair(s) from {source_label}." if count else f"No FC/CN pairs were detected in {source_label}.",
    }


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._row is not None and self._cell is not None:
            self._row.append(cell_text("".join(self._cell)))
            self._cell = None
        elif tag == "tr" and self._table is not None and self._row is not None:
            if any(self._row):
                self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            if self._table:
                self.tables.append(self._table)
            self._table = None


def parse_html_tables(text: str) -> list[list[list[str]]]:
    parser = SimpleTableParser()
    parser.feed(text or "")
    return parser.tables


def local_name(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def parse_spreadsheet_xml(text: str) -> list[list[list[str]]]:
    root = ET.fromstring(text)
    rows = []
    for row_node in root.iter():
        if local_name(row_node.tag) != "Row":
            continue
        cells = []
        expected_index = 1
        for cell_node in list(row_node):
            if local_name(cell_node.tag) != "Cell":
                continue
            idx_attr = None
            for key, value in cell_node.attrib.items():
                if local_name(key) == "Index":
                    idx_attr = value
                    break
            if idx_attr:
                while expected_index < int(idx_attr):
                    cells.append("")
                    expected_index += 1
            data_text = "".join(cell_node.itertext())
            cells.append(cell_text(data_text))
            expected_index += 1
        if any(cells):
            rows.append(cells)
    return [rows] if rows else []


def parse_xlsx(path: Path) -> list[list[list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[list[list[str]]] = []
    try:
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = [cell_text(cell) for cell in row]
                if any(values):
                    rows.append(values)
            if rows:
                tables.append(rows)
    finally:
        workbook.close()
    return tables


def parse_xls(path: Path) -> list[list[list[str]]]:
    if xlrd is None:
        raise RuntimeError("XLS support requires xlrd.")
    book = xlrd.open_workbook(str(path))
    tables = []
    for sheet in book.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            values = [cell_text(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            if any(values):
                rows.append(values)
        if rows:
            tables.append(rows)
    return tables


def import_structured_file(path: Path) -> dict[str, Any]:
    lower = path.name.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        tables = parse_xlsx(path)
        result = extract_name_id_lines_from_tables(tables, path.name)
        return result if result["found_rows"] else extract_id_lines_from_tables(tables, path.name)
    if lower.endswith(".xls"):
        tables = parse_xls(path)
        result = extract_name_id_lines_from_tables(tables, path.name)
        return result if result["found_rows"] else extract_id_lines_from_tables(tables, path.name)

    text = path.read_text(encoding="utf-8-sig", errors="replace")
    if lower.endswith(".xml") or "<Workbook" in text:
        tables = parse_spreadsheet_xml(text)
        result = extract_name_id_lines_from_tables(tables, path.name)
        return result if result["found_rows"] else extract_id_lines_from_tables(tables, path.name)
    if lower.endswith((".html", ".htm")) or re.search(r"<table[\s>]", text, re.I):
        tables = parse_html_tables(text)
        result = extract_name_id_lines_from_tables(tables, path.name)
        return result if result["found_rows"] else extract_id_lines_from_tables(tables, path.name)
    if lower.endswith((".txt", ".csv", ".tsv")):
        cleaned = clean_id_lines_from_text(text, numeric_only=True)
        if cleaned["lines"]:
            return {
                "lines": cleaned["lines"],
                "found_rows": len(cleaned["lines"]),
                "strategy": "detected numeric IDs",
                "message": f"Imported {len(cleaned['lines'])} detected numeric ID row(s) from {path.name}.",
            }
        structured = extract_name_id_lines_from_tables([parse_delimited_text(text)], path.name)
        if structured["found_rows"]:
            return structured
        return {
            "lines": [],
            "found_rows": 0,
            "strategy": "detected IDs",
            "message": f"No clean 8-character IDs were detected in {path.name}.",
        }
    raise ValueError("Unsupported file type.")


def import_unconvert_file(path: Path) -> dict[str, Any]:
    lower = path.name.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return extract_fc_cn_lines_from_tables(parse_xlsx(path), path.name)
    if lower.endswith(".xls"):
        return extract_fc_cn_lines_from_tables(parse_xls(path), path.name)
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    if lower.endswith(".xml") or "<Workbook" in text:
        return extract_fc_cn_lines_from_tables(parse_spreadsheet_xml(text), path.name)
    if lower.endswith((".html", ".htm")) or re.search(r"<table[\s>]", text, re.I):
        return extract_fc_cn_lines_from_tables(parse_html_tables(text), path.name)
    if lower.endswith((".txt", ".csv", ".tsv")):
        cleaned = clean_fc_cn_lines_from_text(text)
        return {
            "lines": cleaned["lines"],
            "found_rows": len(cleaned["lines"]),
            "strategy": "detected FC/CN pairs",
            "message": f"Imported {len(cleaned['lines'])} FC/CN pair(s) from {path.name}." if cleaned["lines"] else f"No FC/CN pairs were detected in {path.name}.",
        }
    raise ValueError("Unsupported file type.")


class ConverterApp(TkRoot):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_DISPLAY_NAME)
        self.geometry("1200x780")
        self.minsize(980, 640)
        self.configure(bg=UI_BG)
        self.settings = self.load_settings()
        self.results: list[ConvertedRow] = []
        self.invalid: list[InvalidRow] = []
        self.unconvert_results: list[UnconvertRow] = []
        self.unconvert_invalid: list[InvalidRow] = []
        self.last_converted_at = ""
        self.last_unconverted_at = ""
        self.last_error_report = ""
        self.row_lookup: dict[str, ConvertedRow | InvalidRow] = {}
        self.unconvert_row_lookup: dict[str, UnconvertRow | InvalidRow] = {}
        self.sort_column = "Line"
        self.sort_reverse = False
        self.active_tab_index = 0
        self.status_target_tab = 0
        self.sidebar_focus_tab = 0
        self.batch_scan_count = 0
        self._batch_scan_after: str | None = None
        self._single_scan_after: str | None = None
        self._sidebar_tip_after: str | None = None
        self._last_sidebar_tip = ""
        self.last_single_auto_value = ""
        self.logo_photo: ImageTk.PhotoImage | None = None
        self.app_icon_photo: ImageTk.PhotoImage | None = None
        self.header_accent_photo: ImageTk.PhotoImage | None = None
        self.nav_icon_photos: list[ImageTk.PhotoImage] = []
        self.icon_photos: dict[tuple[str, int, int], ImageTk.PhotoImage] = {}
        self.empty_results_photo: ImageTk.PhotoImage | None = None
        self._setup_icon()
        self._setup_styles()
        self._build_shell()
        self._setup_shortcuts()
        self._rotate_sidebar_tip()
        self.render_results()
        self.render_unconvert_results()
        self.after(50, self._start_full_view)

    def _start_full_view(self) -> None:
        try:
            self.state("zoomed")
        except tk.TclError:
            width = self.winfo_screenwidth()
            height = self.winfo_screenheight()
            self.geometry(f"{width}x{height}+0+0")

    def load_settings(self) -> dict[str, Any]:
        defaults = {
            "default_export_dir": str(Path.home() / "Desktop"),
            "default_export_type": EXPORT_TYPE_CHOICES[0],
            "recent_files": [],
            "recent_exports": [],
            "history": [],
        }
        try:
            if SETTINGS_FILE.exists():
                data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
                defaults.update({key: data.get(key, value) for key, value in defaults.items()})
        except Exception:
            pass
        if defaults["default_export_type"] not in EXPORT_TYPE_CHOICES:
            defaults["default_export_type"] = EXPORT_TYPE_CHOICES[0]
        return defaults

    def save_settings(self) -> None:
        APP_STATE_DIR.mkdir(parents=True, exist_ok=True)
        SETTINGS_FILE.write_text(json.dumps(self.settings, indent=2), encoding="utf-8")

    def add_recent_file(self, path: Path) -> None:
        recent = [item for item in self.settings.get("recent_files", []) if item != str(path)]
        recent.insert(0, str(path))
        self.settings["recent_files"] = recent[:8]
        self.save_settings()
        if hasattr(self, "recent_menu"):
            self.refresh_recent_menu()

    def add_recent_export(self, path: Path, export_name: str) -> None:
        item = {
            "path": str(path),
            "name": path.name,
            "type": export_name,
            "timestamp": datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
        }
        recent = [entry for entry in self.settings.get("recent_exports", []) if entry.get("path") != str(path)]
        recent.insert(0, item)
        self.settings["recent_exports"] = recent[:10]
        self.save_settings()

    def clear_recent_exports(self) -> None:
        count = len(self.settings.get("recent_exports", []))
        if count == 0:
            self.set_status("No recent exports saved.")
            return
        if not messagebox.askyesno("Clear recent exports", "Remove saved recent export shortcuts from this workstation?"):
            return
        self.settings["recent_exports"] = []
        self.save_settings()
        self.set_status(f"Cleared {count} recent export shortcut(s).")

    def build_error_report(self, title: str, details: list[str]) -> str:
        lines = [
            f"{APP_SHORT_NAME} Error Report",
            f"Version: {APP_VERSION}",
            f"Time: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}",
            f"Issue: {title}",
            "",
            "Details:",
        ]
        lines.extend(details or ["No additional details were captured."])
        return "\n".join(lines)

    def record_error_report(self, title: str, details: list[str]) -> None:
        self.last_error_report = self.build_error_report(title, details)
        self.set_status(f"{title}. Error report ready to copy.", "Needs Review")

    def copy_error_report(self) -> None:
        if not self.last_error_report:
            messagebox.showinfo("No error report", "No import or export error has been captured in this session.")
            return
        self.clipboard_clear()
        self.clipboard_append(self.last_error_report)
        self.set_status("Error report copied to clipboard.", "Needs Review")

    def add_history(self, action: str, total: int, valid: int, invalid: int, warnings: int) -> None:
        entry = HistoryEntry(datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"), action, total, valid, invalid, warnings)
        history = self.settings.get("history", [])
        history.insert(0, entry.__dict__)
        self.settings["history"] = history[:50]
        self.save_settings()
        if hasattr(self, "history_tree") or hasattr(self, "history_text"):
            self.render_history()

    def _setup_icon(self) -> None:
        image_icon = asset_path("macys-ap-icon.png")
        if image_icon.exists():
            try:
                self.app_icon_photo = ImageTk.PhotoImage(Image.open(image_icon).resize((64, 64), Image.LANCZOS))
                self.iconphoto(True, self.app_icon_photo)
            except (tk.TclError, OSError):
                self.app_icon_photo = None
        icon = asset_path("app-icon.ico")
        if icon.exists():
            try:
                self.iconbitmap(str(icon))
            except tk.TclError:
                pass

    def _apply_window_icon(self, window: tk.Toplevel | tk.Tk) -> None:
        if self.app_icon_photo:
            try:
                window.iconphoto(False, self.app_icon_photo)
            except tk.TclError:
                pass
        icon = asset_path("app-icon.ico")
        if icon.exists():
            try:
                window.iconbitmap(str(icon))
            except tk.TclError:
                pass

    def _new_dialog(self, title: str, bg: str = UI_BG) -> tk.Toplevel:
        dialog = tk.Toplevel(self)
        dialog.title(title)
        dialog.configure(bg=bg)
        self._apply_window_icon(dialog)
        dialog.transient(self)
        dialog.grab_set()
        return dialog

    def _setup_styles(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", background=UI_BG, foreground=UI_TEXT, fieldbackground=UI_INPUT)
        style.configure("TNotebook", background=UI_BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(16, 9), background=UI_SURFACE, foreground=UI_MUTED)
        style.map("TNotebook.Tab", background=[("selected", UI_RED_SOFT)], foreground=[("selected", UI_RED)])
        style.configure(
            "Treeview",
            background=UI_SURFACE,
            foreground=UI_TEXT,
            fieldbackground=UI_SURFACE,
            font=("Segoe UI", 10),
            rowheight=34,
            borderwidth=0,
            relief="flat",
            bordercolor=UI_BORDER,
            lightcolor=UI_BORDER,
            darkcolor=UI_BORDER,
        )
        style.configure(
            "Treeview.Heading",
            background="#eef2f6",
            foreground=UI_TEXT,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            borderwidth=1,
            bordercolor=UI_BORDER,
        )
        style.map("Treeview", background=[("selected", UI_RED)], foreground=[("selected", "#ffffff")])
        style.configure("TCombobox", padding=6)
        for scrollbar_style in ("Dark.Vertical.TScrollbar", "Dark.Horizontal.TScrollbar"):
            style.configure(
                scrollbar_style,
                gripcount=0,
                background=UI_BORDER_DARK,
                darkcolor=UI_BORDER,
                lightcolor=UI_SURFACE,
                troughcolor="#eef2f6",
                bordercolor="#eef2f6",
                arrowcolor=UI_TEXT,
                relief="flat",
                width=14,
            )
            style.map(
                scrollbar_style,
                background=[("active", UI_MUTED), ("pressed", UI_RED)],
                arrowcolor=[("active", "#ffffff"), ("pressed", "#ffffff")],
            )

    def _setup_shortcuts(self) -> None:
        self.bind_all("<Control-i>", lambda _event: self.import_file())
        self.bind_all("<Control-r>", lambda _event: self.convert_batch())
        self.bind_all("<Control-e>", lambda _event: self.export_excel())
        self.bind_all("<Control-Shift-E>", lambda _event: self.export_csv())
        self.bind_all("<Control-p>", lambda _event: self.export_pdf())
        self.bind_all("<Control-l>", lambda _event: self.clear_workspace())
        self.bind_all("<Control-f>", lambda _event: self.focus_search())
        self.bind_all("<F9>", self.focus_batch_scanner)
        self.bind_all("<F10>", self.focus_single_lookup)

    def focus_search(self) -> None:
        self.select_tab(0)
        if hasattr(self, "search_entry"):
            self.search_var.set("")
            self.search_entry.focus_set()
        self.set_status("Search is ready.")

    def focus_batch_scanner(self, _event: Any | None = None) -> str:
        self.select_tab(0)

        def focus() -> None:
            if hasattr(self, "batch_scan_entry"):
                self.batch_scan_entry.focus_set()
                self.batch_scan_entry.selection_range(0, "end")
                self.set_status("Batch Scanner Input is ready. Scan an ID or press Enter after typing.", "Ready", target_tab=0)

        self.after(80, focus)
        return "break"

    def focus_single_lookup(self, _event: Any | None = None) -> str:
        self.select_tab(1)

        def focus() -> None:
            if hasattr(self, "single_entry"):
                self.single_entry.focus_set()
                self.single_entry.selection_range(0, "end")
                self.set_status("Single Lookup scanner field is ready. Scan one HEX ID.", "Ready", target_tab=1)

        self.after(80, focus)
        return "break"

    def _menu_options(self) -> dict[str, Any]:
        return {
            "bg": UI_SURFACE,
            "fg": UI_TEXT,
            "activebackground": UI_RED,
            "activeforeground": "#ffffff",
            "disabledforeground": UI_MUTED,
            "bd": 1,
            "relief": "solid",
            "activeborderwidth": 0,
            "font": ("Segoe UI", 10),
        }

    def _styled_menu(self, parent: tk.Widget) -> tk.Menu:
        return tk.Menu(parent, tearoff=False, **self._menu_options())

    def _style_menu(self, menu: tk.Menu) -> None:
        try:
            menu.configure(**self._menu_options())
        except tk.TclError:
            pass

    def _build_menu(self) -> None:
        menubar = tk.Menu(self)
        import_menu = self._styled_menu(menubar)
        import_menu.add_command(label="Browse Files", command=self.import_file)
        import_menu.add_command(label="Paste Clipboard To Queue", command=self.paste_clipboard_to_queue)
        import_menu.add_command(label="Load Sample IDs", command=self.load_sample)
        import_menu.add_separator()
        import_menu.add_command(label="Focus Batch Scanner (F9)", command=self.focus_batch_scanner)
        import_menu.add_command(label="Focus Single Lookup (F10)", command=self.focus_single_lookup)
        menubar.add_cascade(label="Import", menu=import_menu)

        file_menu = self._styled_menu(menubar)
        file_menu.add_command(label="Settings", command=self.show_settings)
        file_menu.add_command(label="Choose Default Export Folder", command=self.choose_export_folder)
        file_menu.add_command(label="Open Export Folder", command=self.open_export_folder)
        file_menu.add_command(label="Create Desktop Shortcut", command=self.create_desktop_shortcut)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.destroy)
        menubar.add_cascade(label="File", menu=file_menu)

        export_menu = self._styled_menu(menubar)
        export_menu.add_command(label="Export Default", command=self.export_default)
        export_menu.add_separator()
        export_menu.add_command(label="Excel Workbook", command=self.export_excel)
        export_menu.add_command(label="CSV Report", command=self.export_csv)
        export_menu.add_command(label="TXT Report", command=self.export_txt)
        export_menu.add_command(label="PDF Report", command=self.export_pdf)
        export_menu.add_separator()
        export_menu.add_command(label="Recent Exports", command=self.show_recent_exports)
        menubar.add_cascade(label="Export", menu=export_menu)

        help_menu = self._styled_menu(menubar)
        help_menu.add_command(label="How To Use", command=self.show_help)
        help_menu.add_command(label="About", command=self.show_about)
        help_menu.add_command(label="Copy Last Error Report", command=self.copy_error_report)
        menubar.add_cascade(label="Help", menu=help_menu)
        self.config(menu=menubar)

    def refresh_recent_menu(self) -> None:
        self.recent_menu.delete(0, "end")
        recent = [item for item in self.settings.get("recent_files", []) if Path(item).exists()]
        if not recent:
            self.recent_menu.add_command(label="No recent files", state="disabled")
            return
        for item in recent:
            self.recent_menu.add_command(label=Path(item).name, command=lambda p=item: self.import_file(Path(p)))

    def _load_icon(self, name: str, size: int = 18) -> ImageTk.PhotoImage | None:
        key = (name, size, size)
        if key in self.icon_photos:
            return self.icon_photos[key]
        path = asset_path(name)
        if not path.exists():
            return None
        image = Image.open(path).resize((size, size), Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        self.icon_photos[key] = photo
        return photo

    def _button(
        self,
        parent: tk.Widget,
        text: str,
        command: Any,
        primary: bool = False,
        icon: str | None = None,
        tooltip: str | None = None,
        compact: bool = False,
    ) -> tk.Button:
        bg = UI_RED if primary else UI_SURFACE_ALT
        hover = UI_RED_DARK if primary else "#eef2f6"
        fg = "#ffffff" if primary else UI_TEXT
        icon_photo = self._load_icon(icon, 16 if compact else 18) if icon else None
        button = tk.Button(
            parent,
            text=text,
            image=icon_photo,
            compound="left" if icon_photo else "none",
            command=command,
            bg=bg,
            fg=fg,
            activebackground=hover,
            activeforeground="#ffffff" if primary else UI_TEXT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=UI_BORDER if not primary else UI_RED_DARK,
            font=("Segoe UI", 8 if compact else 9, "bold"),
            padx=8 if compact else 13,
            pady=7 if compact else 8,
            cursor="hand2",
        )
        button.bind("<Enter>", lambda _event: button.configure(bg=hover))
        button.bind("<Leave>", lambda _event: button.configure(bg=bg))
        if tooltip:
            ToolTip(button, tooltip)
        return button

    def _menu_button(
        self,
        parent: tk.Widget,
        text: str,
        items: list[tuple[str, Any] | None],
        icon: str | None = None,
        tooltip: str | None = None,
    ) -> tk.Menubutton:
        icon_photo = self._load_icon(icon, 18) if icon else None
        bg = UI_SURFACE_ALT
        hover = "#eef2f6"
        button = tk.Menubutton(
            parent,
            text=f"{text} v",
            image=icon_photo,
            compound="left" if icon_photo else "none",
            bg=bg,
            fg=UI_TEXT,
            activebackground=hover,
            activeforeground=UI_TEXT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=8,
            cursor="hand2",
        )
        menu = self._styled_menu(button)
        for item in items:
            if item is None:
                menu.add_separator()
            else:
                label, command = item
                menu.add_command(label=label, command=command)
        button.configure(menu=menu)
        button.bind("<Enter>", lambda _event: button.configure(bg=hover))
        button.bind("<Leave>", lambda _event: button.configure(bg=bg))
        if tooltip:
            ToolTip(button, tooltip)
        return button

    def _panel(self, parent: tk.Widget) -> tk.Frame:
        return tk.Frame(parent, bg=UI_SURFACE, highlightthickness=1, highlightbackground=UI_BORDER)

    def _card(self, parent: tk.Widget, bg: str = UI_SURFACE_ALT, border: str = UI_BORDER) -> tk.Frame:
        return tk.Frame(parent, bg=bg, highlightthickness=1, highlightbackground=border)

    def _info_strip(self, parent: tk.Widget, bg: str) -> tk.Frame:
        return tk.Frame(parent, bg=bg, highlightthickness=1, highlightbackground=UI_CARD_SHADOW)

    def _dialog_header(
        self,
        dialog: tk.Toplevel,
        title: str,
        subtitle: str,
        accent: str = UI_RED,
        right_builder: Any | None = None,
    ) -> tk.Frame:
        header = tk.Frame(dialog, bg=UI_HEADER, highlightthickness=1, highlightbackground=UI_BORDER)
        header.pack(fill="x")
        tk.Frame(header, bg=accent, width=5).pack(side="left", fill="y")
        logo = self._load_icon("macys-ap-icon.png", 48)
        if logo:
            tk.Label(header, image=logo, bg=UI_HEADER).pack(side="left", padx=(18, 12), pady=14)
        title_box = tk.Frame(header, bg=UI_HEADER)
        title_box.pack(side="left", fill="x", expand=True, pady=14)
        tk.Label(title_box, text=title, bg=UI_HEADER, fg=UI_TEXT, font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(title_box, text=subtitle, bg=UI_HEADER, fg=UI_MUTED, font=("Segoe UI", 9)).pack(anchor="w", pady=(2, 0))
        if right_builder:
            right_builder(header)
        return header

    def _apply_corporate_skin(self, root: tk.Widget) -> None:
        bg_map = {
            "#0b0d12": UI_BG,
            "#10141b": UI_SURFACE,
            "#111721": UI_SURFACE,
            "#151922": UI_SURFACE_ALT,
            "#080a0f": UI_INPUT,
            "#090b10": UI_HEADER,
            "#1f2632": UI_BORDER,
            "#202633": UI_SURFACE_ALT,
            "#2c3444": "#eef2f6",
            "#251017": UI_RED_SOFT,
            "#0d1722": "#eef6ff",
            "#2a2119": UI_BAD_SOFT,
            "#1d1a12": UI_WARN_SOFT,
            "#263141": UI_BORDER,
        }
        fg_map = {
            "#ffffff": UI_TEXT,
            "#f5f7fb": UI_TEXT,
            "#d9e3f0": UI_TEXT,
            "#a8b2c2": UI_MUTED,
            "#6f7a8b": UI_MUTED,
            "#46d9ff": UI_BLUE,
            "#8beaff": UI_BLUE,
            "#ffca63": UI_WARN_TEXT,
            "#ffd9a0": UI_BAD_TEXT,
            "#ffe1a3": UI_WARN_TEXT,
            "#ff6d78": UI_RED,
        }
        border_map = {
            "#303845": UI_BORDER,
            "#323b49": UI_BORDER,
            "#344153": UI_BORDER,
            "#3c4656": UI_BORDER,
            "#4a5261": UI_BORDER,
            "#202734": UI_BORDER,
            "#252d3a": UI_BORDER,
            "#2d3848": UI_BORDER,
        }

        def normalize(value: Any) -> str:
            return str(value).lower()

        def recolor(widget: tk.Widget) -> None:
            widget_class = widget.winfo_class()
            try:
                current_bg = normalize(widget.cget("background"))
            except tk.TclError:
                current_bg = ""
            new_bg = bg_map.get(current_bg, None)
            if new_bg:
                try:
                    widget.configure(bg=new_bg)
                except tk.TclError:
                    pass
                current_bg = normalize(new_bg)

            for option in ("highlightbackground", "highlightcolor"):
                try:
                    value = normalize(widget.cget(option))
                    if value in border_map:
                        widget.configure(**{option: border_map[value]})
                except tk.TclError:
                    pass

            try:
                current_fg = normalize(widget.cget("foreground"))
                if current_bg in {normalize(UI_RED), normalize(UI_RED_DARK)}:
                    widget.configure(fg="#ffffff")
                    try:
                        widget.configure(activeforeground="#ffffff")
                    except tk.TclError:
                        pass
                elif current_fg in fg_map:
                    widget.configure(fg=fg_map[current_fg])
            except tk.TclError:
                pass

            if widget_class in {"Text", "Entry"}:
                try:
                    widget.configure(bg=UI_INPUT, fg=UI_TEXT, insertbackground=UI_RED)
                except tk.TclError:
                    pass
            if widget_class == "Menu":
                try:
                    widget.configure(bg=UI_SURFACE, fg=UI_TEXT, activebackground=UI_RED, activeforeground="#ffffff")
                except tk.TclError:
                    pass
            for child in widget.winfo_children():
                recolor(child)

        recolor(root)

    def _enable_mousewheel(self, widget: tk.Widget, y_target: Any | None = None, x_target: Any | None = None) -> None:
        y_target = y_target or widget
        x_target = x_target or y_target

        def units_from_event(event: Any) -> int:
            if getattr(event, "num", None) == 4:
                return -3
            if getattr(event, "num", None) == 5:
                return 3
            delta = getattr(event, "delta", 0)
            if delta == 0:
                return 0
            if abs(delta) < 120:
                return -1 if delta > 0 else 1
            return -int(delta / 120)

        def scroll_y(event: Any) -> str:
            units = units_from_event(event)
            if units and hasattr(y_target, "yview_scroll"):
                y_target.yview_scroll(units, "units")
            return "break"

        def scroll_x(event: Any) -> str:
            units = units_from_event(event)
            if units and hasattr(x_target, "xview_scroll"):
                x_target.xview_scroll(units, "units")
            return "break"

        widget.bind("<MouseWheel>", scroll_y, add="+")
        widget.bind("<Shift-MouseWheel>", scroll_x, add="+")
        widget.bind("<Button-4>", scroll_y, add="+")
        widget.bind("<Button-5>", scroll_y, add="+")

    def _enable_mousewheel_tree(self, widget: tk.Widget, y_target: Any, x_target: Any | None = None) -> None:
        self._enable_mousewheel(widget, y_target, x_target)
        for child in widget.winfo_children():
            self._enable_mousewheel_tree(child, y_target, x_target)

    def _enable_tree_header_tooltips(self, tree: ttk.Treeview, tips: dict[str, str]) -> None:
        tooltip = ToolTip(tree, "")

        def update_tip(event: Any) -> None:
            if tree.identify_region(event.x, event.y) != "heading":
                tooltip.hide()
                return
            column_id = tree.identify_column(event.x)
            try:
                column = tree["columns"][int(column_id.replace("#", "")) - 1]
            except (ValueError, IndexError, tk.TclError):
                tooltip.hide()
                return
            tooltip.text = tips.get(str(column), "")
            if tooltip.text:
                tooltip.schedule()
            else:
                tooltip.hide()

        tree.bind("<Motion>", update_tip, add="+")
        tree.bind("<Leave>", tooltip.hide, add="+")

    def _enable_tree_row_hover(self, tree: ttk.Treeview) -> None:
        tree.tag_configure("hover", background="#eef6ff", foreground=UI_TEXT)
        state: dict[str, Any] = {"item": "", "tags": {}}

        def restore(item_id: str) -> None:
            tags = state["tags"].pop(item_id, None)
            if tags is not None and tree.exists(item_id):
                tree.item(item_id, tags=tags)

        def on_motion(event: Any) -> None:
            item_id = tree.identify_row(event.y)
            previous = state.get("item", "")
            if item_id == previous:
                return
            if previous:
                restore(previous)
            state["item"] = item_id
            if item_id and tree.exists(item_id):
                current_tags = tuple(tree.item(item_id, "tags"))
                state["tags"][item_id] = current_tags
                tree.item(item_id, tags=(*current_tags, "hover"))

        def on_leave(_event: Any) -> None:
            previous = state.get("item", "")
            if previous:
                restore(previous)
            state["item"] = ""

        tree.bind("<Motion>", on_motion, add="+")
        tree.bind("<Leave>", on_leave, add="+")

    def _context_menu(self, items: list[tuple[str, Any] | None]) -> tk.Menu:
        menu = self._styled_menu(self)
        for item in items:
            if item is None:
                menu.add_separator()
            else:
                label, command = item
                menu.add_command(label=label, command=command)
        return menu

    def _mark_edit_widget_changed(self, widget: tk.Widget) -> None:
        if hasattr(self, "multi_text") and widget is self.multi_text:
            self.multi_text.edit_modified(True)
            self.handle_batch_input_changed()
        elif hasattr(self, "unconvert_text") and widget is self.unconvert_text:
            self.unconvert_text.edit_modified(True)
            self.handle_unconvert_input_changed()

    def _paste_raw_into_widget(self, widget: tk.Widget) -> None:
        try:
            widget.event_generate("<<Paste>>")
            self.after(25, lambda: self._mark_edit_widget_changed(widget))
        except tk.TclError:
            pass

    def _cut_from_widget(self, widget: tk.Widget) -> None:
        try:
            widget.event_generate("<<Cut>>")
            self.after(25, lambda: self._mark_edit_widget_changed(widget))
        except tk.TclError:
            pass

    def _copy_from_widget(self, widget: tk.Widget) -> None:
        try:
            widget.event_generate("<<Copy>>")
        except tk.TclError:
            pass

    def _select_all_in_widget(self, widget: tk.Widget) -> None:
        try:
            if widget.winfo_class() == "Text":
                widget.tag_add("sel", "1.0", "end-1c")  # type: ignore[attr-defined]
                widget.mark_set("insert", "1.0")  # type: ignore[attr-defined]
            else:
                widget.selection_range(0, "end")  # type: ignore[attr-defined]
                widget.icursor("end")  # type: ignore[attr-defined]
            widget.focus_set()
        except tk.TclError:
            pass

    def _clear_edit_widget(self, widget: tk.Widget) -> None:
        try:
            if widget.winfo_class() == "Text":
                widget.delete("1.0", "end")  # type: ignore[attr-defined]
            else:
                widget.delete(0, "end")  # type: ignore[attr-defined]
            self._mark_edit_widget_changed(widget)
        except tk.TclError:
            pass

    def _enable_edit_context_menu(
        self,
        widget: tk.Widget,
        clean_paste_command: Any | None = None,
        clean_label: str = "Paste Clean Values",
    ) -> None:
        items: list[tuple[str, Any] | None] = []
        if clean_paste_command:
            items.extend(
                [
                    (clean_label, clean_paste_command),
                    ("Paste Raw Text", lambda: self._paste_raw_into_widget(widget)),
                ]
            )
        else:
            items.append(("Paste", lambda: self._paste_raw_into_widget(widget)))
        items.extend(
            [
                None,
                ("Cut", lambda: self._cut_from_widget(widget)),
                ("Copy", lambda: self._copy_from_widget(widget)),
                ("Select All", lambda: self._select_all_in_widget(widget)),
                None,
                ("Clear", lambda: self._clear_edit_widget(widget)),
            ]
        )
        menu = self._context_menu(items)

        def show_menu(event: Any) -> str:
            widget.focus_set()
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
            return "break"

        widget.bind("<Button-3>", show_menu, add="+")

    def _add_text_resize_handle(
        self,
        parent: tk.Widget,
        text_widget: tk.Text,
        label: str,
        min_height: int = 3,
        max_height: int = 18,
    ) -> None:
        handle = tk.Frame(parent, bg="#18202c", height=20, cursor="sb_v_double_arrow")
        handle.pack(fill="x", padx=12, pady=(4, 0))
        handle.pack_propagate(False)
        tk.Label(
            handle,
            text=label,
            bg="#18202c",
            fg="#a8b2c2",
            font=("Segoe UI", 8, "bold"),
            cursor="sb_v_double_arrow",
        ).pack(expand=True)
        state: dict[str, int] = {"start_y": 0, "start_height": int(text_widget.cget("height"))}

        def start_drag(event: Any) -> None:
            state["start_y"] = int(event.y_root)
            state["start_height"] = int(text_widget.cget("height"))

        def drag(event: Any) -> None:
            delta = int((int(event.y_root) - state["start_y"]) / 18)
            height = max(min_height, min(max_height, state["start_height"] + delta))
            text_widget.configure(height=height)

        handle.bind("<ButtonPress-1>", start_drag)
        handle.bind("<B1-Motion>", drag)
        for child in handle.winfo_children():
            child.bind("<ButtonPress-1>", start_drag)
            child.bind("<B1-Motion>", drag)

    def _show_tree_context_menu(self, event: Any, tree: ttk.Treeview, menu: tk.Menu) -> str:
        item_id = tree.identify_row(event.y)
        if not item_id:
            return "break"
        tree.selection_set(item_id)
        tree.focus(item_id)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def _dialog_footer_accent(self, dialog: tk.Widget, accent: str = UI_RED, padx: int = 18, pady: tuple[int, int] = (0, 8)) -> None:
        tk.Frame(dialog, bg=accent, height=3).pack(fill="x", padx=padx, pady=pady)

    def _link_label(
        self,
        parent: tk.Widget,
        text: str,
        url: str,
        bg: str,
        tooltip: str | None = None,
        font_size: int = 9,
        bold: bool = False,
        padx: int = 10,
    ) -> tk.Label:
        label = tk.Label(
            parent,
            text=text,
            bg=bg,
            fg=UI_BLUE,
            cursor="hand2",
            padx=padx,
            font=("Segoe UI", font_size, "bold underline" if bold else "underline"),
        )
        label.bind("<Button-1>", lambda _event: webbrowser.open(url))
        label.bind("<Enter>", lambda _event: label.configure(fg=UI_RED))
        label.bind("<Leave>", lambda _event: label.configure(fg=UI_BLUE))
        if tooltip:
            ToolTip(label, tooltip)
        return label

    def _copy_value_from_var(self, variable: tk.StringVar, label: str, target_tab: int) -> bool:
        value = variable.get().strip()
        if not value or value == "--":
            self.set_status(f"No {label} value to copy yet.", "Needs Review", target_tab=target_tab)
            return False
        self.clipboard_clear()
        self.clipboard_append(value)
        self.set_status(f"Copied {label}.", "Ready", target_tab=target_tab)
        return True

    def _flash_label(self, label: tk.Label, flash_bg: str = UI_GREEN_SOFT, duration: int = 180) -> None:
        try:
            original_bg = str(label.cget("background"))
            label.configure(bg=flash_bg)
            label.after(duration, lambda: label.configure(bg=original_bg))
        except tk.TclError:
            return

    def _clickable_value_label(
        self,
        parent: tk.Widget,
        variable: tk.StringVar,
        accent: str,
        label: str,
        target_tab: int,
        font_size: int = 13,
    ) -> tk.Label:
        hover = UI_RED if accent != UI_RED else UI_BLUE
        value_label = tk.Label(
            parent,
            textvariable=variable,
            bg=UI_SURFACE_ALT,
            fg=accent,
            font=("Cascadia Mono", font_size, "bold"),
            cursor="hand2",
        )
        normal_font = ("Cascadia Mono", font_size, "bold")
        hover_font = ("Cascadia Mono", font_size, "bold underline")

        def copy_clicked(_event: Any) -> str:
            if self._copy_value_from_var(variable, label, target_tab):
                self._flash_label(value_label)
            return "break"

        value_label.bind("<Button-1>", copy_clicked)
        value_label.bind("<Enter>", lambda _event: value_label.configure(fg=hover, font=hover_font))
        value_label.bind("<Leave>", lambda _event: value_label.configure(fg=accent, font=normal_font))
        ToolTip(value_label, f"Click the {label} number to copy it.")
        return value_label

    def _sidebar_tip_options(self) -> list[tuple[str, str]]:
        return [
            ("F9 Batch Scan", "Focus scanner input. New scans go to the top."),
            ("F10 Single Lookup", "Jump to one-ID lookup and auto-copy FC/CN."),
            ("Ctrl+I Import", "Browse for Excel, CSV, TXT, XML, or HTML data files."),
            ("Ctrl+R Convert", "Run the active batch after checking row colors."),
            ("Ctrl+F Search", "Move to Results search after a long run."),
            ("Click Results", "Click exact output values in lookup cards to copy them."),
            ("Right-Click Rows", "Open copy options for results or unconvert rows."),
            ("Drag Files", "Drop supported files onto a batch input area to import them."),
            ("Keep Valid", "Clean the queue down to rows the app can convert."),
            ("Export Default", "Use the report type saved in Settings."),
        ]

    def _rotate_sidebar_tip(self) -> None:
        if not hasattr(self, "sidebar_tip_title_var"):
            return
        tips = self._sidebar_tip_options()
        choices = [tip for tip in tips if "|".join(tip) != self._last_sidebar_tip] or tips
        title, body = random.choice(choices)
        self._last_sidebar_tip = f"{title}|{body}"
        self.sidebar_tip_title_var.set(title)
        self.sidebar_tip_body_var.set(body)
        accent, _soft = self._tab_theme(getattr(self, "sidebar_focus_tab", getattr(self, "active_tab_index", 0)))
        if hasattr(self, "sidebar_tip_bar"):
            self.sidebar_tip_bar.configure(bg=accent)
        if hasattr(self, "sidebar_tip_title_label"):
            self.sidebar_tip_title_label.configure(fg=accent)
        if hasattr(self, "sidebar_tip_body_label"):
            self.sidebar_tip_body_label.configure(fg=UI_TEXT)
        self._sidebar_tip_after = self.after(15000, self._rotate_sidebar_tip)

    def _tab_theme(self, index: int, state: str = "Ready") -> tuple[str, str]:
        if state == "Needs Review":
            return UI_WARN_TEXT, UI_WARN_SOFT
        if state == "Exported":
            return UI_BLUE, UI_BLUE_SOFT
        themes = {
            0: (UI_RED, UI_RED_SOFT),
            1: (UI_BLUE, UI_BLUE_SOFT),
            2: (UI_WARN_TEXT, UI_WARN_SOFT),
            3: (UI_WARN_TEXT, UI_WARN_SOFT),
            4: (UI_GREEN_TEXT, UI_GREEN_SOFT),
        }
        return themes.get(max(0, min(4, index)), (UI_RED, UI_RED_SOFT))

    def _section_theme(self, section: str) -> tuple[str, str]:
        normalized = section.strip().lower()
        if normalized == "reverse":
            return UI_WARN_TEXT, UI_WARN_SOFT
        if normalized == "review":
            return UI_GREEN_TEXT, UI_GREEN_SOFT
        return UI_RED, UI_RED_SOFT

    def _apply_widget_bg(self, widgets: list[tk.Widget], color: str) -> None:
        for widget in widgets:
            try:
                widget.configure(bg=color)
            except tk.TclError:
                continue

    def _workspace_focus_text(self, index: int, state: str = "Ready") -> str:
        labels = ["Batch Converter", "Single Lookup", "FC/CN to Hex", "Unconvert Batch", "History"]
        workspace = labels[max(0, min(4, index))]
        if state == "Needs Review":
            return f"Needs Review\nOpen {workspace}."
        if state == "Exported":
            return "Export Ready\nUse Recent Exports."
        focus = {
            0: "Batch Converter\nScan, convert, export.",
            1: "Single Lookup\nClick values to copy.",
            2: "FC/CN to Hex\nClick HEX to copy.",
            3: "Unconvert Batch\nBuild HEX from FC/CN.",
            4: "History\nReview local runs.",
        }
        return focus.get(index, workspace)

    def update_sidebar_focus(self, index: int, state: str = "Ready") -> None:
        self.sidebar_focus_tab = max(0, min(4, index))
        if hasattr(self, "nav_status"):
            self.nav_status.set(self._workspace_focus_text(self.sidebar_focus_tab, state))
        accent, soft = self._tab_theme(self.sidebar_focus_tab, state)
        if hasattr(self, "work_focus_bar"):
            self.work_focus_bar.configure(bg=accent)
        if hasattr(self, "nav_status_label"):
            self.nav_status_label.configure(fg=accent)
        if hasattr(self, "focus_card_widgets"):
            self._apply_widget_bg(self.focus_card_widgets, soft)
        if hasattr(self, "guide_card_widgets"):
            self._apply_widget_bg(self.guide_card_widgets, soft)
        if hasattr(self, "sidebar_tip_bar"):
            self.sidebar_tip_bar.configure(bg=accent)
        if hasattr(self, "sidebar_tip_title_label"):
            self.sidebar_tip_title_label.configure(fg=accent)
        if hasattr(self, "sidebar_tip_body_label"):
            self.sidebar_tip_body_label.configure(fg=UI_TEXT)

    def open_sidebar_focus_target(self, _event: Any | None = None) -> str:
        target = max(0, min(4, int(getattr(self, "sidebar_focus_tab", getattr(self, "active_tab_index", 0)))))
        self.select_tab(target)
        return "break"

    def _nav_section(self, parent: tk.Widget, text: str) -> None:
        accent, soft = self._section_theme(text)
        section = tk.Frame(parent, bg=soft, highlightthickness=1, highlightbackground=UI_BORDER)
        section.pack(fill="x", padx=10, pady=(14, 6))
        tk.Frame(section, bg=accent, width=4).pack(side="left", fill="y")
        tk.Label(
            section,
            text=text.upper(),
            bg=soft,
            fg=accent,
            font=("Segoe UI", 8, "bold"),
            anchor="w",
            padx=9,
            pady=5,
        ).pack(fill="x")

    def _nav_button(
        self,
        parent: tk.Widget,
        text: str,
        index: int,
        image: ImageTk.PhotoImage | None = None,
        tooltip: str | None = None,
        accent: str = UI_RED,
        active_bg: str = UI_RED_SOFT,
    ) -> tk.Button:
        button = tk.Button(
            parent,
            text=text,
            image=image,
            compound="left" if image else "none",
            command=lambda: self.select_tab(index),
            anchor="w",
            bg=UI_SURFACE,
            fg=UI_TEXT,
            activebackground=active_bg,
            activeforeground=accent,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
            font=("Segoe UI", 10, "bold"),
            padx=14,
            pady=11,
            cursor="hand2",
        )
        button.bind("<Enter>", lambda _event: button.configure(bg=active_bg))
        button.bind("<Leave>", lambda _event: button.configure(bg=active_bg if getattr(self, "active_tab_index", -1) == index else UI_SURFACE))
        if tooltip:
            ToolTip(button, tooltip)
        return button

    def _load_nav_icon(self, name: str) -> ImageTk.PhotoImage | None:
        path = asset_path(name)
        if not path.exists():
            return None
        image = Image.open(path).resize((24, 24), Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        self.nav_icon_photos.append(photo)
        return photo

    def _workspace_title(
        self,
        parent: tk.Widget,
        title: str,
        subtitle: str,
        chips: list[tuple[str, str]] | None = None,
        accent: str = UI_RED,
    ) -> tk.Frame:
        _accent, soft = self._tab_theme({UI_RED: 0, UI_BLUE: 1, UI_WARN_TEXT: 2, UI_GREEN_TEXT: 4}.get(accent, 0))
        header = self._panel(parent)
        header.configure(bg=soft, highlightbackground=accent)
        header.pack(fill="x", pady=(0, 12))
        tk.Frame(header, bg=accent, width=5).pack(side="left", fill="y")
        copy = tk.Frame(header, bg=soft)
        copy.pack(side="left", fill="both", expand=True, padx=16, pady=14)
        tk.Label(copy, text=title, bg=soft, fg=UI_TEXT, font=("Segoe UI", 18, "bold")).pack(anchor="w")
        tk.Label(copy, text=subtitle, bg=soft, fg=UI_MUTED, font=("Segoe UI", 10)).pack(anchor="w", pady=(3, 0))
        if chips:
            chip_row = tk.Frame(header, bg=soft)
            chip_row.pack(side="right", padx=16, pady=14)
            for text, color in chips:
                tk.Label(
                    chip_row,
                    text=text,
                    bg=UI_SURFACE,
                    fg=color,
                    font=("Segoe UI", 9, "bold"),
                    padx=10,
                    pady=6,
                    highlightthickness=1,
                    highlightbackground=UI_BORDER,
                ).pack(side="left", padx=(0, 8))
        return header

    def _build_shell(self) -> None:
        header = tk.Frame(self, bg=UI_HEADER, height=72)
        header.pack(fill="x")
        header.pack_propagate(False)

        logo_path = asset_path("macys-ap-icon.png")
        if logo_path.exists():
            image = Image.open(logo_path).resize((50, 50), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(image)
            tk.Label(header, image=self.logo_photo, bg=UI_HEADER).pack(side="left", padx=(18, 12))

        title_box = tk.Frame(header, bg=UI_HEADER)
        title_box.pack(side="left", fill="x", expand=True)
        tk.Label(title_box, text=APP_SHORT_NAME, bg=UI_HEADER, fg=UI_TEXT, font=("Segoe UI", 14, "bold")).pack(anchor="w")
        tk.Label(title_box, text="Asset Protection access-control workstation", bg=UI_HEADER, fg=UI_MUTED, font=("Segoe UI", 9)).pack(anchor="w")

        identity = tk.Frame(header, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
        identity.configure(width=292, height=50)
        identity.pack(side="right", padx=(12, 18), pady=10)
        identity.pack_propagate(False)
        tk.Frame(identity, bg=UI_RED, width=4).pack(side="left", fill="y")
        badge = tk.Label(identity, text="AP", bg=UI_RED, fg="#ffffff", font=("Segoe UI", 9, "bold"), width=5)
        badge.pack(side="left", fill="y")
        identity_copy = tk.Frame(identity, bg=UI_SURFACE_ALT)
        identity_copy.pack(side="left", fill="both", expand=True, padx=10, pady=6)
        tk.Label(identity_copy, text="CHINA GROVE", bg=UI_SURFACE_ALT, fg=UI_TEXT, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        tk.Label(identity_copy, text=f"HEX CONVERTER | v{APP_VERSION}", bg=UI_SURFACE_ALT, fg=UI_MUTED, font=("Segoe UI", 8)).pack(anchor="w", pady=(1, 0))
        accent_lines = tk.Frame(identity, bg=UI_SURFACE_ALT)
        accent_lines.pack(side="right", padx=(0, 10), pady=12)
        tk.Frame(accent_lines, bg=UI_RED, width=54, height=3).pack(anchor="e", pady=(0, 5))
        tk.Frame(accent_lines, bg=UI_BLUE, width=74, height=3).pack(anchor="e")

        self.mode_var = tk.StringVar(value="Batch Converter")

        toolbar = tk.Frame(self, bg=UI_SURFACE_ALT, height=50, highlightthickness=1, highlightbackground=UI_BORDER)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)
        self._menu_button(toolbar, "File", [
            ("Settings", self.show_settings),
            ("Default Export Folder", self.choose_export_folder),
            ("Open Export Folder", self.open_export_folder),
            ("Create Desktop Shortcut", self.create_desktop_shortcut),
            None,
            ("Exit", self.destroy),
        ], icon="icon-folder.png", tooltip="Open settings, export folder tools, or create a desktop shortcut.").pack(side="left", padx=(14, 8), pady=8)
        self._menu_button(toolbar, "Import", [
            ("Browse Files", self.import_file),
            ("Paste Clipboard To Queue", self.paste_clipboard_to_queue),
            ("Load Sample IDs", self.load_sample),
            None,
            ("Focus Batch Scanner (F9)", self.focus_batch_scanner),
            ("Focus Single Lookup (F10)", self.focus_single_lookup),
        ], icon="icon-import.png", tooltip="Import files, paste clipboard text, or drag files onto the Input Queue box.").pack(side="left", padx=(0, 8), pady=8)
        self._menu_button(toolbar, "Export", [
            ("Export Default", self.export_default),
            None,
            ("Excel Workbook", self.export_excel),
            ("CSV Report", self.export_csv),
            ("TXT Report", self.export_txt),
            ("PDF Report", self.export_pdf),
            None,
            ("Recent Exports", self.show_recent_exports),
        ], icon="icon-excel.png", tooltip="Save the current conversion results as a report. Settings controls the default export type.").pack(side="left", padx=(0, 8), pady=8)
        self._menu_button(toolbar, "Help", [
            ("How To Use", self.show_help),
            ("About", self.show_about),
            None,
            ("Copy Last Error Report", self.copy_error_report),
        ], icon="icon-help.png", tooltip="Open usage help, app details, and support links.").pack(side="right", padx=(0, 16), pady=8)
        self._button(
            toolbar,
            "BlueWave",
            self.open_bluewave,
            icon="icon-bluewave.png",
            tooltip="Open the BlueWave access-control site in your browser.",
        ).pack(side="right", padx=(0, 8), pady=8)

        body = tk.Frame(self, bg=UI_BG)
        body.pack(fill="both", expand=True, padx=16, pady=16)
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=1)

        nav = self._panel(body)
        nav.grid(row=0, column=0, sticky="ns", padx=(0, 14))
        nav.configure(width=244, bg=UI_SURFACE)
        nav.grid_propagate(False)
        workspace_intro = tk.Frame(nav, bg=UI_SURFACE)
        workspace_intro.pack(fill="x", padx=10, pady=(12, 8))
        workspace_badge = tk.Label(workspace_intro, text="AP", bg=UI_RED, fg="#ffffff", font=("Segoe UI", 9, "bold"), width=4, pady=5)
        workspace_badge.pack(side="left", padx=(0, 9))
        workspace_copy = tk.Frame(workspace_intro, bg=UI_SURFACE)
        workspace_copy.pack(side="left", fill="x", expand=True)
        tk.Label(workspace_copy, text="Workspace Menu", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        tk.Label(workspace_copy, text="Select a task below", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 8)).pack(anchor="w", pady=(1, 0))
        self.nav_buttons: list[tk.Button] = []
        self.nav_button_themes: list[tuple[str, str]] = []
        nav_specs = [
            ("Convert", "Batch Converter", 0, "nav-batch.png"),
            ("Convert", "Single Lookup", 1, "nav-single.png"),
            ("Reverse", "FC/CN to Hex", 2, "nav-reverse.png"),
            ("Reverse", "Unconvert Batch", 3, "nav-unconvert.png"),
            ("Review", "History", 4, "nav-history.png"),
        ]
        nav_tips = {
            "Batch Converter": "Convert many HEX IDs into Facility Code and Card Number rows.",
            "Single Lookup": "Quickly convert one HEX ID and copy the FC/CN pair.",
            "FC/CN to Hex": "Create one HEX ID from a Facility Code and Card Number.",
            "Unconvert Batch": "Convert many FC/CN pairs back into HEX IDs.",
            "History": "Review recent conversion activity saved by this utility.",
        }
        last_section = ""
        for section, label, index, icon_name in nav_specs:
            if section != last_section:
                self._nav_section(nav, section)
                last_section = section
            accent, soft = self._tab_theme(index)
            button = self._nav_button(nav, label, index, self._load_nav_icon(icon_name), nav_tips.get(label), accent, soft)
            self.nav_buttons.append(button)
            self.nav_button_themes.append((accent, soft))
            button.pack(fill="x", padx=10, pady=(0, 8))
        tk.Frame(nav, bg=UI_BORDER, height=1).pack(fill="x", padx=12, pady=(8, 10))
        self.nav_status = tk.StringVar(value="Batch Converter\nScan, import, convert, export.")
        status_card = self._card(nav, bg=UI_SURFACE, border=UI_BORDER)
        status_card.pack(fill="x", padx=10, pady=(0, 8))
        status_card.configure(height=96)
        status_card.pack_propagate(False)
        self.work_focus_bar = tk.Frame(status_card, bg=UI_RED, height=3)
        self.work_focus_bar.pack(fill="x")
        status_head = tk.Frame(status_card, bg=UI_SURFACE)
        status_head.pack(fill="x", padx=10, pady=(9, 4))
        status_icon = self._load_icon("icon-status.png", 18)
        if status_icon:
            self.focus_icon_label = tk.Label(status_head, image=status_icon, bg=UI_SURFACE)
            self.focus_icon_label.pack(side="left", padx=(0, 6))
        else:
            self.focus_icon_label = None
        self.focus_heading_label = tk.Label(status_head, text="WORKSPACE STATUS", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 8, "bold"))
        self.focus_heading_label.pack(side="left")
        self.focus_value_frame = self._info_strip(status_card, UI_RED_SOFT)
        self.focus_value_frame.pack(fill="x", padx=10, pady=(0, 9))
        self.nav_status_label = tk.Label(
            self.focus_value_frame,
            textvariable=self.nav_status,
            bg=UI_RED_SOFT,
            fg=UI_BLUE,
            wraplength=166,
            justify="left",
            anchor="w",
            height=2,
            font=("Segoe UI", 8, "bold"),
            padx=8,
            pady=5,
        )
        self.nav_status_label.pack(fill="x")
        self.focus_card_widgets = [self.focus_value_frame, self.nav_status_label]
        for widget in (status_card, status_head, self.focus_value_frame, self.nav_status_label):
            widget.configure(cursor="hand2")
            widget.bind("<Button-1>", self.open_sidebar_focus_target, add="+")
        ToolTip(status_card, "Shows the current work focus. Click to open that workspace.")

        self.sidebar_tip_title_var = tk.StringVar(value="F9 Batch Scan")
        self.sidebar_tip_body_var = tk.StringVar(value="Focus scanner input so the next scan lands at the top of the queue.")
        tip_card = self._card(nav, bg=UI_SURFACE, border=UI_BORDER)
        tip_card.pack(fill="x", padx=10, pady=(0, 10))
        tip_card.configure(height=132)
        tip_card.pack_propagate(False)
        self.sidebar_tip_bar = tk.Frame(tip_card, bg=UI_BLUE, height=3)
        self.sidebar_tip_bar.pack(fill="x")
        tip_head = tk.Frame(tip_card, bg=UI_SURFACE)
        tip_head.pack(fill="x", padx=10, pady=(9, 4))
        tip_icon = self._load_icon("icon-help.png", 18)
        if tip_icon:
            self.guide_icon_label = tk.Label(tip_head, image=tip_icon, bg=UI_SURFACE)
            self.guide_icon_label.pack(side="left", padx=(0, 6))
        else:
            self.guide_icon_label = None
        self.guide_heading_label = tk.Label(tip_head, text="FIELD GUIDE", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 8, "bold"))
        self.guide_heading_label.pack(side="left")
        self.guide_value_frame = self._info_strip(tip_card, UI_RED_SOFT)
        self.guide_value_frame.pack(fill="x", padx=10, pady=(0, 9))
        self.sidebar_tip_title_label = tk.Label(
            self.guide_value_frame,
            textvariable=self.sidebar_tip_title_var,
            bg=UI_RED_SOFT,
            fg=UI_RED,
            wraplength=166,
            justify="left",
            anchor="w",
            font=("Segoe UI", 8, "bold"),
            padx=8,
            pady=3,
        )
        self.sidebar_tip_title_label.pack(fill="x")
        self.sidebar_tip_body_label = tk.Label(
            self.guide_value_frame,
            textvariable=self.sidebar_tip_body_var,
            bg=UI_RED_SOFT,
            fg=UI_TEXT,
            wraplength=194,
            justify="left",
            anchor="nw",
            font=("Segoe UI", 8),
            padx=8,
            pady=3,
        )
        self.sidebar_tip_body_label.pack(fill="both", expand=True)
        self.guide_card_widgets = [self.guide_value_frame, self.sidebar_tip_title_label, self.sidebar_tip_body_label]
        ToolTip(tip_card, "Shows a new shortcut or usage note about every 15 seconds.")

        content = tk.Frame(body, bg=UI_BG)
        content.grid(row=0, column=1, sticky="nsew")
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)

        self.batch_tab = tk.Frame(content, bg=UI_BG)
        self.single_tab = tk.Frame(content, bg=UI_BG)
        self.reverse_tab = tk.Frame(content, bg=UI_BG)
        self.unconvert_tab = tk.Frame(content, bg=UI_BG)
        self.history_tab = tk.Frame(content, bg=UI_BG)
        self.tab_frames = [self.batch_tab, self.single_tab, self.reverse_tab, self.unconvert_tab, self.history_tab]
        for frame in self.tab_frames:
            frame.grid(row=0, column=0, sticky="nsew")

        self._build_batch_tab()
        self._build_single_tab()
        self._build_reverse_tab()
        self._build_unconvert_tab()
        self._build_history_tab()
        self.select_tab(0)

        footer = tk.Frame(self, bg=UI_HEADER, height=34, highlightthickness=1, highlightbackground=UI_BORDER)
        footer.pack(side="bottom", fill="x")
        footer.pack_propagate(False)
        self.status_var = tk.StringVar(value="Scan, paste, import, or choose a workspace.")
        self.status_state_var = tk.StringVar(value="READY")
        self.status_state_chip = tk.Label(
            footer,
            textvariable=self.status_state_var,
            bg=UI_GREEN_SOFT,
            fg=UI_GREEN_TEXT,
            font=("Segoe UI", 8, "bold"),
            padx=10,
            pady=3,
            width=13,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
        )
        self.status_state_chip.pack(side="left", padx=(12, 8))
        self.busy_var = tk.StringVar(value="")
        self.busy_label = tk.Label(
            footer,
            textvariable=self.busy_var,
            bg=UI_RED_SOFT,
            fg=UI_RED,
            font=("Segoe UI", 8, "bold"),
            padx=9,
            pady=3,
            highlightthickness=1,
            highlightbackground=UI_RED_SOFT,
        )
        self.footer_status_label = tk.Label(footer, textvariable=self.status_var, bg=UI_HEADER, fg=UI_MUTED, anchor="w")
        self.footer_status_label.pack(side="left", fill="x", expand=True)
        for widget in (self.status_state_chip, self.footer_status_label):
            widget.configure(cursor="hand2")
            widget.bind("<Button-1>", self.open_status_target, add="+")
            ToolTip(widget, "Click to open the workspace related to this status message.")
        tk.Label(
            footer,
            text=f"Version {APP_VERSION} | Contact and project links are in Help > About",
            bg=UI_HEADER,
            fg=UI_MUTED,
            font=("Segoe UI", 8),
            anchor="e",
        ).pack(side="right", padx=(12, 12))
        self._apply_corporate_skin(self)

    def _build_batch_tab(self) -> None:
        workspace_header = self._panel(self.batch_tab)
        workspace_header.configure(bg=UI_RED_SOFT, highlightbackground=UI_RED)
        workspace_header.pack(fill="x", pady=(0, 12))
        tk.Frame(workspace_header, bg=UI_RED, width=5).pack(side="left", fill="y")
        header_copy = tk.Frame(workspace_header, bg=UI_RED_SOFT)
        header_copy.pack(side="left", fill="both", expand=True, padx=16, pady=14)
        tk.Label(header_copy, text="Batch Converter", bg=UI_RED_SOFT, fg=UI_TEXT, font=("Segoe UI", 18, "bold")).pack(anchor="w")
        tk.Label(
            header_copy,
            text="Import, clean, convert, review, and export access-control IDs from one workspace.",
            bg=UI_RED_SOFT,
            fg=UI_MUTED,
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(3, 0))
        header_tools = tk.Frame(workspace_header, bg=UI_RED_SOFT)
        header_tools.pack(side="right", padx=16, pady=14)
        for text, color in [("Drag/drop ready", UI_BLUE), (f"Version {APP_VERSION}", UI_RED)]:
            chip = tk.Label(
                header_tools,
                text=text,
                bg=UI_SURFACE,
                fg=color,
                font=("Segoe UI", 9, "bold"),
                padx=10,
                pady=6,
                highlightthickness=1,
                highlightbackground=UI_BORDER,
            )
            chip.pack(side="left", padx=(0, 8))

        top = tk.Frame(self.batch_tab, bg="#0b0d12")
        top.pack(fill="x", pady=(0, 12))

        input_panel = self._panel(top)
        input_panel.pack(side="left", fill="both", expand=True, padx=(0, 12))
        queue_header = tk.Frame(input_panel, bg="#10141b")
        queue_header.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(queue_header, text="Input Queue", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(side="left")
        legend = tk.Frame(queue_header, bg="#10141b")
        legend.pack(side="right")
        for label, bg, fg in [("Valid", UI_GREEN_SOFT, UI_GREEN_TEXT), ("Warning", UI_WARN_SOFT, UI_WARN_TEXT), ("Invalid", UI_BAD_SOFT, UI_BAD_TEXT)]:
            tk.Label(
                legend,
                text=label,
                bg=bg,
                fg=fg,
                font=("Segoe UI", 8, "bold"),
                padx=8,
                pady=3,
                highlightthickness=1,
                highlightbackground=UI_BORDER,
            ).pack(side="left", padx=(6, 0))
        self.queue_count_var = tk.StringVar(value="0 rows")
        tk.Label(
            queue_header,
            textvariable=self.queue_count_var,
            bg=UI_SURFACE_ALT,
            fg=UI_BLUE,
            font=("Segoe UI", 8, "bold"),
            padx=8,
            pady=3,
            width=10,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
        ).pack(side="right", padx=(0, 10))
        tk.Label(
            input_panel,
            text="Paste HEX IDs, import files, drag supported files onto this box, or use Scanner Input for handheld scans.",
            bg="#10141b",
            fg="#a8b2c2",
            font=("Segoe UI", 9),
            wraplength=520,
            justify="left",
        ).pack(anchor="w", padx=12, pady=(0, 8))
        scan_panel = tk.Frame(input_panel, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
        scan_panel.pack(fill="x", padx=12, pady=(0, 8))
        tk.Frame(scan_panel, bg=UI_RED, width=4).pack(side="left", fill="y")
        scan_copy = tk.Frame(scan_panel, bg=UI_SURFACE_ALT)
        scan_copy.pack(side="left", fill="x", expand=True, padx=10, pady=8)
        scan_title_row = tk.Frame(scan_copy, bg=UI_SURFACE_ALT)
        scan_title_row.pack(fill="x")
        tk.Label(scan_title_row, text="Scanner Input", bg=UI_SURFACE_ALT, fg=UI_TEXT, font=("Segoe UI", 10, "bold")).pack(side="left")
        self.batch_scan_count_var = tk.StringVar(value="0 scans this session")
        tk.Label(
            scan_title_row,
            textvariable=self.batch_scan_count_var,
            bg="#eef6ff",
            fg=UI_BLUE,
            font=("Segoe UI", 8, "bold"),
            padx=8,
            pady=3,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
        ).pack(side="left", padx=(10, 0))
        tk.Label(
            scan_title_row,
            text="F9 focuses scanner",
            bg=UI_SURFACE,
            fg=UI_GREEN_TEXT,
            font=("Segoe UI", 8, "bold"),
            padx=8,
            pady=3,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
        ).pack(side="left", padx=(6, 0))
        tk.Label(
            scan_copy,
            text="Focus here, scan an ID, and press Enter/Tab from the scanner. New scans go to the top of the queue.",
            bg=UI_SURFACE_ALT,
            fg=UI_MUTED,
            font=("Segoe UI", 8),
            wraplength=430,
            justify="left",
        ).pack(anchor="w", pady=(2, 0))
        self.batch_scan_var = tk.StringVar()
        scan_entry_shell = tk.Frame(scan_panel, bg=UI_INPUT, highlightthickness=1, highlightbackground=UI_RED)
        scan_entry_shell.pack(side="right", fill="x", expand=False, padx=(8, 10), pady=10)
        scan_icon = self._load_icon("icon-search.png", 16)
        if scan_icon:
            tk.Label(scan_entry_shell, image=scan_icon, bg=UI_INPUT).pack(side="left", padx=(8, 3))
        self.batch_scan_entry = tk.Entry(
            scan_entry_shell,
            textvariable=self.batch_scan_var,
            bg=UI_INPUT,
            fg=UI_TEXT,
            insertbackground=UI_RED,
            relief="flat",
            width=22,
            font=("Cascadia Mono", 11, "bold"),
        )
        self.batch_scan_entry.pack(side="left", fill="x", ipady=7, padx=(0, 8))
        self.batch_scan_entry.bind("<Return>", self.handle_batch_scan_submit)
        self.batch_scan_entry.bind("<KP_Enter>", self.handle_batch_scan_submit)
        self.batch_scan_entry.bind("<Tab>", self.handle_batch_scan_submit)
        self.batch_scan_entry.bind("<KeyRelease>", self.schedule_batch_scan_autofill)
        self._enable_edit_context_menu(self.batch_scan_entry)
        ToolTip(self.batch_scan_entry, "Most handheld scanners type like a keyboard. Configure the scanner suffix to Enter or Tab when possible.")
        multi_frame = tk.Frame(input_panel, bg="#10141b")
        multi_frame.pack(fill="both", expand=True, padx=12)
        multi_frame.rowconfigure(0, weight=1)
        multi_frame.columnconfigure(0, weight=1)
        self.multi_text = tk.Text(
            multi_frame,
            height=4,
            bg=UI_INPUT,
            fg=UI_TEXT,
            insertbackground=UI_RED,
            relief="flat",
            highlightthickness=1,
            highlightbackground=UI_BLUE,
            highlightcolor=UI_RED,
            padx=10,
            pady=8,
            font=("Cascadia Mono", 10),
            width=48,
            wrap="none",
        )
        self.multi_text.tag_configure("input_valid", background=UI_GREEN_SOFT, foreground=UI_GREEN_TEXT)
        self.multi_text.tag_configure("input_warning", background=UI_WARN_SOFT, foreground=UI_WARN_TEXT)
        self.multi_text.tag_configure("input_invalid", background=UI_BAD_SOFT, foreground=UI_BAD_TEXT)
        multi_scroll_y = ttk.Scrollbar(multi_frame, orient="vertical", command=self.multi_text.yview, style="Dark.Vertical.TScrollbar")
        multi_scroll_x = ttk.Scrollbar(multi_frame, orient="horizontal", command=self.multi_text.xview, style="Dark.Horizontal.TScrollbar")
        self.multi_text.configure(yscrollcommand=multi_scroll_y.set, xscrollcommand=multi_scroll_x.set)
        self.multi_text.grid(row=0, column=0, sticky="nsew")
        multi_scroll_y.grid(row=0, column=1, sticky="ns")
        multi_scroll_x.grid(row=1, column=0, sticky="ew")
        self.multi_text.insert("1.0", "")
        self.multi_text.bind("<<Modified>>", self.handle_batch_input_changed)
        self._enable_mousewheel(self.multi_text, self.multi_text, self.multi_text)
        self._enable_edit_context_menu(self.multi_text, lambda: self.paste_clipboard_to_queue(clean_only=True), "Paste Clean IDs")
        self._enable_drop_target(self.multi_text)
        self._enable_drop_target(multi_frame)
        self._add_text_resize_handle(input_panel, self.multi_text, "Drag to resize Input Queue")
        btns = tk.Frame(input_panel, bg="#10141b")
        btns.pack(fill="x", padx=12, pady=8)
        batch_actions = [
            ("Import", self.import_file, False, "icon-import.png", "Browse for one or more supported files and add extracted IDs to the queue."),
            ("Sample", self.load_sample, False, "icon-sample.png", "Load numeric sample data so you can see conversion results."),
            ("Convert", self.convert_batch, True, "icon-convert.png", "Convert every queued HEX ID into Facility Code and Card Number."),
            ("Remove Duplicates", self.remove_duplicate_input_lines, False, "icon-clear.png", "Keep the first matching HEX ID and remove repeated valid IDs from the queue."),
            ("Keep Valid", self.keep_only_valid_input_lines, False, "icon-status.png", "Remove rows that cannot be read as valid 8-character HEX IDs."),
            ("Clear", self.clear_workspace, False, "icon-clear.png", "Clear input, results, and single-lookup fields."),
        ]
        for idx, (text, command, primary, icon, tooltip) in enumerate(batch_actions):
            button = self._button(btns, text, command, primary, icon=icon, tooltip=tooltip, compact=True)
            button.pack(side="left", padx=(0, 6), pady=0)

        summary = self._panel(top)
        summary.pack(side="right", fill="y")
        summary.configure(width=262)
        summary.pack_propagate(False)
        tk.Label(summary, text="Run Summary", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(10, 2))
        tk.Label(summary, text="Live queue and conversion status.", bg="#10141b", fg="#a8b2c2", font=("Segoe UI", 9), wraplength=230, justify="left").pack(anchor="w", padx=12, pady=(0, 6))
        self.stat_vars = {name: tk.StringVar(value="0") for name in ["Input", "Valid", "Invalid", "Warnings"]}
        grid = tk.Frame(summary, bg="#10141b")
        grid.pack(fill="x", padx=12, pady=(4, 8))
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)
        stat_colors = {"Input": UI_BLUE, "Valid": UI_GREEN_TEXT, "Invalid": UI_BAD_TEXT, "Warnings": UI_WARN_TEXT}
        for idx, name in enumerate(self.stat_vars):
            card = tk.Frame(grid, bg=UI_SURFACE_ALT, width=104, height=58, highlightthickness=1, highlightbackground=UI_BORDER)
            card.grid(row=idx // 2, column=idx % 2, sticky="ew", padx=5, pady=5)
            card.grid_propagate(False)
            tk.Frame(card, bg=stat_colors[name], height=3).pack(fill="x")
            tk.Label(card, text=name, bg=UI_SURFACE_ALT, fg=UI_MUTED, font=("Segoe UI", 8, "bold")).pack(pady=(5, 0))
            tk.Label(card, textvariable=self.stat_vars[name], bg=UI_SURFACE_ALT, fg=stat_colors[name], font=("Segoe UI", 14, "bold")).pack()
        self.time_var = tk.StringVar(value="No run yet")
        tk.Label(
            summary,
            textvariable=self.time_var,
            bg=UI_SURFACE_ALT,
            fg=UI_TEXT,
            wraplength=220,
            justify="center",
            padx=10,
            pady=6,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
        ).pack(fill="x", padx=12, pady=(3, 6))
        self.notice_var = tk.StringVar(value="")
        tk.Label(summary, textvariable=self.notice_var, bg="#10141b", fg="#ffca63", wraplength=230, justify="left").pack(anchor="w", padx=12, pady=(0, 12))
        results_panel = self._panel(self.batch_tab)
        results_panel.pack(fill="both", expand=True)
        header = tk.Frame(results_panel, bg="#10141b")
        header.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(header, text="Results", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(side="left")
        copy_row = tk.Frame(header, bg="#10141b")
        copy_row.pack(side="right")
        self._button(copy_row, "Copy All", self.copy_all_pairs, icon="icon-copy.png", tooltip="Copy all valid FC,CN pairs to the clipboard.").pack(side="left", padx=(0, 6))
        self._button(copy_row, "Clear Invalid", self.clear_invalid_rows, icon="icon-clear.png", tooltip="Remove invalid rows from the review table.").pack(side="left")

        filter_row = tk.Frame(results_panel, bg="#10141b")
        filter_row.pack(fill="x", padx=12, pady=(0, 8))
        tk.Label(filter_row, text="Search", bg="#10141b", fg="#a8b2c2", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 8))
        self.search_var = tk.StringVar()
        search_shell = tk.Frame(filter_row, bg=UI_INPUT, highlightthickness=1, highlightbackground=UI_BLUE)
        search_shell.pack(side="left", fill="x", expand=True, padx=(0, 10))
        search_icon = self._load_icon("icon-search.png", 16)
        if search_icon:
            tk.Label(search_shell, image=search_icon, bg=UI_INPUT).pack(side="left", padx=(9, 4))
        self.search_entry = tk.Entry(
            search_shell,
            textvariable=self.search_var,
            bg=UI_INPUT,
            fg=UI_TEXT,
            insertbackground=UI_RED,
            relief="flat",
            width=32,
            font=("Segoe UI", 10),
        )
        self.search_entry.pack(side="left", fill="x", expand=True, ipady=7, padx=(0, 8))
        self._enable_edit_context_menu(self.search_entry)
        self.search_var.trace_add("write", lambda *_args: self.render_results())
        tk.Label(filter_row, text="Status", bg="#10141b", fg="#a8b2c2", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 8))
        self.status_filter_var = tk.StringVar(value="All")
        status_menu = tk.OptionMenu(filter_row, self.status_filter_var, "All", "Valid", "Warning", "Invalid", command=lambda _value: self.render_results())
        status_menu.configure(bg=UI_SURFACE_ALT, fg=UI_TEXT, activebackground="#eef2f6", activeforeground=UI_TEXT, relief="flat", highlightthickness=1, highlightbackground=UI_BORDER, font=("Segoe UI", 9, "bold"), padx=10, pady=5)
        self._style_menu(status_menu["menu"])
        status_menu.pack(side="left")

        self.results_empty_frame = tk.Frame(results_panel, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
        self.results_empty_frame.pack(fill="x", padx=12, pady=(0, 8))
        empty_badge = tk.Frame(self.results_empty_frame, bg=UI_RED, width=54, height=54)
        empty_badge.pack(side="left", padx=14, pady=14)
        empty_badge.pack_propagate(False)
        tk.Label(empty_badge, text="AP", bg=UI_RED, fg="#ffffff", font=("Segoe UI", 14, "bold")).pack(expand=True)
        empty_copy = tk.Frame(self.results_empty_frame, bg=UI_SURFACE_ALT)
        empty_copy.pack(side="left", fill="both", expand=True, padx=(0, 12), pady=14)
        tk.Label(empty_copy, text="Ready for results", bg=UI_SURFACE_ALT, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(empty_copy, text="Paste or import access-control IDs, then use Convert. Valid, warning, and invalid rows will appear here.", bg=UI_SURFACE_ALT, fg=UI_MUTED, wraplength=720, justify="left", font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 0))

        table_frame = tk.Frame(results_panel, bg="#10141b")
        self.results_table_frame = table_frame
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        columns = ("Line", "Hex ID", "FC", "CN", "Status", "Notes")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        result_heading_text = {"Notes": "Notes / Details"}
        result_anchors = {"Line": "center", "Hex ID": "center", "FC": "center", "CN": "center", "Status": "center", "Notes": "w"}
        for col in columns:
            self.tree.heading(col, text=result_heading_text.get(col, col), anchor=result_anchors[col], command=lambda c=col: self.sort_results_by(c))
        widths = {"Line": 60, "Hex ID": 150, "FC": 90, "CN": 90, "Status": 120, "Notes": 420}
        for col, width in widths.items():
            self.tree.column(col, width=width, minwidth=60, anchor=result_anchors[col], stretch=col == "Notes")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview, style="Dark.Vertical.TScrollbar")
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview, style="Dark.Horizontal.TScrollbar")
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.tree.tag_configure("valid", background=UI_GREEN_SOFT, foreground=UI_GREEN_TEXT)
        self.tree.tag_configure("invalid", background=UI_BAD_SOFT, foreground=UI_BAD_TEXT)
        self.tree.tag_configure("warning", background=UI_WARN_SOFT, foreground=UI_WARN_TEXT)
        self.tree.tag_configure("empty", foreground=UI_MUTED)
        self.tree.bind("<Double-1>", lambda _event: self.copy_selected("pair"))
        self.results_context_menu = self._context_menu(
            [
                ("Copy Full Row", lambda: self.copy_selected("row")),
                ("Copy FC,CN Pair", lambda: self.copy_selected("pair")),
                ("Copy Facility Code", lambda: self.copy_selected("fc")),
                ("Copy Card Number", lambda: self.copy_selected("cn")),
                ("Copy HEX / Raw", lambda: self.copy_selected("hex")),
                None,
                ("Copy Notes", lambda: self.copy_selected("notes")),
            ]
        )
        self.tree.bind("<Button-3>", lambda event: self._show_tree_context_menu(event, self.tree, self.results_context_menu))
        self._enable_mousewheel(self.tree, self.tree, self.tree)
        self._enable_tree_row_hover(self.tree)
        self._enable_tree_header_tooltips(
            self.tree,
            {
                "Line": "Original row number from the Input Queue.",
                "Hex ID": "The 8-character access-control HEX value.",
                "FC": "Facility Code. This is the high 16-bit value from the HEX ID.",
                "CN": "Card Number. This is the low 16-bit value from the HEX ID.",
                "Status": "Valid, Warning, or Invalid conversion result.",
                "Notes": "Details appear only when the app cleaned a row, found a duplicate, flagged an unusual value, or explains invalid input.",
            },
        )

    def _build_single_tab(self) -> None:
        self._workspace_title(
            self.single_tab,
            "Single Hex Lookup",
            "Convert one HEX ID and copy the FC/CN pair for quick checks.",
            [("Quick check", UI_BLUE), ("Copies FC,CN", UI_GREEN_TEXT)],
            UI_BLUE,
        )
        top = tk.Frame(self.single_tab, bg=UI_BG)
        top.pack(fill="x", pady=(0, 12))

        panel = self._panel(top)
        panel.pack(side="left", fill="both", expand=True, padx=(0, 12))
        tk.Frame(panel, bg=UI_BLUE, height=3).pack(fill="x")
        tk.Label(panel, text="Scanner / Single Lookup", bg="#10141b", fg="#ffffff", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        tk.Label(
            panel,
            text="Scan or type one 8-character HEX ID. A handheld scanner can send Enter or Tab after the scan; the app converts automatically.",
            bg="#10141b",
            fg="#a8b2c2",
            font=("Segoe UI", 9),
            wraplength=560,
            justify="left",
        ).pack(anchor="w", padx=12, pady=(0, 10))
        self.single_var = tk.StringVar()
        entry_shell = tk.Frame(panel, bg=UI_INPUT, highlightthickness=1, highlightbackground=UI_BLUE)
        entry_shell.configure(width=360, height=48)
        entry_shell.pack(anchor="w", padx=12, pady=(0, 10))
        entry_shell.pack_propagate(False)
        single_icon = self._load_icon("nav-single.png", 18)
        if single_icon:
            tk.Label(entry_shell, image=single_icon, bg=UI_INPUT).pack(side="left", padx=(10, 5))
        entry = tk.Entry(
            entry_shell,
            textvariable=self.single_var,
            bg=UI_INPUT,
            fg=UI_TEXT,
            insertbackground=UI_RED,
            relief="flat",
            font=("Cascadia Mono", 12),
            width=24,
        )
        entry.pack(side="left", fill="x", expand=True, ipady=10, padx=(0, 10))
        self.single_entry = entry
        entry.bind("<Return>", lambda _event: self.convert_single(source="scan"))
        entry.bind("<KP_Enter>", lambda _event: self.convert_single(source="scan"))
        entry.bind("<Tab>", self.handle_single_scan_tab)
        entry.bind("<KeyRelease>", self.schedule_single_scan_autoconvert)
        self._enable_edit_context_menu(entry)
        row = tk.Frame(panel, bg="#10141b")
        row.pack(fill="x", padx=12, pady=(0, 12))
        self._button(row, "Convert", self.convert_single, True, icon="icon-convert.png", tooltip="Convert one HEX ID and copy the FC,CN pair.", compact=True).pack(side="left", padx=(0, 6))
        self._button(row, "Clear", self.clear_single, icon="icon-clear.png", tooltip="Clear the single lookup field.", compact=True).pack(side="left", padx=(0, 6))
        self._button(row, "Focus Scanner", lambda: self.single_entry.focus_set(), icon="icon-search.png", tooltip="Move the cursor back to the scanner input field.", compact=True).pack(side="left")

        info = self._card(panel, bg=UI_SURFACE_ALT, border=UI_BORDER)
        info.pack(fill="x", padx=12, pady=(0, 12))
        tk.Frame(info, bg=UI_BLUE, width=4).pack(side="left", fill="y")
        tk.Label(
            info,
            text="Scanner tip: most USB handheld scanners work as keyboard input. Enter or Tab as the scanner suffix is recommended.",
            bg=UI_SURFACE_ALT,
            fg=UI_MUTED,
            justify="left",
            wraplength=560,
            padx=12,
            pady=10,
            font=("Segoe UI", 9),
        ).pack(side="left", fill="x", expand=True)

        result_panel = self._panel(top)
        result_panel.pack(side="right", fill="both")
        result_panel.configure(width=330, height=300)
        result_panel.pack_propagate(False)
        tk.Frame(result_panel, bg=UI_BLUE, height=3).pack(fill="x")
        tk.Label(result_panel, text="Lookup Result", bg="#10141b", fg="#ffffff", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        tk.Label(result_panel, text="Result is copied as FC,CN after conversion.", bg="#10141b", fg="#a8b2c2", font=("Segoe UI", 9), wraplength=290, justify="left").pack(anchor="w", padx=12, pady=(0, 10))
        self.single_result = tk.StringVar(value="Waiting for one 8-character HEX ID.")
        self.single_hex_var = tk.StringVar(value="--")
        self.single_fc_var = tk.StringVar(value="--")
        self.single_cn_var = tk.StringVar(value="--")
        self.single_note_var = tk.StringVar(value="No lookup has run yet.")

        def result_row(label_text: str, variable: tk.StringVar, accent: str) -> None:
            item = tk.Frame(result_panel, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
            item.pack(fill="x", padx=12, pady=(0, 8))
            tk.Frame(item, bg=accent, width=4).pack(side="left", fill="y")
            copy = tk.Frame(item, bg=UI_SURFACE_ALT)
            copy.pack(side="left", fill="x", expand=True, padx=10, pady=8)
            tk.Label(copy, text=label_text.upper(), bg=UI_SURFACE_ALT, fg=UI_MUTED, font=("Segoe UI", 8, "bold")).pack(anchor="w")
            self._clickable_value_label(copy, variable, accent, label_text, 1).pack(anchor="w", pady=(2, 0))

        result_row("Hex ID", self.single_hex_var, UI_BLUE)
        result_row("Facility Code", self.single_fc_var, UI_GREEN_TEXT)
        result_row("Card Number", self.single_cn_var, UI_RED)
        tk.Label(result_panel, textvariable=self.single_note_var, bg="#10141b", fg=UI_WARN_TEXT, wraplength=290, justify="left", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12, pady=(2, 12))

    def _build_reverse_tab(self) -> None:
        self._workspace_title(
            self.reverse_tab,
            "FC/CN to Hex",
            "Build one 8-character HEX ID from Facility Code and Card Number.",
            [("Reverse lookup", UI_WARN_TEXT), ("Local conversion", UI_GREEN_TEXT)],
            UI_WARN_TEXT,
        )
        top = tk.Frame(self.reverse_tab, bg=UI_BG)
        top.pack(fill="x", pady=(0, 12))

        panel = self._panel(top)
        panel.pack(side="left", fill="both", expand=True, padx=(0, 12))
        tk.Frame(panel, bg=UI_WARN_TEXT, height=3).pack(fill="x")
        tk.Label(panel, text="Facility Code and Card Number", bg="#10141b", fg="#ffffff", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        tk.Label(
            panel,
            text="Enter whole numbers from 0 to 65535. Press Enter in either field or use Convert to build the HEX ID.",
            bg="#10141b",
            fg="#a8b2c2",
            font=("Segoe UI", 9),
            wraplength=560,
            justify="left",
        ).pack(anchor="w", padx=12, pady=(0, 10))
        self.fc_var = tk.StringVar()
        self.cn_var = tk.StringVar()
        entry_grid = tk.Frame(panel, bg="#10141b")
        entry_grid.pack(anchor="w", padx=12, pady=(0, 10))
        for column, (label, var) in enumerate([("Facility Code", self.fc_var), ("Card Number", self.cn_var)]):
            shell = tk.Frame(entry_grid, bg=UI_SURFACE_ALT)
            shell.configure(width=260, height=78)
            shell.pack(side="left", padx=(0, 12) if column == 0 else (0, 0))
            shell.pack_propagate(False)
            tk.Label(shell, text=label.upper(), bg=UI_SURFACE_ALT, fg=UI_MUTED, font=("Segoe UI", 8, "bold")).pack(anchor="w", padx=10, pady=(8, 3))
            entry = tk.Entry(
                shell,
                textvariable=var,
                bg=UI_INPUT,
                fg=UI_TEXT,
                insertbackground=UI_RED,
                relief="flat",
                highlightthickness=1,
                highlightbackground=UI_WARN_TEXT,
                highlightcolor=UI_RED,
                font=("Cascadia Mono", 12),
                width=14,
            )
            entry.pack(fill="x", padx=10, pady=(0, 10), ipady=9)
            entry.bind("<Return>", lambda _event: self.convert_reverse())
            entry.bind("<KP_Enter>", lambda _event: self.convert_reverse())
            self._enable_edit_context_menu(entry)
            if label == "Facility Code":
                self.fc_entry = entry
            else:
                self.cn_entry = entry
        row = tk.Frame(panel, bg="#10141b")
        row.pack(fill="x", padx=12, pady=(0, 12))
        self._button(row, "Convert", self.convert_reverse, True, icon="icon-convert.png", tooltip="Build one 8-character HEX ID from FC and CN.", compact=True).pack(side="left", padx=(0, 6))
        self._button(row, "Clear", self.clear_reverse, icon="icon-clear.png", tooltip="Clear the FC and CN fields.", compact=True).pack(side="left")

        guide = self._card(panel, bg=UI_SURFACE_ALT, border=UI_BORDER)
        guide.pack(fill="x", padx=12, pady=(0, 12))
        tk.Frame(guide, bg=UI_WARN_TEXT, width=4).pack(side="left", fill="y")
        tk.Label(
            guide,
            text="Formula: Facility Code uses the high 16 bits. Card Number uses the low 16 bits. The app combines both into one 8-character HEX value.",
            bg=UI_SURFACE_ALT,
            fg=UI_MUTED,
            justify="left",
            wraplength=560,
            padx=12,
            pady=10,
            font=("Segoe UI", 9),
        ).pack(side="left", fill="x", expand=True)

        result_panel = self._panel(top)
        result_panel.pack(side="right", fill="both")
        result_panel.configure(width=330, height=240)
        result_panel.pack_propagate(False)
        tk.Frame(result_panel, bg=UI_WARN_TEXT, height=3).pack(fill="x")
        tk.Label(result_panel, text="HEX Output", bg="#10141b", fg="#ffffff", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        tk.Label(result_panel, text="HEX is copied after conversion.", bg="#10141b", fg="#a8b2c2", font=("Segoe UI", 9), wraplength=290, justify="left").pack(anchor="w", padx=12, pady=(0, 10))
        self.reverse_result = tk.StringVar(value="Waiting for Facility Code and Card Number.")
        self.reverse_hex_var = tk.StringVar(value="--")
        self.reverse_note_var = tk.StringVar(value="No reverse lookup has run yet.")
        hex_card = tk.Frame(result_panel, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
        hex_card.pack(fill="x", padx=12, pady=(0, 8))
        tk.Frame(hex_card, bg=UI_WARN_TEXT, width=4).pack(side="left", fill="y")
        hex_copy = tk.Frame(hex_card, bg=UI_SURFACE_ALT)
        hex_copy.pack(side="left", fill="x", expand=True, padx=10, pady=10)
        tk.Label(hex_copy, text="HEX ID", bg=UI_SURFACE_ALT, fg=UI_MUTED, font=("Segoe UI", 8, "bold")).pack(anchor="w")
        self._clickable_value_label(hex_copy, self.reverse_hex_var, UI_WARN_TEXT, "HEX ID", 2, font_size=16).pack(anchor="w", pady=(2, 0))
        tk.Label(result_panel, textvariable=self.reverse_note_var, bg="#10141b", fg=UI_WARN_TEXT, wraplength=290, justify="left", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=12, pady=(2, 12))

    def _build_unconvert_tab(self) -> None:
        self._workspace_title(
            self.unconvert_tab,
            "Unconvert Batch",
            "Convert multiple Facility Code/Card Number pairs back into HEX IDs.",
            [("Batch reverse", UI_WARN_TEXT), ("Copy HEX output", UI_BLUE)],
            UI_WARN_TEXT,
        )
        top = tk.Frame(self.unconvert_tab, bg="#0b0d12")
        top.pack(fill="x", pady=(0, 12))

        input_panel = self._panel(top)
        input_panel.pack(side="left", fill="both", expand=True, padx=(0, 12))
        tk.Label(input_panel, text="Unconvert FC/CN Batch", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(12, 8))
        tk.Label(input_panel, text="Paste one pair per line. Examples: 34968,18199 or FC 34968 CN 18199.", bg="#10141b", fg="#a8b2c2", font=("Segoe UI", 9)).pack(anchor="w", padx=12, pady=(0, 8))
        unconvert_input_frame = tk.Frame(input_panel, bg="#10141b")
        unconvert_input_frame.pack(fill="both", expand=True, padx=12)
        unconvert_input_frame.rowconfigure(0, weight=1)
        unconvert_input_frame.columnconfigure(0, weight=1)
        self.unconvert_text = tk.Text(
            unconvert_input_frame,
            height=5,
            bg=UI_INPUT,
            fg=UI_TEXT,
            insertbackground=UI_RED,
            relief="flat",
            highlightthickness=1,
            highlightbackground=UI_WARN_TEXT,
            highlightcolor=UI_RED,
            padx=12,
            pady=8,
            font=("Cascadia Mono", 10),
            width=42,
            wrap="none",
        )
        unconvert_scroll_y = ttk.Scrollbar(unconvert_input_frame, orient="vertical", command=self.unconvert_text.yview, style="Dark.Vertical.TScrollbar")
        unconvert_scroll_x = ttk.Scrollbar(unconvert_input_frame, orient="horizontal", command=self.unconvert_text.xview, style="Dark.Horizontal.TScrollbar")
        self.unconvert_text.configure(yscrollcommand=unconvert_scroll_y.set, xscrollcommand=unconvert_scroll_x.set)
        self.unconvert_text.grid(row=0, column=0, sticky="nsew")
        unconvert_scroll_y.grid(row=0, column=1, sticky="ns")
        unconvert_scroll_x.grid(row=1, column=0, sticky="ew")
        self.unconvert_text.bind("<<Modified>>", self.handle_unconvert_input_changed)
        self._enable_mousewheel(self.unconvert_text, self.unconvert_text, self.unconvert_text)
        self._enable_edit_context_menu(self.unconvert_text, self.paste_clipboard_to_unconvert, "Paste FC/CN Pairs")
        self._enable_drop_target(self.unconvert_text)
        self._enable_drop_target(unconvert_input_frame)
        self._add_text_resize_handle(input_panel, self.unconvert_text, "Drag to resize FC/CN input")
        row = tk.Frame(input_panel, bg="#10141b")
        row.pack(fill="x", padx=12, pady=12)
        unconvert_actions = [
            ("Sample", self.load_unconvert_sample, False, "icon-sample.png", "Load numeric FC/CN sample pairs for the unconvert workflow."),
            ("Unconvert", self.convert_unconvert_batch, True, "icon-convert.png", "Convert every queued FC/CN pair back into HEX."),
            ("Copy All HEX", self.copy_all_unconverted_hex, False, "icon-copy.png", "Copy all valid unconverted HEX IDs."),
            ("Clear", self.clear_unconvert, False, "icon-clear.png", "Clear the unconvert input and results."),
        ]
        for idx, (text, command, primary, icon, tooltip) in enumerate(unconvert_actions):
            button = self._button(row, text, command, primary, icon=icon, tooltip=tooltip, compact=True)
            button.pack(side="left", padx=(0, 6), pady=0)

        summary = self._panel(top)
        summary.pack(side="right", fill="y")
        summary.configure(width=280)
        summary.pack_propagate(False)
        tk.Label(summary, text="Unconvert Summary", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(12, 8))
        self.unconvert_summary_var = tk.StringVar(value="No Unconvert run yet")
        tk.Label(summary, textvariable=self.unconvert_summary_var, bg="#10141b", fg="#46d9ff", wraplength=235, justify="left", font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=12, pady=(8, 12))
        tk.Label(summary, text="This tab reverses FC/CN pairs back into 8-character HEX IDs.", bg="#10141b", fg="#a8b2c2", wraplength=235, justify="left").pack(anchor="w", padx=12, pady=(0, 12))

        results_panel = self._panel(self.unconvert_tab)
        results_panel.pack(fill="both", expand=True)
        header = tk.Frame(results_panel, bg="#10141b")
        header.pack(fill="x", padx=12, pady=(12, 8))
        tk.Label(header, text="Unconvert Results", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(side="left")
        self._button(header, "Copy Selected HEX", self.copy_selected_unconvert_hex, icon="icon-copy.png", tooltip="Copy the selected unconverted HEX ID.").pack(side="right")

        self.unconvert_empty_frame = tk.Frame(results_panel, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
        self.unconvert_empty_frame.pack(fill="x", padx=12, pady=(0, 8))
        unconvert_badge = tk.Frame(self.unconvert_empty_frame, bg=UI_RED, width=54, height=54)
        unconvert_badge.pack(side="left", padx=14, pady=14)
        unconvert_badge.pack_propagate(False)
        tk.Label(unconvert_badge, text="HEX", bg=UI_RED, fg="#ffffff", font=("Segoe UI", 11, "bold")).pack(expand=True)
        empty_copy = tk.Frame(self.unconvert_empty_frame, bg=UI_SURFACE_ALT)
        empty_copy.pack(side="left", fill="both", expand=True, padx=(0, 12), pady=14)
        tk.Label(empty_copy, text="Ready to rebuild HEX IDs", bg=UI_SURFACE_ALT, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(empty_copy, text="Paste FC/CN pairs, run Unconvert, then copy returned HEX IDs from the review area.", bg=UI_SURFACE_ALT, fg=UI_MUTED, wraplength=720, justify="left", font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 0))

        table_frame = tk.Frame(results_panel, bg="#10141b")
        self.unconvert_table_frame = table_frame
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        columns = ("Line", "FC", "CN", "Hex ID", "Status", "Notes")
        self.unconvert_tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        unconvert_heading_text = {"Notes": "Notes / Details"}
        unconvert_anchors = {"Line": "center", "FC": "center", "CN": "center", "Hex ID": "center", "Status": "center", "Notes": "w"}
        for col in columns:
            self.unconvert_tree.heading(col, text=unconvert_heading_text.get(col, col), anchor=unconvert_anchors[col])
        widths = {"Line": 60, "FC": 120, "CN": 120, "Hex ID": 150, "Status": 120, "Notes": 430}
        for col, width in widths.items():
            self.unconvert_tree.column(col, width=width, minwidth=60, anchor=unconvert_anchors[col], stretch=col == "Notes")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.unconvert_tree.yview, style="Dark.Vertical.TScrollbar")
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.unconvert_tree.xview, style="Dark.Horizontal.TScrollbar")
        self.unconvert_tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.unconvert_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.unconvert_tree.tag_configure("valid", background=UI_GREEN_SOFT, foreground=UI_GREEN_TEXT)
        self.unconvert_tree.tag_configure("invalid", background=UI_BAD_SOFT, foreground=UI_BAD_TEXT)
        self.unconvert_tree.tag_configure("warning", background=UI_WARN_SOFT, foreground=UI_WARN_TEXT)
        self.unconvert_tree.tag_configure("empty", foreground=UI_MUTED)
        self.unconvert_context_menu = self._context_menu(
            [
                ("Copy Full Row", lambda: self.copy_selected_unconvert("row")),
                ("Copy HEX / Raw", lambda: self.copy_selected_unconvert("hex")),
                ("Copy Facility Code", lambda: self.copy_selected_unconvert("fc")),
                ("Copy Card Number", lambda: self.copy_selected_unconvert("cn")),
                None,
                ("Copy Notes", lambda: self.copy_selected_unconvert("notes")),
            ]
        )
        self.unconvert_tree.bind("<Button-3>", lambda event: self._show_tree_context_menu(event, self.unconvert_tree, self.unconvert_context_menu))
        self._enable_mousewheel(self.unconvert_tree, self.unconvert_tree, self.unconvert_tree)
        self._enable_tree_row_hover(self.unconvert_tree)
        self._enable_tree_header_tooltips(
            self.unconvert_tree,
            {
                "Line": "Original row number from the FC/CN input queue.",
                "FC": "Facility Code used to rebuild the HEX ID.",
                "CN": "Card Number used to rebuild the HEX ID.",
                "Hex ID": "The 8-character HEX ID created from FC and CN.",
                "Status": "Valid, Warning, or Invalid unconvert result.",
                "Notes": "Details appear when an FC/CN pair is duplicate, unusual, or invalid.",
            },
        )

    def _build_history_tab(self) -> None:
        self._workspace_title(
            self.history_tab,
            "Conversion History",
            "Review recent conversion activity saved by this utility.",
            [("Local history", UI_BLUE), ("Newest first", UI_GREEN_TEXT)],
            UI_GREEN_TEXT,
        )
        panel = self._panel(self.history_tab)
        panel.pack(fill="both", expand=True, padx=2, pady=2)
        header = tk.Frame(panel, bg="#10141b")
        header.pack(fill="x", padx=12, pady=(12, 8))
        heading = tk.Frame(header, bg="#10141b")
        heading.pack(side="left", fill="x", expand=True)
        tk.Label(heading, text="Recent Activity", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(
            heading,
            text="Newest conversions appear first. Rows with problems are softly highlighted for review.",
            bg="#10141b",
            fg="#a8b2c2",
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(3, 0))
        self._button(header, "Clear History", self.clear_history, icon="icon-clear.png", tooltip="Delete saved conversion history from this app.").pack(side="right")

        self.history_empty_frame = tk.Frame(panel, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
        self.history_empty_frame.pack(fill="x", padx=12, pady=(0, 8))
        empty_badge = tk.Frame(self.history_empty_frame, bg=UI_GREEN_TEXT, width=54, height=54)
        empty_badge.pack(side="left", padx=14, pady=14)
        empty_badge.pack_propagate(False)
        tk.Label(empty_badge, text="LOG", bg=UI_GREEN_TEXT, fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(expand=True)
        empty_copy = tk.Frame(self.history_empty_frame, bg=UI_SURFACE_ALT)
        empty_copy.pack(side="left", fill="both", expand=True, padx=(0, 12), pady=14)
        tk.Label(empty_copy, text="No history yet", bg=UI_SURFACE_ALT, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(
            empty_copy,
            text="Run a batch conversion or unconvert batch and this page will keep a local review trail.",
            bg=UI_SURFACE_ALT,
            fg=UI_MUTED,
            wraplength=720,
            justify="left",
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(5, 0))

        table_frame = tk.Frame(panel, bg="#10141b")
        self.history_table_frame = table_frame
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        columns = ("Date/Time", "Action", "Input", "Valid", "Invalid", "Warnings")
        self.history_tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        history_anchors = {"Date/Time": "w", "Action": "w", "Input": "center", "Valid": "center", "Invalid": "center", "Warnings": "center"}
        for col in columns:
            self.history_tree.heading(col, text=col, anchor=history_anchors[col])
        widths = {"Date/Time": 230, "Action": 220, "Input": 90, "Valid": 90, "Invalid": 90, "Warnings": 110}
        for col, width in widths.items():
            self.history_tree.column(col, width=width, minwidth=70, anchor=history_anchors[col], stretch=col in {"Date/Time", "Action"})
        history_scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.history_tree.yview, style="Dark.Vertical.TScrollbar")
        history_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.history_tree.xview, style="Dark.Horizontal.TScrollbar")
        self.history_tree.configure(yscrollcommand=history_scroll_y.set, xscrollcommand=history_scroll_x.set)
        self.history_tree.grid(row=0, column=0, sticky="nsew")
        history_scroll_y.grid(row=0, column=1, sticky="ns")
        history_scroll_x.grid(row=1, column=0, sticky="ew")
        self.history_tree.tag_configure("history_even", background="#ffffff", foreground=UI_TEXT)
        self.history_tree.tag_configure("history_odd", background="#f8fafc", foreground=UI_TEXT)
        self.history_tree.tag_configure("invalid", background=UI_BAD_SOFT, foreground=UI_BAD_TEXT)
        self.history_tree.tag_configure("warning", background=UI_WARN_SOFT, foreground=UI_WARN_TEXT)
        self.history_tree.tag_configure("empty", foreground=UI_MUTED)
        self._enable_mousewheel(self.history_tree, self.history_tree, self.history_tree)
        self._enable_tree_row_hover(self.history_tree)
        self._enable_tree_header_tooltips(
            self.history_tree,
            {
                "Date/Time": "When the conversion run happened.",
                "Action": "The workspace used for the run.",
                "Input": "Total input rows reviewed.",
                "Valid": "Rows that converted successfully.",
                "Invalid": "Rows that need review.",
                "Warnings": "Rows that converted but had unusual or duplicate values.",
            },
        )
        self.render_history()

    def _mode_changed(self, _event: Any) -> None:
        self.select_tab(["Batch Converter", "Single Hex Lookup", "FC/CN to Hex", "Unconvert Batch", "History"].index(self.mode_var.get()))

    def _tab_changed(self, _event: Any) -> None:
        return

    def select_tab(self, index: int) -> None:
        labels = ["Batch Converter", "Single Hex Lookup", "FC/CN to Hex", "Unconvert Batch", "History"]
        self.active_tab_index = index
        self.tab_frames[index].tkraise()
        self.mode_var.set(labels[index])
        self.update_sidebar_focus(index)
        for i, button in enumerate(self.nav_buttons):
            accent, soft = self.nav_button_themes[i] if i < len(self.nav_button_themes) else self._tab_theme(i)
            if i == index:
                button.configure(bg=soft, fg=accent, highlightbackground=accent)
            else:
                button.configure(bg=UI_SURFACE, fg=UI_TEXT, highlightbackground=UI_BORDER)
        if index == 4:
            self.render_history()
        self.after(120, lambda idx=index: self.focus_primary_control(idx))

    def focus_primary_control(self, index: int | None = None) -> None:
        target = getattr(self, "active_tab_index", 0) if index is None else index
        try:
            if target == 0 and hasattr(self, "batch_scan_entry"):
                self.batch_scan_entry.focus_set()
            elif target == 1 and hasattr(self, "single_entry"):
                self.single_entry.focus_set()
            elif target == 2 and hasattr(self, "fc_entry"):
                self.fc_entry.focus_set()
            elif target == 3 and hasattr(self, "unconvert_text"):
                self.unconvert_text.focus_set()
            elif target == 4 and hasattr(self, "history_tree"):
                self.history_tree.focus_set()
        except tk.TclError:
            pass

    def set_status(self, message: str, state: str | None = None, target_tab: int | None = None) -> None:
        if target_tab is not None:
            self.status_target_tab = target_tab
        else:
            self.status_target_tab = self._infer_status_target_tab(message)
        self.status_var.set(message)
        label = self._status_state_from_message(message, state)
        focus_target = self.status_target_tab if label == "Needs Review" else getattr(self, "active_tab_index", 0)
        self.update_sidebar_focus(focus_target, label)
        if not hasattr(self, "status_state_var"):
            return
        self.status_state_var.set(label.upper())
        bg, fg = self._status_state_colors(label)
        self.status_state_chip.configure(bg=bg, fg=fg, highlightbackground=fg if label != "Ready" else UI_BORDER)

    def _infer_status_target_tab(self, message: str) -> int:
        text = str(message or "").lower()
        if "single" in text:
            return 1
        if "unconvert" in text or "fc/cn pair" in text or "fc/cn pairs" in text:
            return 3
        if "hex value created" in text or "fc/cn to hex" in text or "facility code" in text:
            return 2
        if "history" in text:
            return 4
        return getattr(self, "active_tab_index", 0)

    def open_status_target(self, _event: Any | None = None) -> str:
        target = max(0, min(4, int(getattr(self, "status_target_tab", getattr(self, "active_tab_index", 0)))))
        self.select_tab(target)

        def focus_target() -> None:
            if target == 0:
                if (self.results or self.invalid) and hasattr(self, "tree"):
                    self.tree.focus_set()
                elif hasattr(self, "batch_scan_entry"):
                    self.batch_scan_entry.focus_set()
            elif target == 1 and hasattr(self, "single_entry"):
                self.single_entry.focus_set()
            elif target == 2 and hasattr(self, "fc_entry"):
                self.fc_entry.focus_set()
            elif target == 3:
                if (self.unconvert_results or self.unconvert_invalid) and hasattr(self, "unconvert_tree"):
                    self.unconvert_tree.focus_set()
                elif hasattr(self, "unconvert_text"):
                    self.unconvert_text.focus_set()
            elif target == 4 and hasattr(self, "history_tree"):
                self.history_tree.focus_set()

        self.after(180, focus_target)
        return "break"

    def set_busy(self, message: str) -> None:
        if not hasattr(self, "busy_label"):
            return
        self.busy_var.set(message)
        if not self.busy_label.winfo_ismapped():
            self.busy_label.pack(side="left", padx=(0, 8), before=self.footer_status_label)
        self.configure(cursor="watch")
        self.update_idletasks()

    def clear_busy(self) -> None:
        if not hasattr(self, "busy_label"):
            return
        self.busy_var.set("")
        if self.busy_label.winfo_ismapped():
            self.busy_label.pack_forget()
        self.configure(cursor="")
        self.update_idletasks()

    def _status_state_from_message(self, message: str, state: str | None = None) -> str:
        if state in {"Ready", "Needs Review", "Exported"}:
            return state
        text = message.lower()
        run_complete = re.match(
            r"^(?:converted|unconverted)\s+\d+\s+valid\s+(?:line|pair)\(s\);\s+"
            r"(\d+)\s+invalid\s+line\(s\)(?:;\s+(\d+)\s+warning\(s\))?\.",
            text,
        )
        if run_complete:
            invalid_count = int(run_complete.group(1))
            warning_count = int(run_complete.group(2) or 0)
            return "Needs Review" if invalid_count or warning_count else "Ready"
        if text.startswith(("cleared", "removed", "kept", "workspace", "settings", "sample", "clipboard", "opened", "copied")):
            return "Ready"
        if any(word in text for word in ("failed", "error", "invalid", "warning", "review", "unavailable")):
            return "Needs Review"
        if any(word in text for word in ("saved", "exported", "report")):
            return "Exported"
        return "Ready"

    def _status_state_colors(self, state: str) -> tuple[str, str]:
        if state == "Needs Review":
            return UI_WARN_SOFT, UI_WARN_TEXT
        if state == "Exported":
            return UI_BLUE, "#ffffff"
        return UI_GREEN_SOFT, UI_GREEN_TEXT

    def _enable_drop_target(self, widget: tk.Widget) -> None:
        if DND_FILES is None or not hasattr(widget, "drop_target_register"):
            return
        try:
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", self.handle_file_drop)
            widget.dnd_bind("<<DragEnter>>", self.handle_drag_enter)
            widget.dnd_bind("<<DragLeave>>", self.handle_drag_leave)
        except Exception:
            return

    def handle_drag_enter(self, event: Any) -> str | None:
        target = self.unconvert_text if getattr(self, "active_tab_index", 0) == 3 and hasattr(self, "unconvert_text") else self.multi_text if hasattr(self, "multi_text") else None
        if target is not None:
            target.configure(bg="#eef6ff")
        self.set_status("Drop supported files onto this batch input area to import them.")
        return getattr(event, "action", None)

    def handle_drag_leave(self, event: Any) -> str | None:
        for name in ("multi_text", "unconvert_text"):
            if hasattr(self, name):
                getattr(self, name).configure(bg=UI_INPUT)
        return getattr(event, "action", None)

    def handle_file_drop(self, event: Any) -> str | None:
        for name in ("multi_text", "unconvert_text"):
            if hasattr(self, name):
                getattr(self, name).configure(bg=UI_INPUT)
        try:
            paths = [Path(item) for item in self.tk.splitlist(event.data)]
        except Exception:
            paths = [Path(str(event.data))]
        self.import_paths_to_queue(paths, preview=False)
        return getattr(event, "action", None)

    def append_text_to_queue(self, text: str) -> None:
        incoming = text.strip()
        if not incoming:
            return
        current = self.multi_text.get("1.0", "end").strip()
        self.multi_text.delete("1.0", "end")
        self.multi_text.insert("1.0", f"{current}\n{incoming}".strip() if current else incoming)
        self.multi_text.edit_modified(True)
        self.handle_batch_input_changed()

    def prepend_text_to_queue(self, text: str) -> None:
        incoming = text.strip()
        if not incoming:
            return
        current = self.multi_text.get("1.0", "end").strip()
        self.multi_text.delete("1.0", "end")
        self.multi_text.insert("1.0", f"{incoming}\n{current}".strip() if current else incoming)
        self.multi_text.edit_modified(True)
        self.handle_batch_input_changed()

    def handle_batch_scan_submit(self, _event: Any | None = None) -> str:
        if self._batch_scan_after:
            self.after_cancel(self._batch_scan_after)
            self._batch_scan_after = None
        text = self.batch_scan_var.get() if hasattr(self, "batch_scan_var") else ""
        cleanup = clean_id_lines_from_text(text, numeric_only=True)
        if not cleanup["lines"]:
            self.set_status("Scanner input needs review: no clean 8-character ID found.", "Needs Review", target_tab=0)
            return "break"
        self.prepend_text_to_queue("\n".join(cleanup["lines"]))
        if hasattr(self, "batch_scan_count_var"):
            self.batch_scan_count += len(cleanup["lines"])
            scan_label = "scan" if self.batch_scan_count == 1 else "scans"
            self.batch_scan_count_var.set(f"{self.batch_scan_count} {scan_label} this session")
        self.batch_scan_var.set("")
        if hasattr(self, "batch_scan_entry"):
            self.batch_scan_entry.focus_set()
        label = "ID" if len(cleanup["lines"]) == 1 else "IDs"
        self.notice_var.set(f"Scanner added {len(cleanup['lines'])} clean {label} to the top of the queue.")
        self.set_status(f"Scanner added {len(cleanup['lines'])} clean {label} to Batch Converter.", "Ready", target_tab=0)
        return "break"

    def schedule_batch_scan_autofill(self, event: Any | None = None) -> None:
        key = getattr(event, "keysym", "")
        if key in {"Return", "KP_Enter", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R", "Alt_L", "Alt_R"}:
            return
        if self._batch_scan_after:
            self.after_cancel(self._batch_scan_after)
        self._batch_scan_after = self.after(260, self.autofill_batch_scan_if_ready)

    def autofill_batch_scan_if_ready(self) -> None:
        self._batch_scan_after = None
        text = self.batch_scan_var.get().strip() if hasattr(self, "batch_scan_var") else ""
        if not text:
            return
        cleanup = clean_id_lines_from_text(text, numeric_only=True)
        if cleanup["lines"]:
            self.handle_batch_scan_submit()

    def append_text_to_unconvert(self, text: str) -> None:
        incoming = text.strip()
        if not incoming:
            return
        current = self.unconvert_text.get("1.0", "end").strip()
        self.unconvert_text.delete("1.0", "end")
        self.unconvert_text.insert("1.0", f"{current}\n{incoming}".strip() if current else incoming)
        self.unconvert_text.edit_modified(True)
        self.handle_unconvert_input_changed()

    def paste_clipboard_to_queue(self, clean_only: bool = False) -> None:
        if getattr(self, "active_tab_index", 0) == 3:
            self.paste_clipboard_to_unconvert()
            return
        if getattr(self, "active_tab_index", 0) not in {0, 3}:
            messagebox.showinfo("Batch import only", "Import and paste are available on Batch Converter and Unconvert Batch.")
            return
        try:
            text = self.clipboard_get()
        except tk.TclError:
            messagebox.showinfo("Clipboard empty", "There is no text on the clipboard to paste.")
            return
        if not text.strip():
            messagebox.showinfo("Clipboard empty", "There is no text on the clipboard to paste.")
            return
        cleanup = clean_id_lines_from_text(text, numeric_only=True)
        if clean_only:
            if cleanup["lines"]:
                self.append_text_to_queue("\n".join(cleanup["lines"]))
                self.set_status(f"Added {len(cleanup['lines'])} clean ID row(s) from clipboard.")
                return
            messagebox.showinfo("No IDs found", "No clean 8-character IDs were detected on the clipboard.")
            self.set_status("Clipboard did not contain clean IDs.", "Needs Review")
            return
        if cleanup["changed"] and cleanup["lines"]:
            action = self.confirm_paste_cleanup_preview(text, cleanup)
            if action == "cancel":
                self.set_status("Clipboard paste canceled.")
                return
            if action == "clean":
                self.append_text_to_queue("\n".join(cleanup["lines"]))
                self.set_status(f"Added {len(cleanup['lines'])} cleaned ID row(s) from clipboard.")
                return
        self.append_text_to_queue(text)
        self.set_status("Clipboard text added to the Input Queue.")

    def paste_clipboard_to_unconvert(self) -> None:
        if getattr(self, "active_tab_index", 0) not in {3}:
            self.select_tab(3)
        try:
            text = self.clipboard_get()
        except tk.TclError:
            messagebox.showinfo("Clipboard empty", "There is no text on the clipboard to paste.")
            return
        if not text.strip():
            messagebox.showinfo("Clipboard empty", "There is no text on the clipboard to paste.")
            return
        cleanup = clean_fc_cn_lines_from_text(text)
        if cleanup["lines"]:
            self.append_text_to_unconvert("\n".join(cleanup["lines"]))
            self.set_status(f"Added {len(cleanup['lines'])} FC/CN pair(s) from clipboard.")
            return
        messagebox.showinfo("No FC/CN pairs found", "No Facility Code/Card Number pairs were detected on the clipboard.")
        self.set_status("Clipboard did not contain FC/CN pairs.", "Needs Review")

    def _count_nonempty_text_lines(self, widget: tk.Text) -> int:
        text = widget.get("1.0", "end").strip()
        return len([line for line in text.splitlines() if line.strip()]) if text else 0

    def update_queue_count(self, queued: int | None = None) -> None:
        if not hasattr(self, "queue_count_var") or not hasattr(self, "multi_text"):
            return
        count = self._count_nonempty_text_lines(self.multi_text) if queued is None else queued
        label = "row" if count == 1 else "rows"
        self.queue_count_var.set(f"{count} {label}")

    def handle_batch_input_changed(self, _event: Any | None = None) -> None:
        if not hasattr(self, "multi_text") or not self.multi_text.edit_modified():
            return
        self.multi_text.edit_modified(False)
        queued = self._count_nonempty_text_lines(self.multi_text)
        self.update_queue_count(queued)
        self.highlight_input_queue()
        if hasattr(self, "stat_vars"):
            self.stat_vars["Input"].set(str(queued))
        if not hasattr(self, "time_var"):
            return
        if queued == 0:
            if not self.results and not self.invalid:
                self.time_var.set("No run yet")
                self.notice_var.set("")
            return
        if self.results or self.invalid:
            self.time_var.set(f"{queued} line(s) in queue")
            self.notice_var.set("Input changed. Run Convert to refresh results.")
        else:
            self.time_var.set(f"{queued} line(s) queued")
            self.notice_var.set("")

    def highlight_input_queue(self) -> None:
        if not hasattr(self, "multi_text"):
            return
        for tag in ("input_valid", "input_warning", "input_invalid"):
            self.multi_text.tag_remove(tag, "1.0", "end")
        lines = self.multi_text.get("1.0", "end-1c").splitlines()
        prepared_rows: list[tuple[str, dict[str, Any], str]] = []
        hex_counts: dict[str, int] = {}
        for line in lines:
            prepared = clean_candidate_line(line)
            hex_value = prepared["extracted"].strip().upper()
            prepared_rows.append((line, prepared, hex_value))
            if valid_hex8(hex_value):
                hex_counts[hex_value] = hex_counts.get(hex_value, 0) + 1
        for idx, (line, prepared, hex_value) in enumerate(prepared_rows, start=1):
            if not line.strip():
                continue
            tag = "input_invalid"
            if valid_hex8(hex_value):
                facility, card = hex_to_fc_cn(hex_value)
                has_warning = bool(prepared["suggestions"] or unusual_warnings(hex_value, facility, card) or hex_counts.get(hex_value, 0) > 1)
                tag = "input_warning" if has_warning else "input_valid"
            self.multi_text.tag_add(tag, f"{idx}.0", f"{idx}.end")
        self.update_queue_count(len([line for line in lines if line.strip()]))

    def remove_duplicate_input_lines(self) -> None:
        if not hasattr(self, "multi_text"):
            return
        lines = self.multi_text.get("1.0", "end-1c").splitlines()
        kept: list[str] = []
        seen: set[str] = set()
        removed = 0
        for line in lines:
            if not line.strip():
                continue
            prepared = clean_candidate_line(line)
            hex_value = prepared["extracted"].strip().upper()
            if valid_hex8(hex_value):
                if hex_value in seen:
                    removed += 1
                    continue
                seen.add(hex_value)
            kept.append(line)
        self.multi_text.delete("1.0", "end")
        self.multi_text.insert("1.0", "\n".join(kept))
        self.multi_text.edit_modified(True)
        self.handle_batch_input_changed()
        self.set_status(f"Removed {removed} duplicate row(s)." if removed else "No duplicate HEX IDs found.")

    def keep_only_valid_input_lines(self) -> None:
        if not hasattr(self, "multi_text"):
            return
        lines = self.multi_text.get("1.0", "end-1c").splitlines()
        kept: list[str] = []
        removed = 0
        for line in lines:
            if not line.strip():
                continue
            prepared = clean_candidate_line(line)
            if valid_hex8(prepared["extracted"].strip().upper()):
                kept.append(line)
            else:
                removed += 1
        self.multi_text.delete("1.0", "end")
        self.multi_text.insert("1.0", "\n".join(kept))
        self.multi_text.edit_modified(True)
        self.handle_batch_input_changed()
        self.set_status(f"Kept {len(kept)} valid row(s); removed {removed} invalid row(s).")

    def handle_unconvert_input_changed(self, _event: Any | None = None) -> None:
        if not hasattr(self, "unconvert_text") or not self.unconvert_text.edit_modified():
            return
        self.unconvert_text.edit_modified(False)
        queued = self._count_nonempty_text_lines(self.unconvert_text)
        if not hasattr(self, "unconvert_summary_var"):
            return
        if queued:
            self.unconvert_summary_var.set(f"{queued} line(s) queued")
        elif not self.unconvert_results and not self.unconvert_invalid:
            self.unconvert_summary_var.set("No Unconvert run yet")

    def load_sample(self) -> None:
        self.multi_text.delete("1.0", "end")
        self.multi_text.insert(
            "1.0",
            "\n".join(
                [
                    "88984717",
                    "88984765",
                    "88984130",
                    "88981234",
                    "88984717",
                ]
            ),
        )
        self.multi_text.edit_modified(True)
        self.handle_batch_input_changed()
        self.set_status("Sample IDs loaded.")

    def load_unconvert_sample(self) -> None:
        self.unconvert_text.delete("1.0", "end")
        self.unconvert_text.insert(
            "1.0",
            "\n".join(
                [
                    "34968,18199",
                    "34968,18277",
                    "34968,16700",
                    "34968,4660",
                ]
            ),
        )
        self.unconvert_text.edit_modified(True)
        self.handle_unconvert_input_changed()
        self.set_status("Unconvert sample loaded.")

    def clear_workspace(self) -> None:
        self.multi_text.delete("1.0", "end")
        self.highlight_input_queue()
        self.results.clear()
        self.invalid.clear()
        self.unconvert_results.clear()
        self.unconvert_invalid.clear()
        self.last_converted_at = ""
        self.last_unconverted_at = ""
        self.notice_var.set("")
        self.time_var.set("No run yet")
        if hasattr(self, "batch_scan_var"):
            self.batch_scan_var.set("")
        if hasattr(self, "batch_scan_count_var"):
            self.batch_scan_count = 0
            self.batch_scan_count_var.set("0 scans this session")
        self.single_var.set("")
        self.single_result.set("Waiting for one 8-character HEX ID.")
        if hasattr(self, "single_hex_var"):
            self.single_hex_var.set("--")
            self.single_fc_var.set("--")
            self.single_cn_var.set("--")
            self.single_note_var.set("No lookup has run yet.")
        self.reverse_result.set("Waiting for Facility Code and Card Number.")
        if hasattr(self, "reverse_hex_var"):
            self.reverse_hex_var.set("--")
            self.reverse_note_var.set("No reverse lookup has run yet.")
        self.fc_var.set("")
        self.cn_var.set("")
        if hasattr(self, "unconvert_text"):
            self.unconvert_text.delete("1.0", "end")
            self.unconvert_summary_var.set("No Unconvert run yet")
            self.render_unconvert_results()
        self.render_results()
        self.set_status("Workspace cleared.")

    def convert_batch(self) -> None:
        converted_at = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
        self.results, self.invalid = convert_lines(self.multi_text.get("1.0", "end"), converted_at)
        self.last_converted_at = converted_at
        self.render_results()
        total = len(self.results) + len(self.invalid)
        warn_count = sum(len(row.warnings) for row in self.results)
        if total == 0:
            self.set_status("No input: add at least one ID before converting.", "Needs Review", target_tab=0)
            messagebox.showinfo("No input", "Add at least one ID before converting.")
            return
        self.add_history("HEX to FC/CN", total, len(self.results), len(self.invalid), warn_count)
        state = "Needs Review" if self.invalid or warn_count else "Ready"
        self.set_status(
            f"Converted {len(self.results)} valid line(s); {len(self.invalid)} invalid line(s); {warn_count} warning(s).",
            state,
            target_tab=0,
        )

    def convert_unconvert_batch(self) -> None:
        converted_at = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
        self.unconvert_results, self.unconvert_invalid = unconvert_lines(self.unconvert_text.get("1.0", "end"), converted_at)
        self.last_unconverted_at = converted_at
        self.render_unconvert_results()
        total = len(self.unconvert_results) + len(self.unconvert_invalid)
        warn_count = sum(len(row.warnings) for row in self.unconvert_results)
        if total == 0:
            self.set_status("No input: add at least one FC/CN pair before running Unconvert.", "Needs Review", target_tab=3)
            messagebox.showinfo("No input", "Add at least one FC/CN pair before running Unconvert.")
            return
        self.add_history("FC/CN to HEX", total, len(self.unconvert_results), len(self.unconvert_invalid), warn_count)
        state = "Needs Review" if self.unconvert_invalid or warn_count else "Ready"
        self.set_status(
            f"Unconverted {len(self.unconvert_results)} valid pair(s); {len(self.unconvert_invalid)} invalid line(s); {warn_count} warning(s).",
            state,
            target_tab=3,
        )

    def clear_unconvert(self) -> None:
        self.unconvert_text.delete("1.0", "end")
        self.unconvert_results.clear()
        self.unconvert_invalid.clear()
        self.last_unconverted_at = ""
        self.unconvert_summary_var.set("No Unconvert run yet")
        self.render_unconvert_results()
        self.set_status("Unconvert workspace cleared.")

    def sort_results_by(self, column: str) -> None:
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        self.render_results()

    def render_results(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.row_lookup.clear()
        combined: list[tuple[int, str, ConvertedRow | InvalidRow]] = [(row.line, "valid", row) for row in self.results] + [(row.line, "invalid", row) for row in self.invalid]
        query = self.search_var.get().strip().lower() if hasattr(self, "search_var") else ""
        status_filter = self.status_filter_var.get() if hasattr(self, "status_filter_var") else "All"
        filtered: list[tuple[int, str, ConvertedRow | InvalidRow]] = []
        if hasattr(self, "results_empty_frame"):
            if combined:
                self.results_empty_frame.pack_forget()
            elif not self.results_empty_frame.winfo_ismapped():
                self.results_empty_frame.pack(fill="x", padx=12, pady=(0, 8), before=self.results_table_frame)
        for line, kind, row in combined:
            values = [str(line)]
            if isinstance(row, InvalidRow):
                status = "Invalid"
                values.extend([row.raw, row.reason])
            else:
                status = "Warning" if row.warnings else "Valid"
                values.extend([row.hex_value, str(row.facility), str(row.card), " ".join([*row.suggestions, *row.warnings])])
            if status_filter != "All" and status != status_filter:
                continue
            if query and query not in " ".join(values).lower():
                continue
            filtered.append((line, kind, row))
        def sort_key(item: tuple[int, str, ConvertedRow | InvalidRow]) -> Any:
            _line, kind, row = item
            if self.sort_column == "Line":
                return row.line
            if isinstance(row, InvalidRow):
                mapping = {"Hex ID": row.raw, "FC": "", "CN": "", "Status": "Invalid", "Notes": row.reason}
            else:
                mapping = {"Hex ID": row.hex_value, "FC": row.facility, "CN": row.card, "Status": "Warning" if row.warnings else "Valid", "Notes": " ".join([*row.suggestions, *row.warnings])}
            return mapping.get(self.sort_column, row.line)
        filtered.sort(key=sort_key, reverse=self.sort_reverse)
        if not filtered and combined:
            self.tree.insert("", "end", values=("", "No matching results.", "", "", "", ""), tags=("empty",))
        for _line, kind, row in filtered:
            if kind == "invalid":
                item_id = self.tree.insert("", "end", values=(row.line, row.raw, "", "", "Invalid", row.reason), tags=("invalid",))
            else:
                notes = " | ".join([*row.suggestions, *row.warnings])
                status = "Warning" if row.warnings else "Valid"
                tags = ("warning",) if row.warnings else ("valid",)
                item_id = self.tree.insert("", "end", values=(row.line, row.hex_value, row.facility, row.card, status, notes), tags=tags)
            self.row_lookup[item_id] = row
        warn_count = sum(len(row.warnings) for row in self.results)
        self.stat_vars["Input"].set(str(len(self.results) + len(self.invalid)))
        self.stat_vars["Valid"].set(str(len(self.results)))
        self.stat_vars["Invalid"].set(str(len(self.invalid)))
        self.stat_vars["Warnings"].set(str(warn_count))
        self.time_var.set(f"Converted {self.last_converted_at}" if self.last_converted_at else "No run yet")
        self.notice_var.set(f"{warn_count} warning(s) | {len(self.invalid)} invalid line(s)" if self.results or self.invalid else "")

    def render_unconvert_results(self) -> None:
        for item in self.unconvert_tree.get_children():
            self.unconvert_tree.delete(item)
        self.unconvert_row_lookup.clear()
        combined: list[tuple[int, str, UnconvertRow | InvalidRow]] = [(row.line, "valid", row) for row in self.unconvert_results] + [(row.line, "invalid", row) for row in self.unconvert_invalid]
        if hasattr(self, "unconvert_empty_frame"):
            if combined:
                self.unconvert_empty_frame.pack_forget()
            elif not self.unconvert_empty_frame.winfo_ismapped():
                self.unconvert_empty_frame.pack(fill="x", padx=12, pady=(0, 8), before=self.unconvert_table_frame)
        if not combined:
            return
        for _line, kind, row in sorted(combined, key=lambda item: item[0]):
            if kind == "invalid":
                item_id = self.unconvert_tree.insert("", "end", values=(row.line, "", "", row.raw, "Invalid", row.reason), tags=("invalid",))
            else:
                notes = " | ".join(row.warnings)
                status = "Warning" if row.warnings else "Valid"
                tags = ("warning",) if row.warnings else ("valid",)
                item_id = self.unconvert_tree.insert("", "end", values=(row.line, row.facility, row.card, row.hex_value, status, notes), tags=tags)
            self.unconvert_row_lookup[item_id] = row
        warn_count = sum(len(row.warnings) for row in self.unconvert_results)
        total = len(self.unconvert_results) + len(self.unconvert_invalid)
        self.unconvert_summary_var.set(f"Input: {total}\nValid: {len(self.unconvert_results)}\nInvalid: {len(self.unconvert_invalid)}\nWarnings: {warn_count}")

    def copy_selected(self, mode: str) -> None:
        selected = self.tree.selection()
        if not selected:
            self.set_status("Select a result row first.")
            return
        row = self.row_lookup.get(selected[0])
        if row is None:
            self.set_status("Select a converted result row first.")
            return
        status_message = "Copied to clipboard."
        if isinstance(row, InvalidRow):
            values = [str(row.line), row.raw, "", "", "Invalid", row.reason]
            if mode == "row":
                text = "\t".join(values)
                status_message = "Copied full invalid result row."
            elif mode == "notes":
                text = row.reason
                status_message = "Copied row notes."
            else:
                text = row.raw
                status_message = "Copied raw invalid input."
        elif mode == "fc":
            text = str(row.facility)
            status_message = "Copied Facility Code."
        elif mode == "cn":
            text = str(row.card)
            status_message = "Copied Card Number."
        elif mode == "hex":
            text = row.hex_value
            status_message = "Copied HEX ID."
        elif mode == "notes":
            text = " | ".join([*row.suggestions, *row.warnings])
            if not text:
                self.set_status("Selected row has no notes.")
                return
            status_message = "Copied row notes."
        elif mode == "row":
            notes = " | ".join([*row.suggestions, *row.warnings])
            status = "Warning" if row.warnings else "Valid"
            text = "\t".join([str(row.line), row.hex_value, str(row.facility), str(row.card), status, notes])
            status_message = "Copied full result row."
        else:
            text = f"{row.facility},{row.card}"
            status_message = "Copied FC,CN pair."
        self.clipboard_clear()
        self.clipboard_append(text)
        self.set_status(status_message)

    def copy_all_pairs(self) -> None:
        if not self.results:
            self.set_status("No valid FC/CN pairs to copy.")
            return
        text = "\n".join(f"{row.facility},{row.card}" for row in self.results)
        self.clipboard_clear()
        self.clipboard_append(text)
        self.set_status(f"Copied {len(self.results)} FC/CN pair(s).")

    def clear_invalid_rows(self) -> None:
        if not self.invalid:
            self.set_status("No invalid rows to clear.")
            return
        count = len(self.invalid)
        self.invalid.clear()
        self.render_results()
        self.set_status(f"Cleared {count} invalid row(s).")

    def copy_selected_unconvert_hex(self) -> None:
        self.copy_selected_unconvert("hex", require_valid=True)

    def copy_selected_unconvert(self, mode: str, require_valid: bool = False) -> None:
        selected = self.unconvert_tree.selection()
        if not selected:
            self.set_status("Select an unconvert result first.")
            return
        row = self.unconvert_row_lookup.get(selected[0])
        if row is None:
            self.set_status("Select an unconvert result first.")
            return
        if require_valid and not isinstance(row, UnconvertRow):
            self.set_status("Select a valid unconvert result first.")
            return
        status_message = "Copied to clipboard."
        if isinstance(row, InvalidRow):
            values = [str(row.line), "", "", row.raw, "Invalid", row.reason]
            if mode == "row":
                text = "\t".join(values)
                status_message = "Copied full invalid unconvert row."
            elif mode == "notes":
                text = row.reason
                status_message = "Copied row notes."
            else:
                text = row.raw
                status_message = "Copied raw invalid input."
        elif mode == "row":
            notes = " | ".join(row.warnings)
            status = "Warning" if row.warnings else "Valid"
            text = "\t".join([str(row.line), str(row.facility), str(row.card), row.hex_value, status, notes])
            status_message = "Copied full unconvert row."
        elif mode == "fc":
            text = str(row.facility)
            status_message = "Copied Facility Code."
        elif mode == "cn":
            text = str(row.card)
            status_message = "Copied Card Number."
        elif mode == "notes":
            text = " | ".join(row.warnings)
            if not text:
                self.set_status("Selected unconvert row has no notes.")
                return
            status_message = "Copied row notes."
        else:
            text = row.hex_value
            status_message = "Copied HEX to clipboard."
        self.clipboard_clear()
        self.clipboard_append(text)
        self.set_status(status_message)

    def copy_all_unconverted_hex(self) -> None:
        if not self.unconvert_results:
            self.set_status("No unconverted HEX values to copy.")
            return
        self.clipboard_clear()
        self.clipboard_append("\n".join(row.hex_value for row in self.unconvert_results))
        self.set_status(f"Copied {len(self.unconvert_results)} HEX value(s).")

    def choose_export_folder(self) -> None:
        folder = filedialog.askdirectory(
            title="Choose default export folder",
            initialdir=self.settings.get("default_export_dir") or str(Path.home() / "Desktop"),
        )
        if not folder:
            return
        self.settings["default_export_dir"] = folder
        self.save_settings()
        self.set_status(f"Default export folder set: {folder}")

    def open_export_folder(self) -> None:
        folder = Path(self.default_export_dir())
        folder.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(folder)  # type: ignore[attr-defined]
            self.set_status(f"Opened export folder: {folder}")
        except OSError as exc:
            messagebox.showerror("Folder unavailable", f"Could not open the export folder.\n\n{exc}")

    def default_export_dir(self) -> str:
        folder = Path(self.settings.get("default_export_dir") or Path.home() / "Desktop")
        return str(folder if folder.exists() else Path.home())

    def show_recent_exports(self) -> None:
        dialog = self._new_dialog("Recent Exports")
        dialog.geometry("720x460")
        dialog.minsize(640, 380)

        self._dialog_header(dialog, "Recent Exports", "Reopen recently saved reports from this workstation.", UI_BLUE)

        body = tk.Frame(dialog, bg=UI_BG)
        body.pack(fill="both", expand=True, padx=18, pady=18)

        recent = [
            entry
            for entry in self.settings.get("recent_exports", [])
            if entry.get("path") and Path(entry.get("path", "")).exists()
        ]
        if len(recent) != len(self.settings.get("recent_exports", [])):
            self.settings["recent_exports"] = recent
            self.save_settings()

        if not recent:
            empty = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
            empty.pack(fill="both", expand=True)
            tk.Frame(empty, bg=UI_RED, height=3).pack(fill="x")
            inner = tk.Frame(empty, bg=UI_SURFACE)
            inner.pack(fill="both", expand=True, padx=18, pady=18)
            tk.Label(inner, text="No saved exports yet", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 13, "bold")).pack(anchor="w")
            tk.Label(inner, text="After you export a report, it will appear here for quick reopen access.", bg=UI_SURFACE, fg=UI_MUTED, wraplength=620, justify="left").pack(anchor="w", pady=(6, 0))
        else:
            canvas = tk.Canvas(body, bg=UI_BG, highlightthickness=0)
            scrollbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview, style="Dark.Vertical.TScrollbar")
            list_frame = tk.Frame(canvas, bg=UI_BG)
            list_id = canvas.create_window((0, 0), window=list_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            list_frame.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.bind("<Configure>", lambda event: canvas.itemconfigure(list_id, width=event.width))
            self._enable_mousewheel(canvas, canvas)
            self._enable_mousewheel_tree(list_frame, canvas)
            for entry in recent:
                path = Path(entry.get("path", ""))
                card = self._card(list_frame, bg=UI_SURFACE, border=UI_BORDER)
                card.pack(fill="x", pady=(0, 10))
                tk.Frame(card, bg=UI_RED if entry.get("type") != "PDF" else UI_BLUE, height=3).pack(fill="x")
                row = tk.Frame(card, bg=UI_SURFACE)
                row.pack(fill="x", padx=12, pady=10)
                copy = tk.Frame(row, bg=UI_SURFACE)
                copy.pack(side="left", fill="x", expand=True)
                tk.Label(copy, text=entry.get("name", path.name), bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w")
                tk.Label(copy, text=f"{entry.get('type', 'Report')} | {entry.get('timestamp', '')}", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))
                tk.Label(copy, text=str(path), bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 8), wraplength=440, justify="left").pack(anchor="w", pady=(3, 0))
                self._button(row, "Open", lambda p=path: self.open_path(p, "export"), True, icon="icon-status.png").pack(side="right", padx=(8, 0))
                self._button(row, "Folder", lambda p=path: self.open_path(p.parent, "folder"), icon="icon-folder.png").pack(side="right")

        self._dialog_footer_accent(dialog)
        row = tk.Frame(dialog, bg=UI_BG)
        row.pack(fill="x", padx=18, pady=(0, 18))
        self._button(row, "Close", dialog.destroy, True).pack(side="right")

    def show_settings(self) -> None:
        dialog = self._new_dialog("Settings")
        dialog.geometry("720x700")
        dialog.minsize(640, 620)

        self._dialog_header(dialog, "Settings", "Defaults for exports and workstation shortcuts.", UI_RED)

        body_shell = tk.Frame(dialog, bg=UI_BG)
        body_shell.pack(fill="both", expand=True, padx=18, pady=18)
        body_shell.configure(height=500)
        body_shell.pack_propagate(False)
        body_shell.rowconfigure(0, weight=1)
        body_shell.columnconfigure(0, weight=1)
        settings_canvas = tk.Canvas(body_shell, bg=UI_BG, highlightthickness=0)
        settings_scroll = ttk.Scrollbar(body_shell, orient="vertical", command=settings_canvas.yview, style="Dark.Vertical.TScrollbar")
        body = tk.Frame(settings_canvas, bg=UI_BG)
        body_id = settings_canvas.create_window((0, 0), window=body, anchor="nw")
        settings_canvas.configure(yscrollcommand=settings_scroll.set)
        settings_canvas.grid(row=0, column=0, sticky="nsew")
        settings_scroll.grid(row=0, column=1, sticky="ns")
        body.bind("<Configure>", lambda _event: settings_canvas.configure(scrollregion=settings_canvas.bbox("all")))
        settings_canvas.bind("<Configure>", lambda event: settings_canvas.itemconfigure(body_id, width=event.width))
        self._enable_mousewheel(settings_canvas, settings_canvas)
        self._enable_mousewheel_tree(body, settings_canvas)

        folder_var = tk.StringVar(value=self.default_export_dir())
        export_type_var = tk.StringVar(value=self.settings.get("default_export_type", EXPORT_TYPE_CHOICES[0]))

        export_card = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
        export_card.pack(fill="x", pady=(0, 12))
        tk.Frame(export_card, bg=UI_RED, height=3).pack(fill="x")
        export_inner = tk.Frame(export_card, bg=UI_SURFACE)
        export_inner.pack(fill="x", padx=14, pady=14)
        tk.Label(export_inner, text="Export Defaults", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(export_inner, text="These settings control where Save dialogs start and what Export Default creates.", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 9), wraplength=620, justify="left").pack(anchor="w", pady=(4, 10))

        tk.Label(export_inner, text="Default export type", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))
        type_menu_button = tk.Menubutton(
            export_inner,
            text=f"{export_type_var.get()} v",
            bg=UI_INPUT,
            fg=UI_TEXT,
            activebackground="#eef2f6",
            activeforeground=UI_TEXT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
            anchor="w",
            font=("Segoe UI", 10, "bold"),
            padx=10,
            pady=9,
            cursor="hand2",
        )
        type_menu = tk.Menu(type_menu_button, tearoff=False, bg=UI_SURFACE, fg=UI_TEXT, activebackground=UI_RED, activeforeground="#ffffff", bd=0, relief="flat")

        def choose_export_type(value: str) -> None:
            export_type_var.set(value)
            type_menu_button.configure(text=f"{value} v")

        for choice in EXPORT_TYPE_CHOICES:
            type_menu.add_command(label=choice, command=lambda value=choice: choose_export_type(value))
        type_menu_button.configure(menu=type_menu)
        type_menu_button.bind("<Enter>", lambda _event: type_menu_button.configure(bg="#eef2f6"))
        type_menu_button.bind("<Leave>", lambda _event: type_menu_button.configure(bg=UI_INPUT))
        ToolTip(type_menu_button, "Choose the report format used by Export Default.")
        type_menu_button.pack(fill="x", pady=(0, 10))

        tk.Label(export_inner, text="Default export folder", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))
        folder_row = tk.Frame(export_inner, bg=UI_SURFACE)
        folder_row.pack(fill="x")
        folder_entry = tk.Entry(folder_row, textvariable=folder_var, bg=UI_INPUT, fg=UI_TEXT, insertbackground=UI_RED, relief="flat", highlightthickness=1, highlightbackground=UI_BORDER, font=("Segoe UI", 10))
        folder_entry.pack(side="left", fill="x", expand=True, ipady=8)
        self._enable_edit_context_menu(folder_entry)

        def browse_folder() -> None:
            folder = filedialog.askdirectory(title="Choose default export folder", initialdir=folder_var.get() or self.default_export_dir())
            if folder:
                folder_var.set(folder)

        def open_folder_from_settings() -> None:
            folder = Path(folder_var.get() or self.default_export_dir())
            folder.mkdir(parents=True, exist_ok=True)
            self.open_path(folder, "folder")

        action_row = tk.Frame(export_inner, bg=UI_SURFACE)
        action_row.pack(fill="x", pady=(10, 0))
        self._button(action_row, "Browse", browse_folder, icon="icon-folder.png", tooltip="Choose the folder where export dialogs should start.").pack(side="left", padx=(0, 8))
        self._button(action_row, "Open Folder", open_folder_from_settings, icon="icon-folder.png", tooltip="Open the folder shown above.").pack(side="left")

        shortcut_card = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
        shortcut_card.pack(fill="x")
        tk.Frame(shortcut_card, bg=UI_BLUE, height=3).pack(fill="x")
        shortcut_inner = tk.Frame(shortcut_card, bg=UI_SURFACE)
        shortcut_inner.pack(fill="x", padx=14, pady=14)
        tk.Label(shortcut_inner, text="Desktop Shortcut", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(shortcut_inner, text="Create a Windows desktop shortcut for the current utility so it is easy to launch.", bg=UI_SURFACE, fg=UI_MUTED, wraplength=620, justify="left", font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 10))
        self._button(shortcut_inner, "Create Shortcut", self.create_desktop_shortcut, icon="icon-status.png", tooltip="Add a shortcut to your Windows desktop.").pack(anchor="w")

        maintenance_card = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
        maintenance_card.pack(fill="x", pady=(12, 0))
        tk.Frame(maintenance_card, bg=UI_WARN_TEXT, height=3).pack(fill="x")
        maintenance_inner = tk.Frame(maintenance_card, bg=UI_SURFACE)
        maintenance_inner.pack(fill="x", padx=14, pady=14)
        maintenance_header = tk.Frame(maintenance_inner, bg=UI_SURFACE)
        maintenance_header.pack(fill="x")
        tk.Label(maintenance_header, text="Recent Export List", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(side="left", anchor="w")
        self._button(maintenance_header, "Clear Recent Exports", self.clear_recent_exports, icon="icon-clear.png", tooltip="Clear the Recent Exports list only.").pack(side="right")
        tk.Label(
            maintenance_inner,
            text="Clear saved export shortcuts from this workstation without deleting the exported files.",
            bg=UI_SURFACE,
            fg=UI_MUTED,
            wraplength=620,
            justify="left",
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(4, 10))

        def save_and_close() -> None:
            folder = Path(folder_var.get() or self.default_export_dir())
            folder.mkdir(parents=True, exist_ok=True)
            self.settings["default_export_dir"] = str(folder)
            self.settings["default_export_type"] = export_type_var.get() if export_type_var.get() in EXPORT_TYPE_CHOICES else EXPORT_TYPE_CHOICES[0]
            self.save_settings()
            self.set_status("Settings saved.")
            dialog.destroy()

        row = tk.Frame(dialog, bg=UI_BG)
        row.pack(side="bottom", fill="x", padx=18, pady=(0, 18))
        self._button(row, "Save Settings", save_and_close, True, icon="icon-status.png").pack(side="right", padx=(8, 0))
        self._button(row, "Cancel", dialog.destroy, icon="icon-clear.png").pack(side="right")
        tk.Frame(dialog, bg=UI_RED, height=3).pack(side="bottom", fill="x", padx=18, pady=(0, 8))

    def open_path(self, path: Path, item_name: str = "item") -> None:
        try:
            os.startfile(path)  # type: ignore[attr-defined]
            self.set_status(f"Opened {item_name}: {path}")
        except OSError as exc:
            self.record_error_report(f"Open {item_name} failed", [f"Path: {path}", f"Error: {exc}"])
            messagebox.showerror("Open failed", f"Could not open this {item_name}.\n\n{exc}")

    def app_launch_target(self) -> Path:
        if getattr(sys, "frozen", False):
            return Path(sys.executable)
        built_exe = Path(__file__).resolve().parent / "dist" / "Macys_AP_China_Grove_Hex_Utility.exe"
        return built_exe if built_exe.exists() else Path(sys.executable)

    def create_desktop_shortcut(self) -> None:
        desktop = Path(os.environ.get("USERPROFILE", str(Path.home()))) / "Desktop"
        desktop.mkdir(parents=True, exist_ok=True)
        target = self.app_launch_target()
        icon = asset_path("app-icon.ico")
        shortcut_path = desktop / "Macy's AP China Grove Hex Utility.lnk"

        def ps_quote(value: str) -> str:
            return "'" + value.replace("'", "''") + "'"

        script = "\n".join(
            [
                "$shell = New-Object -ComObject WScript.Shell",
                f"$shortcut = $shell.CreateShortcut({ps_quote(str(shortcut_path))})",
                f"$shortcut.TargetPath = {ps_quote(str(target))}",
                f"$shortcut.WorkingDirectory = {ps_quote(str(target.parent))}",
                f"$shortcut.IconLocation = {ps_quote(str(icon))}",
                f"$shortcut.Description = {ps_quote(APP_DISPLAY_NAME)}",
                "$shortcut.Save()",
            ]
        )
        try:
            subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
                check=True,
                capture_output=True,
                text=True,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
            self.set_status(f"Desktop shortcut created: {shortcut_path}")
            messagebox.showinfo("Shortcut created", f"Desktop shortcut created:\n\n{shortcut_path}")
        except Exception:
            fallback = desktop / "Macy's AP China Grove Hex Utility.url"
            icon_text = str(icon).replace("\\", "/")
            target_text = str(target).replace("\\", "/")
            fallback.write_text(f"[InternetShortcut]\nURL=file:///{target_text}\nIconFile={icon_text}\nIconIndex=0\n", encoding="utf-8")
            self.set_status(f"Desktop launcher created: {fallback}")
            messagebox.showinfo("Launcher created", f"A desktop launcher was created:\n\n{fallback}")

    def open_bluewave(self) -> None:
        webbrowser.open(BLUEWAVE_URL)
        self.set_status("Opened BlueWave in your browser.")

    def render_history(self) -> None:
        if hasattr(self, "history_tree"):
            history = self.settings.get("history", [])
            for item_id in self.history_tree.get_children():
                self.history_tree.delete(item_id)
            if hasattr(self, "history_empty_frame"):
                if history:
                    self.history_empty_frame.pack_forget()
                elif not self.history_empty_frame.winfo_ismapped():
                    self.history_empty_frame.pack(fill="x", padx=12, pady=(0, 8), before=self.history_table_frame)
            for row_index, item in enumerate(history):
                invalid_count = int(item.get("invalid", 0) or 0)
                warning_count = int(item.get("warnings", 0) or 0)
                tags: tuple[str, ...] = ("history_even",) if row_index % 2 == 0 else ("history_odd",)
                if invalid_count:
                    tags = ("invalid",)
                elif warning_count:
                    tags = ("warning",)
                self.history_tree.insert(
                    "",
                    "end",
                    values=(
                        item.get("timestamp", ""),
                        item.get("action", ""),
                        item.get("total", 0),
                        item.get("valid", 0),
                        invalid_count,
                        warning_count,
                    ),
                    tags=tags,
                )
            return
        if not hasattr(self, "history_text"):
            return
        history = self.settings.get("history", [])
        lines = [f"{APP_SHORT_NAME} Conversion History", ""]
        if not history:
            lines.append("No conversion history yet.")
        for item in history:
            lines.append(
                f"{item.get('timestamp', '')} | {item.get('action', '')} | "
                f"input {item.get('total', 0)} | valid {item.get('valid', 0)} | "
                f"invalid {item.get('invalid', 0)} | warnings {item.get('warnings', 0)}"
            )
        self.history_text.configure(state="normal")
        self.history_text.delete("1.0", "end")
        self.history_text.insert("1.0", "\n".join(lines))
        self.history_text.configure(state="disabled")

    def clear_history(self) -> None:
        self.settings["history"] = []
        self.save_settings()
        self.render_history()
        self.set_status("Conversion history cleared.")

    def handle_single_scan_tab(self, _event: Any | None = None) -> str:
        self.convert_single(source="scan")
        return "break"

    def schedule_single_scan_autoconvert(self, event: Any | None = None) -> None:
        key = getattr(event, "keysym", "")
        if key in {"Return", "KP_Enter", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R", "Alt_L", "Alt_R"}:
            return
        if self._single_scan_after:
            self.after_cancel(self._single_scan_after)
        self._single_scan_after = self.after(260, self.auto_convert_single_if_ready)

    def auto_convert_single_if_ready(self) -> None:
        self._single_scan_after = None
        prepared = clean_candidate_line(self.single_var.get())
        if valid_hex8(prepared["extracted"].strip().upper()):
            self.convert_single(show_errors=False, source="scan")

    def convert_single(self, show_errors: bool = True, source: str = "manual") -> None:
        if self._single_scan_after:
            self.after_cancel(self._single_scan_after)
            self._single_scan_after = None
        prepared = clean_candidate_line(self.single_var.get())
        hex_value = prepared["extracted"].strip().upper()
        if not valid_hex8(hex_value):
            reason = invalid_reason(prepared["original"])
            self.set_status(f"Single lookup needs review: {reason}.", "Needs Review", target_tab=1)
            if show_errors:
                messagebox.showerror("Invalid hex", f"Single hex is invalid: {reason}.")
            return
        facility, card = hex_to_fc_cn(hex_value)
        notes = [*prepared["suggestions"], *unusual_warnings(hex_value, facility, card)]
        result_lines = [
            f"HEX  {hex_value}",
            f"FC   {facility}",
            f"CN   {card}",
        ]
        if notes:
            result_lines.append("NOTE " + " ".join(notes))
        self.single_result.set("\n".join(result_lines))
        if hasattr(self, "single_hex_var"):
            self.single_hex_var.set(hex_value)
            self.single_fc_var.set(str(facility))
            self.single_cn_var.set(str(card))
            self.single_note_var.set(" ".join(notes) if notes else "Clean lookup. FC,CN copied to clipboard.")
        self.clipboard_clear()
        self.clipboard_append(f"{facility},{card}")
        if hasattr(self, "single_entry"):
            self.single_entry.focus_set()
            self.single_entry.selection_range(0, "end")
        self.set_status("Single scan converted and FC,CN copied." if source == "scan" else "Single ID converted and FC,CN copied.", "Ready", target_tab=1)

    def clear_single(self) -> None:
        self.single_var.set("")
        self.single_result.set("Waiting for one 8-character HEX ID.")
        if hasattr(self, "single_hex_var"):
            self.single_hex_var.set("--")
            self.single_fc_var.set("--")
            self.single_cn_var.set("--")
            self.single_note_var.set("No lookup has run yet.")
        if hasattr(self, "single_entry"):
            self.single_entry.focus_set()
        self.set_status("Single lookup cleared.", "Ready", target_tab=1)

    def clear_reverse(self) -> None:
        self.fc_var.set("")
        self.cn_var.set("")
        self.reverse_result.set("Waiting for Facility Code and Card Number.")
        if hasattr(self, "reverse_hex_var"):
            self.reverse_hex_var.set("--")
            self.reverse_note_var.set("No reverse lookup has run yet.")
        if hasattr(self, "fc_entry"):
            self.fc_entry.focus_set()
        self.set_status("FC/CN to Hex cleared.", "Ready", target_tab=2)

    def convert_reverse(self) -> None:
        try:
            hex_value = fc_cn_to_hex(self.fc_var.get(), self.cn_var.get())
            warnings = unusual_warnings(hex_value, int(self.fc_var.get()), int(self.cn_var.get()))
        except ValueError as exc:
            self.set_status(f"FC/CN to Hex needs review: {exc}", "Needs Review", target_tab=2)
            messagebox.showerror("Invalid FC/CN", str(exc))
            return
        result_lines = [
            f"FC   {self.fc_var.get().strip()}",
            f"CN   {self.cn_var.get().strip()}",
            f"HEX  {hex_value}",
        ]
        if warnings:
            result_lines.append("NOTE " + " ".join(warnings))
        self.reverse_result.set("\n".join(result_lines))
        if hasattr(self, "reverse_hex_var"):
            self.reverse_hex_var.set(hex_value)
            self.reverse_note_var.set(" ".join(warnings) if warnings else "Clean reverse lookup. HEX copied to clipboard.")
        self.clipboard_clear()
        self.clipboard_append(hex_value)
        self.set_status("Hex value created and copied.", "Ready", target_tab=2)

    def confirm_paste_cleanup_preview(self, raw_text: str, cleanup: dict[str, Any]) -> str:
        preview = "\n".join(cleanup.get("lines", [])[:60])
        dialog = self._new_dialog("Paste Cleanup Preview", "#10141b")
        dialog.geometry("700x460")
        dialog.minsize(620, 380)
        choice = tk.StringVar(value="cancel")
        tk.Label(dialog, text="Paste Cleanup Preview", bg="#10141b", fg="#ffffff", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=14, pady=(14, 5))
        tk.Label(
            dialog,
            text=f"{cleanup.get('message', '')} Review the cleaned numeric IDs before adding them to the queue.",
            bg="#10141b",
            fg="#a8b2c2",
            wraplength=660,
            justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 8))
        preview_frame = tk.Frame(dialog, bg="#10141b")
        preview_frame.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)
        text = tk.Text(preview_frame, bg="#080a0f", fg="#f5f7fb", relief="flat", padx=10, pady=10, font=("Cascadia Mono", 10), wrap="none")
        preview_scroll_y = ttk.Scrollbar(preview_frame, orient="vertical", command=text.yview, style="Dark.Vertical.TScrollbar")
        preview_scroll_x = ttk.Scrollbar(preview_frame, orient="horizontal", command=text.xview, style="Dark.Horizontal.TScrollbar")
        text.configure(yscrollcommand=preview_scroll_y.set, xscrollcommand=preview_scroll_x.set)
        text.grid(row=0, column=0, sticky="nsew")
        preview_scroll_y.grid(row=0, column=1, sticky="ns")
        preview_scroll_x.grid(row=1, column=0, sticky="ew")
        text.insert("1.0", preview)
        text.configure(state="disabled")
        self._enable_mousewheel(text, text, text)
        self._dialog_footer_accent(dialog, padx=14, pady=(0, 8))
        row = tk.Frame(dialog, bg="#10141b")
        row.pack(fill="x", padx=14, pady=(0, 14))
        self._button(row, "Add Clean IDs", lambda: (choice.set("clean"), dialog.destroy()), True, icon="icon-import.png").pack(side="right", padx=(8, 0))
        self._button(row, "Paste Raw", lambda: (choice.set("raw"), dialog.destroy()), icon="icon-copy.png", tooltip="Paste the original clipboard text instead.").pack(side="right", padx=(8, 0))
        self._button(row, "Cancel", dialog.destroy, icon="icon-clear.png").pack(side="right")
        self._apply_corporate_skin(dialog)
        self.wait_window(dialog)
        return choice.get()

    def confirm_import_preview(self, path: Path, result: dict[str, Any], queue_lines: list[str] | None = None, value_label: str = "IDs") -> bool:
        lines = queue_lines if queue_lines is not None else import_result_queue_lines(result)
        preview = "\n".join(lines[:30])
        if not preview:
            messagebox.showinfo("Import preview", f"{result.get('message', '')}\n\nNo {value_label} were detected, so nothing was added.")
            return False
        dialog = self._new_dialog("Import Preview", "#10141b")
        dialog.geometry("680x420")
        approved = tk.BooleanVar(value=False)
        tk.Label(dialog, text=f"Preview: {path.name}", bg="#10141b", fg="#ffffff", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=14, pady=(14, 6))
        tk.Label(dialog, text=result.get("message", ""), bg="#10141b", fg="#a8b2c2", wraplength=640, justify="left").pack(anchor="w", padx=14, pady=(0, 8))
        preview_frame = tk.Frame(dialog, bg="#10141b")
        preview_frame.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)
        text = tk.Text(preview_frame, bg="#080a0f", fg="#f5f7fb", relief="flat", padx=10, pady=10, font=("Cascadia Mono", 10), wrap="none")
        preview_scroll_y = ttk.Scrollbar(preview_frame, orient="vertical", command=text.yview, style="Dark.Vertical.TScrollbar")
        preview_scroll_x = ttk.Scrollbar(preview_frame, orient="horizontal", command=text.xview, style="Dark.Horizontal.TScrollbar")
        text.configure(yscrollcommand=preview_scroll_y.set, xscrollcommand=preview_scroll_x.set)
        text.grid(row=0, column=0, sticky="nsew")
        preview_scroll_y.grid(row=0, column=1, sticky="ns")
        preview_scroll_x.grid(row=1, column=0, sticky="ew")
        text.insert("1.0", preview)
        text.configure(state="disabled")
        self._enable_mousewheel(text, text, text)
        self._dialog_footer_accent(dialog, padx=14, pady=(0, 8))
        row = tk.Frame(dialog, bg="#10141b")
        row.pack(fill="x", padx=14, pady=(0, 14))
        self._button(row, "Add To Queue", lambda: (approved.set(True), dialog.destroy()), True, icon="icon-import.png").pack(side="right", padx=(8, 0))
        self._button(row, "Cancel", dialog.destroy, icon="icon-clear.png").pack(side="right")
        self._apply_corporate_skin(dialog)
        self.wait_window(dialog)
        return bool(approved.get())

    def import_paths_to_queue(self, paths: list[Path], preview: bool = True) -> None:
        files = [path for path in paths if path.is_file()]
        if not files:
            messagebox.showinfo("No files", "Drop or choose one or more supported files.")
            return
        unconvert_mode = getattr(self, "active_tab_index", 0) == 3
        self.set_busy(f"Importing {len(files)} file(s)...")
        all_lines: list[str] = []
        imported = 0
        errors: list[str] = []
        messages: list[str] = []
        try:
            for path in files:
                self.set_busy(f"Reading {path.name}...")
                try:
                    result = import_unconvert_file(path) if unconvert_mode else import_structured_file(path)
                except Exception as exc:
                    errors.append(f"{path.name}: {exc}")
                    continue
                lines = [str(line) for line in result.get("lines", [])] if unconvert_mode else import_result_queue_lines(result)
                if preview and len(files) == 1:
                    self.clear_busy()
                    if not self.confirm_import_preview(path, result, lines, "FC/CN pairs" if unconvert_mode else "IDs"):
                        self.set_status("Import canceled.")
                        return
                    self.set_busy("Adding import to queue...")
                if lines:
                    all_lines.extend(lines)
                    imported += 1
                messages.append(f"{path.name}: {result.get('message', '')}")
                self.add_recent_file(path)
            if all_lines:
                if unconvert_mode:
                    self.append_text_to_unconvert("\n".join(all_lines))
                else:
                    self.append_text_to_queue("\n".join(all_lines))
            if errors:
                self.record_error_report("Import failed for one or more files", errors)
                messagebox.showwarning(
                    "Some imports failed",
                    "\n".join(errors[:8]) + "\n\nUse Help > Copy Last Error Report if you need to share troubleshooting details.",
                )
            if len(files) == 1 and messages:
                self.notice_var.set(messages[0])
                self.set_status(messages[0], "Needs Review" if errors else "Ready")
            else:
                item_label = "FC/CN pair(s)" if unconvert_mode else "ID row(s)"
                status = f"Imported {len(all_lines)} {item_label} from {imported} file(s)."
                if errors:
                    status += f" {len(errors)} file(s) failed."
                self.notice_var.set(status)
                self.set_status(status, "Needs Review" if errors else "Ready")
        finally:
            self.clear_busy()

    def import_file(self, path: Path | None = None) -> None:
        if getattr(self, "active_tab_index", 0) not in {0, 3}:
            messagebox.showinfo("Batch import only", "Import is available on Batch Converter and Unconvert Batch.")
            return
        if path is None:
            path_texts = filedialog.askopenfilenames(
                title="Import FC/CN pairs" if getattr(self, "active_tab_index", 0) == 3 else "Import access control data",
                filetypes=[
                    ("Supported files", "*.txt *.csv *.tsv *.xlsx *.xlsm *.xls *.xml *.html *.htm"),
                    ("Excel files", "*.xlsx *.xlsm *.xls"),
                    ("Text files", "*.txt *.csv *.tsv"),
                    ("Table files", "*.xml *.html *.htm"),
                    ("All files", "*.*"),
                ],
            )
            if not path_texts:
                return
            paths = [Path(item) for item in path_texts]
            self.import_paths_to_queue(paths, preview=len(paths) == 1)
            return
        self.import_paths_to_queue([path], preview=True)

    def _exporting_unconvert(self) -> bool:
        return getattr(self, "active_tab_index", 0) == 3

    def _export_workspace_name(self) -> str:
        return "Unconvert Batch" if self._exporting_unconvert() else "Batch Converter"

    def _export_report_title(self) -> str:
        return "Unconvert Report" if self._exporting_unconvert() else "Conversion Report"

    def _export_file_stem(self) -> str:
        return "Macys_AP_China_Grove_Unconvert_Report" if self._exporting_unconvert() else "Macys_AP_China_Grove_Hex_Utility"

    def _export_last_run(self) -> str:
        return self.last_unconverted_at if self._exporting_unconvert() else self.last_converted_at

    def _export_counts(self) -> tuple[int, int, int]:
        if self._exporting_unconvert():
            return (
                len(self.unconvert_results),
                len(self.unconvert_invalid),
                sum(len(row.warnings) for row in self.unconvert_results),
            )
        return (
            len(self.results),
            len(self.invalid),
            sum(len(row.warnings) for row in self.results),
        )

    def _ensure_export_ready(self) -> bool:
        if getattr(self, "active_tab_index", 0) not in {0, 3}:
            messagebox.showinfo("Batch export only", "Export is available from Batch Converter or Unconvert Batch.")
            return False
        if self._exporting_unconvert():
            if not self.unconvert_results and not self.unconvert_invalid:
                messagebox.showinfo("No report", "Run Unconvert Batch before exporting.")
                return False
            return True
        if not self.results and not self.invalid:
            messagebox.showinfo("No report", "Run a Batch Converter conversion before exporting.")
            return False
        return True

    def _export_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        if self._exporting_unconvert():
            for item in self.unconvert_results:
                records.append(
                    {
                        "line": item.line,
                        "hex": item.hex_value,
                        "facility": item.facility,
                        "card": item.card,
                        "status": "Warning" if item.warnings else "Valid",
                        "notes": " | ".join(item.warnings),
                        "converted_at": item.converted_at,
                    }
                )
            for item in self.unconvert_invalid:
                records.append(
                    {
                        "line": item.line,
                        "hex": item.raw,
                        "facility": "",
                        "card": "",
                        "status": "Invalid",
                        "notes": item.reason,
                        "converted_at": item.converted_at,
                    }
                )
            return sorted(records, key=lambda row: row["line"])
        for item in self.results:
            records.append(
                {
                    "line": item.line,
                    "hex": item.hex_value,
                    "facility": item.facility,
                    "card": item.card,
                    "status": "Warning" if item.warnings else "Valid",
                    "notes": " | ".join([*item.suggestions, *item.warnings]),
                    "converted_at": item.converted_at,
                }
            )
        for item in self.invalid:
            records.append(
                {
                    "line": item.line,
                    "hex": item.raw,
                    "facility": "",
                    "card": "",
                    "status": "Invalid",
                    "notes": item.reason,
                    "converted_at": item.converted_at,
                }
            )
        return sorted(records, key=lambda row: row["line"])

    def export_default(self) -> None:
        export_type = self.settings.get("default_export_type", EXPORT_TYPE_CHOICES[0])
        actions = {
            EXPORT_TYPE_CHOICES[0]: self.export_excel,
            EXPORT_TYPE_CHOICES[1]: self.export_csv,
            EXPORT_TYPE_CHOICES[2]: self.export_txt,
            EXPORT_TYPE_CHOICES[3]: self.export_pdf,
        }
        actions.get(export_type, self.export_excel)()

    def show_export_complete(self, path: Path, export_name: str) -> None:
        dialog = self._new_dialog("Export Complete")
        dialog.geometry("640x360")
        dialog.minsize(580, 320)

        self._dialog_header(dialog, "Export Complete", f"{export_name} report saved successfully.", UI_GREEN_TEXT)

        body = tk.Frame(dialog, bg=UI_BG)
        body.pack(fill="both", expand=True, padx=18, pady=18)
        card = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
        card.pack(fill="both", expand=True)
        tk.Frame(card, bg=UI_GREEN_TEXT, height=3).pack(fill="x")
        inner = tk.Frame(card, bg=UI_SURFACE)
        inner.pack(fill="both", expand=True, padx=14, pady=14)
        tk.Label(inner, text=path.name, bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 12, "bold"), anchor="w").pack(fill="x")
        tk.Label(inner, text=str(path), bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 9), anchor="w", justify="left", wraplength=560).pack(fill="x", pady=(5, 12))
        valid_count, invalid_count, warning_count = self._export_counts()
        detail = (
            f"{self._export_workspace_name()} | Version {APP_VERSION} | "
            f"{valid_count} valid | {invalid_count} invalid | "
            f"{warning_count} warning(s)"
        )
        tk.Label(
            inner,
            text=detail,
            bg=UI_SURFACE_ALT,
            fg=UI_TEXT,
            font=("Segoe UI", 10, "bold"),
            anchor="w",
            padx=10,
            pady=9,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
        ).pack(fill="x")

        self._dialog_footer_accent(dialog, pady=(0, 8))
        row = tk.Frame(dialog, bg=UI_BG)
        row.pack(fill="x", padx=18, pady=(0, 18))
        self._button(row, "Open File", lambda: self.open_path(path, "file"), True, icon="icon-status.png").pack(side="left", padx=(0, 8))
        self._button(row, "Open Folder", lambda: self.open_path(path.parent, "folder"), icon="icon-folder.png").pack(side="left")
        self._button(row, "Close", dialog.destroy).pack(side="right")

    def complete_export(self, path: Path, export_name: str) -> None:
        self.add_recent_export(path, export_name)
        self.set_status(f"Saved {export_name} report: {path}", "Exported")
        self.show_export_complete(path, export_name)

    def handle_export_failure(self, export_name: str, path: Path, exc: Exception) -> None:
        self.record_error_report(
            f"{export_name} export failed",
            [
                f"Export type: {export_name}",
                f"Target file: {path}",
                f"Error: {exc}",
            ],
        )
        messagebox.showerror(
            "Export failed",
            f"The {export_name} export could not be saved.\n\n{exc}\n\nUse Help > Copy Last Error Report if you need to share troubleshooting details.",
        )

    def export_excel(self) -> None:
        if not self._ensure_export_ready():
            return
        stamp = datetime.now().strftime("%Y-%m-%d")
        path_text = filedialog.asksaveasfilename(
            title="Save Excel report",
            defaultextension=".xlsx",
            initialdir=self.default_export_dir(),
            initialfile=f"{self._export_file_stem()}_{stamp}.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not path_text:
            return
        path = Path(path_text)
        try:
            self._write_excel(path)
        except Exception as exc:
            self.handle_export_failure("Excel Workbook", path, exc)
            return
        self.complete_export(path, "Excel Workbook")

    def export_csv(self) -> None:
        if not self._ensure_export_ready():
            return
        stamp = datetime.now().strftime("%Y-%m-%d")
        path_text = filedialog.asksaveasfilename(
            title="Save CSV report",
            defaultextension=".csv",
            initialdir=self.default_export_dir(),
            initialfile=f"{self._export_file_stem()}_{stamp}.csv",
            filetypes=[("CSV Report", "*.csv")],
        )
        if not path_text:
            return
        path = Path(path_text)
        try:
            self._write_csv(path)
        except Exception as exc:
            self.handle_export_failure("CSV", path, exc)
            return
        self.complete_export(path, "CSV")

    def _write_csv(self, path: Path) -> None:
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Line", "Hex ID / Raw Input", "Facility Code", "Card Number", "Status", "Notes / Details", "Converted At", "App Version"])
            for row in self._export_records():
                writer.writerow([row["line"], row["hex"], row["facility"], row["card"], row["status"], row["notes"], row["converted_at"], APP_VERSION])

    def export_pdf(self) -> None:
        if not self._ensure_export_ready():
            return
        stamp = datetime.now().strftime("%Y-%m-%d")
        path_text = filedialog.asksaveasfilename(
            title="Save PDF report",
            defaultextension=".pdf",
            initialdir=self.default_export_dir(),
            initialfile=f"{self._export_file_stem()}_{stamp}.pdf",
            filetypes=[("PDF Report", "*.pdf")],
        )
        if not path_text:
            return
        if SimpleDocTemplate is None:
            self.record_error_report("PDF export unavailable", ["PDF export needs reportlab installed."])
            messagebox.showerror("PDF export unavailable", "PDF export needs reportlab installed.")
            return
        path = Path(path_text)
        try:
            self._write_pdf(path)
        except Exception as exc:
            self.handle_export_failure("PDF", path, exc)
            return
        self.complete_export(path, "PDF")

    def _write_pdf(self, path: Path) -> None:
        doc = SimpleDocTemplate(str(path), pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("APTitle", parent=styles["Title"], textColor=colors.HexColor("#E51B2D"), fontName="Helvetica-Bold", fontSize=18, leading=22)
        subtitle_style = ParagraphStyle("APSubtitle", parent=styles["Normal"], textColor=colors.HexColor("#323B49"), fontSize=9, leading=12)
        cell_style = ParagraphStyle("APCell", parent=styles["BodyText"], fontSize=7, leading=9)
        story = [
            Paragraph(f"{APP_SHORT_NAME} {self._export_report_title()}", title_style),
            Paragraph(f"Macy's Asset Protection - China Grove | {self._export_workspace_name()} report", subtitle_style),
            Spacer(1, 12),
        ]
        valid_count, invalid_count, warning_count = self._export_counts()

        summary_data = [
            ["Generated", datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"), "Last Run", self._export_last_run() or "No run timestamp"],
            ["App Version", APP_VERSION, "Utility", APP_SHORT_NAME],
            ["Valid", str(valid_count), "Invalid", str(invalid_count)],
            ["Warnings", str(warning_count), "Total Input", str(valid_count + invalid_count)],
        ]
        summary = Table(summary_data, colWidths=[80, 210, 90, 210])
        summary.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F6F8FB")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111721")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD3DF")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.extend([summary, Spacer(1, 12)])

        data = [["Line", "Hex ID / Raw Input", "Facility Code", "Card Number", "Status", "Notes / Details"]]
        for row in self._export_records():
            data.append([
                row["line"],
                Paragraph(str(row["hex"]), cell_style),
                row["facility"],
                row["card"],
                row["status"],
                Paragraph(str(row["notes"] or ""), cell_style),
            ])
        table = Table(data, repeatRows=1, colWidths=[42, 120, 78, 78, 72, 354])
        style_commands: list[tuple[Any, ...]] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E51B2D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFFFFF")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F7F9FC")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD3DF")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        for idx, row in enumerate(self._export_records(), start=1):
            if row["status"] == "Invalid":
                style_commands.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FFF0E0")))
            elif row["status"] == "Warning":
                style_commands.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FFF7D7")))
        table.setStyle(TableStyle(style_commands))
        story.append(table)
        doc.build(story)

    def _write_excel(self, path: Path) -> None:
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        summary.sheet_view.showGridLines = False

        red_fill = PatternFill("solid", fgColor="E51B2D")
        dark_fill = PatternFill("solid", fgColor="111721")
        light_fill = PatternFill("solid", fgColor="F6F8FB")
        header_fill = PatternFill("solid", fgColor="E51B2D")
        valid_fill = PatternFill("solid", fgColor="EAF7EF")
        warning_fill = PatternFill("solid", fgColor="FFF4CC")
        invalid_fill = PatternFill("solid", fgColor="FFE8D2")
        zebra_fill = PatternFill("solid", fgColor="F7F9FC")
        thin = Side(style="thin", color="CBD3DF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        summary.column_dimensions["A"].width = 22
        summary.column_dimensions["B"].width = 32
        summary.column_dimensions["C"].width = 18
        summary.column_dimensions["D"].width = 28
        summary.merge_cells("A1:D1")
        title = summary["A1"]
        title.value = f"{APP_SHORT_NAME} {self._export_report_title()}"
        title.font = Font(bold=True, size=16, color="FFFFFF")
        title.alignment = Alignment(horizontal="center")
        title.fill = red_fill

        summary.merge_cells("A2:D2")
        subtitle = summary["A2"]
        subtitle.value = f"Macy's Asset Protection - China Grove | {self._export_workspace_name()}"
        subtitle.font = Font(bold=True, color="FFFFFF")
        subtitle.alignment = Alignment(horizontal="center")
        subtitle.fill = dark_fill

        valid_count, invalid_count, warning_count = self._export_counts()
        meta = [
            ("Generated", datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")),
            ("App Version", APP_VERSION),
            ("Last Run", self._export_last_run()),
            ("Total Input Lines", valid_count + invalid_count),
            ("Valid Lines", valid_count),
            ("Invalid Lines", invalid_count),
            ("Warnings", warning_count),
            ("Note", "Facility Code = high 16 bits; Card Number = low 16 bits"),
        ]
        row_index = 4
        for key, value in meta:
            key_cell = summary.cell(row_index, 1, key)
            value_cell = summary.cell(row_index, 2, value)
            key_cell.font = Font(bold=True, color="111721")
            key_cell.fill = light_fill
            value_cell.fill = light_fill
            key_cell.border = value_cell.border = border
            key_cell.alignment = Alignment(horizontal="left")
            value_cell.alignment = Alignment(horizontal="left", wrap_text=True)
            row_index += 1

        summary["A13"] = "Report Use"
        summary["A13"].font = Font(bold=True, color="E51B2D")
        summary["A14"] = "Review the Results sheet for valid, warning, and invalid rows. Use filters on the Status column to focus the report."
        summary["A14"].alignment = Alignment(wrap_text=True)
        summary.merge_cells("A14:D15")

        results_ws = wb.create_sheet("Results")
        results_ws.sheet_view.showGridLines = False
        widths = {"A": 8, "B": 22, "C": 16, "D": 16, "E": 14, "F": 48, "G": 24}
        for column, width in widths.items():
            results_ws.column_dimensions[column].width = width
        headers = ["Line", "Hex ID / Raw Input", "Facility Code", "Card Number", "Status", "Notes / Details", "Converted At"]
        row_index = 1
        for col, label in enumerate(headers, start=1):
            cell = results_ws.cell(row_index, col, label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row_index += 1

        for idx, row in enumerate(self._export_records(), start=1):
            values = [row["line"], row["hex"], row["facility"], row["card"], row["status"], row["notes"], row["converted_at"]]
            for col, value in enumerate(values, start=1):
                cell = results_ws.cell(row_index, col, value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col in {1, 3, 4, 5, 7} else "left", wrap_text=col == 6)
                if idx % 2 == 0:
                    cell.fill = zebra_fill
            status_cell = results_ws.cell(row_index, 5)
            if row["status"] == "Valid":
                status_cell.fill = valid_fill
            elif row["status"] == "Warning":
                status_cell.fill = warning_fill
            else:
                status_cell.fill = invalid_fill
            row_index += 1

        last_row = max(row_index - 1, 1)
        results_ws.freeze_panes = "A2"
        results_ws.auto_filter.ref = f"A1:G{last_row}"
        results_ws.sheet_properties.tabColor = "E51B2D"
        summary.sheet_properties.tabColor = "111721"
        wb.save(path)

    def export_txt(self) -> None:
        if not self._ensure_export_ready():
            return
        stamp = datetime.now().strftime("%Y-%m-%d")
        path_text = filedialog.asksaveasfilename(
            title="Save TXT report",
            defaultextension=".txt",
            initialdir=self.default_export_dir(),
            initialfile=f"{self._export_file_stem()}_{stamp}.txt",
            filetypes=[("Text Report", "*.txt")],
        )
        if not path_text:
            return
        path = Path(path_text)
        try:
            path.write_text(self._build_text_report(), encoding="utf-8")
        except Exception as exc:
            self.handle_export_failure("TXT", path, exc)
            return
        self.complete_export(path, "TXT")

    def _build_text_report(self) -> str:
        records = self._export_records()
        valid_count, invalid_count, warning_count = self._export_counts()
        lines = [
            f"{APP_SHORT_NAME} {self._export_report_title()}",
            f"Macy's Asset Protection - China Grove | {self._export_workspace_name()}",
            "=" * 88,
            f"Generated       : {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}",
            f"App Version     : {APP_VERSION}",
            f"Last Run        : {self._export_last_run() or 'No run timestamp'}",
            f"Valid Records   : {valid_count}",
            f"Invalid Lines   : {invalid_count}",
            f"Warnings        : {warning_count}",
            "",
            "Facility Code = high 16 bits; Card Number = low 16 bits",
            "=" * 88,
            "",
        ]
        if records:
            lines.append(f"{'LINE':<6} {'HEX / RAW INPUT':<20} {'FC':<10} {'CN':<10} {'STATUS':<10} NOTES / DETAILS")
            lines.append("-" * 108)
            for row in records:
                lines.append(
                    f"{str(row['line']):<6} {str(row['hex']):<20} "
                    f"{str(row['facility']):<10} {str(row['card']):<10} {row['status']:<10} {row['notes']}"
                )
        return "\n".join(lines) + "\n"

    def show_help(self) -> None:
        dialog = self._new_dialog("How To Use")
        dialog.geometry("900x660")
        dialog.minsize(820, 580)

        self._dialog_header(dialog, "How To Use This Utility", "Fast guide for imports, conversions, exports, and review actions.", UI_BLUE)

        content = tk.Frame(dialog, bg=UI_BG)
        content.pack(fill="both", expand=True, padx=16, pady=16)

        quick = self._card(content, bg=UI_SURFACE, border=UI_BORDER)
        quick.pack(side="left", fill="y", padx=(0, 14))
        quick.configure(width=260)
        quick.pack_propagate(False)
        tk.Frame(quick, bg=UI_RED, height=3).pack(fill="x")
        tk.Label(quick, text="Quick Start", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=14, pady=(12, 3))
        tk.Label(quick, text="Most work follows these steps.", bg=UI_SURFACE, fg=UI_MUTED, wraplength=220, justify="left").pack(anchor="w", padx=14, pady=(0, 8))

        quick_steps = [
            ("1", "Add IDs", "Scan, paste, import, or drag files onto the Input Queue."),
            ("2", "Convert", "Use Convert to create FC/CN results."),
            ("3", "Review", "Check warnings, invalid lines, and duplicate notices."),
            ("4", "Export", "Save Excel, CSV, TXT, or PDF from Export."),
        ]
        for number, title_text, body_text in quick_steps:
            step = tk.Frame(quick, bg=UI_SURFACE_ALT, highlightthickness=1, highlightbackground=UI_BORDER)
            step.pack(fill="x", padx=14, pady=(0, 8))
            tk.Label(step, text=number, bg=UI_RED, fg="#ffffff", width=3, font=("Segoe UI", 10, "bold")).pack(side="left", fill="y")
            text_box = tk.Frame(step, bg=UI_SURFACE_ALT)
            text_box.pack(side="left", fill="both", expand=True, padx=10, pady=6)
            tk.Label(text_box, text=title_text, bg=UI_SURFACE_ALT, fg=UI_TEXT, font=("Segoe UI", 9, "bold")).pack(anchor="w")
            tk.Label(text_box, text=body_text, bg=UI_SURFACE_ALT, fg=UI_MUTED, wraplength=175, justify="left", font=("Segoe UI", 8)).pack(anchor="w")

        contact = self._card(quick, bg=UI_RED_SOFT, border=UI_BORDER)
        contact.pack(fill="x", padx=14, pady=(4, 8), side="bottom")
        tk.Label(contact, text="Need help?", bg=UI_RED_SOFT, fg=UI_TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10, pady=(7, 1))
        tk.Label(
            contact,
            text="Open Help > About for contact, version, and project details.",
            bg=UI_RED_SOFT,
            fg=UI_MUTED,
            wraplength=210,
            justify="left",
            font=("Segoe UI", 9),
        ).pack(anchor="w", padx=10, pady=(0, 7))

        scroll_shell = tk.Frame(content, bg=UI_BG)
        scroll_shell.pack(side="left", fill="both", expand=True)
        canvas = tk.Canvas(scroll_shell, bg=UI_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_shell, orient="vertical", command=canvas.yview, style="Dark.Vertical.TScrollbar")
        cards = tk.Frame(canvas, bg=UI_BG)
        cards_id = canvas.create_window((0, 0), window=cards, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        cards.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(cards_id, width=event.width))
        self._enable_mousewheel(canvas, canvas)
        self._enable_mousewheel(cards, canvas)

        def add_help_card(title_text: str, body_text: str, accent: str) -> None:
            card = self._card(cards, bg=UI_SURFACE, border=UI_BORDER)
            card.pack(fill="x", pady=(0, 10))
            tk.Frame(card, bg=accent, height=3).pack(fill="x")
            body = tk.Frame(card, bg=UI_SURFACE)
            body.pack(fill="x", padx=14, pady=12)
            tk.Label(body, text=title_text, bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
            tk.Label(body, text=body_text, bg=UI_SURFACE, fg=UI_MUTED, wraplength=450, justify="left", font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 0))

        add_help_card("Import Options", "Use Import > Browse Files for TXT, CSV, TSV, XLS, XLSX, XLSM, XML, HTML, or HTM files. Batch Converter imports detected IDs. Unconvert Batch imports detected FC/CN pairs. Drag files directly onto a batch input area when you want the fastest import.", "#46d9ff")
        add_help_card("Scanner Input", "Most USB handheld scanners type like a keyboard. Click the Batch Scanner Input field or press F9, scan an ID, and use an Enter or Tab scanner suffix. Each new scan is cleaned, counted, and placed at the top of the Input Queue.", "#e51b2d")
        add_help_card("Batch Converter", "Paste HEX IDs one per line, scan IDs, or paste full employee lines. The app highlights valid rows, warning rows, and invalid rows directly in the queue before you convert.", "#e51b2d")
        add_help_card("Single Lookup", "Single Lookup also works with a scanner. Press F10 or focus the HEX ID field and scan one value; the app automatically converts it and copies the FC,CN pair. Click the HEX ID, Facility Code, or Card Number result to copy that one value.", "#0b66c3")
        add_help_card("Queue Cleanup", "Use Remove Duplicates to keep the first valid matching HEX ID. Use Keep Valid to remove rows that cannot be read as valid 8-character HEX IDs.", "#35d07f")
        add_help_card("Results Review", "Valid rows show HEX, Facility Code, Card Number, status, and Notes / Details. Notes stay blank for clean rows and appear when the app cleaned imported text, found a duplicate, flagged an unusual value, or explains why an input row is invalid.", "#f1b84b")
        add_help_card("Reverse Tools", "Use FC/CN to Hex for one pair. Click the HEX output number to copy it again after conversion. Use Unconvert Batch when you have many FC/CN pairs. Accepted batch examples include 34968,18199, tab-separated values, or FC 34968 CN 18199.", "#35d07f")
        add_help_card("Exports", "Use Export to save Excel, CSV, TXT, or PDF reports from Batch Converter or Unconvert Batch. Export Default uses your saved default report type. After saving, Open File and Open Folder are available from the completion window.", "#8beaff")
        add_help_card("Status Navigation", "The bottom status strip stays clickable for review messages. The sidebar Workspace Status box shows the active work area, and the Field Guide box rotates useful shortcuts with short explanations every 15 seconds.", "#f1b84b")
        add_help_card("Settings", "Use File > Settings to choose the default export type, default export folder, and create a desktop shortcut for the utility.", "#ff6d78")
        add_help_card("Shortcuts", "F9 focuses Batch Scanner Input, F10 focuses Single Lookup, Ctrl+I imports, Ctrl+R converts, Ctrl+E exports Excel, Ctrl+P exports PDF, Ctrl+F jumps to search, and Ctrl+L clears the workspace.", "#ff6d78")
        self._enable_mousewheel_tree(cards, canvas)

        self._dialog_footer_accent(dialog, padx=16, pady=(0, 8))
        row = tk.Frame(dialog, bg=UI_BG)
        row.pack(fill="x", padx=16, pady=(0, 16))
        self._button(row, "Close", dialog.destroy, True).pack(side="right")

    def show_about(self) -> None:
        dialog = self._new_dialog("About")
        dialog.geometry("820x720")
        dialog.minsize(760, 680)

        self._dialog_header(
            dialog,
            APP_SHORT_NAME,
            f"Professional access-control conversion utility | Version {APP_VERSION}",
            UI_RED,
        )

        body = tk.Frame(dialog, bg=UI_BG)
        body.pack(fill="both", expand=True, padx=18, pady=14)
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=1)

        intro = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
        intro.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        tk.Frame(intro, bg=UI_RED, height=3).pack(fill="x")
        intro_body = tk.Frame(intro, bg=UI_SURFACE)
        intro_body.pack(fill="x", padx=16, pady=14)
        logo = self._load_icon("macys-ap-icon.png", 58)
        if logo:
            tk.Label(intro_body, image=logo, bg=UI_SURFACE).pack(side="left", padx=(0, 14))
        intro_copy = tk.Frame(intro_body, bg=UI_SURFACE)
        intro_copy.pack(side="left", fill="x", expand=True)
        tk.Label(intro_copy, text="Macy's Asset Protection China Grove Hex Converter Utility", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 14, "bold")).pack(anchor="w")
        tk.Label(
            intro_copy,
            text="Created for Macy's Asset Protection operations in China Grove, North Carolina. It supports batch conversion, scanner-assisted lookup, reverse FC/CN conversion, and professional exports.",
            bg=UI_SURFACE,
            fg=UI_MUTED,
            wraplength=620,
            justify="left",
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(5, 0))
        chip_row = tk.Frame(intro_copy, bg=UI_SURFACE)
        chip_row.pack(anchor="w", pady=(10, 0))
        for chip_text, chip_color in [("Local desktop utility", UI_BLUE), ("Scanner ready", UI_GREEN_TEXT), ("Version " + APP_VERSION, UI_RED)]:
            tk.Label(
                chip_row,
                text=chip_text,
                bg=UI_SURFACE_ALT,
                fg=chip_color,
                font=("Segoe UI", 8, "bold"),
                padx=9,
                pady=4,
                highlightthickness=1,
                highlightbackground=UI_BORDER,
            ).pack(side="left", padx=(0, 6))

        def about_card(row_index: int, column: int, title_text: str, value_text: str, accent: str) -> None:
            card = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
            card.grid(row=row_index, column=column, sticky="nsew", padx=(0, 8) if column == 0 else (8, 0), pady=(0, 12))
            tk.Frame(card, bg=accent, height=3).pack(fill="x")
            inner = tk.Frame(card, bg=UI_SURFACE)
            inner.pack(fill="both", expand=True, padx=14, pady=12)
            tk.Label(inner, text=title_text, bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 8, "bold")).pack(anchor="w")
            tk.Label(inner, text=value_text, bg=UI_SURFACE, fg=UI_TEXT, wraplength=315, justify="left", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(5, 0))

        about_card(1, 0, "PURPOSE", "Convert access-control HEX IDs to Facility Code/Card Number values, and rebuild HEX IDs from FC/CN when needed.", UI_BLUE)
        about_card(1, 1, "SCANNER SUPPORT", "Batch Scanner Input and Single Lookup support keyboard-style handheld scanners with Enter or Tab suffixes.", UI_GREEN_TEXT)
        about_card(2, 0, "LOCATION", "Macy's Asset Protection - China Grove, North Carolina", UI_RED)
        about_card(2, 1, "VERSION", f"Version {APP_VERSION}. Current test build with scanner and layout polish.", UI_WARN_TEXT)

        contact = self._card(body, bg=UI_SURFACE, border=UI_BORDER)
        contact.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        tk.Frame(contact, bg=UI_GREEN_TEXT, height=3).pack(fill="x")
        contact_inner = tk.Frame(contact, bg=UI_SURFACE)
        contact_inner.pack(fill="x", padx=16, pady=14)
        tk.Label(contact_inner, text="Contact and Project", bg=UI_SURFACE, fg=UI_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(
            contact_inner,
            text="For testing notes, project files, and source history, use the GitHub project. For app questions, contact Christopher Schumacher.",
            bg=UI_SURFACE,
            fg=UI_MUTED,
            wraplength=700,
            justify="left",
            font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(4, 8))
        made = tk.Frame(contact_inner, bg=UI_SURFACE)
        made.pack(anchor="w", pady=(0, 8))
        tk.Label(made, text="Made by ", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 10, "bold")).pack(side="left")
        self._link_label(
            made,
            "Christopher Schumacher",
            f"mailto:{CONTACT_EMAIL}",
            UI_SURFACE,
            "Email Christopher Schumacher at Macy's.",
            font_size=10,
            bold=True,
            padx=0,
        ).pack(side="left")
        tk.Label(made, text=" - Asset Protection FLO", bg=UI_SURFACE, fg=UI_MUTED, font=("Segoe UI", 10, "bold")).pack(side="left")
        links = tk.Frame(contact_inner, bg=UI_SURFACE)
        links.pack(fill="x", pady=(2, 0))
        email_link = self._link_label(links, "Email Christopher Schumacher", f"mailto:{CONTACT_EMAIL}", UI_SURFACE, "Open an email draft.", font_size=10, bold=True, padx=0)
        email_link.pack(side="left", padx=(0, 18))
        github_link = self._link_label(links, "Open GitHub Project", PROJECT_URL, UI_SURFACE, "Open the GitHub project repository.", font_size=10, bold=True, padx=0)
        github_link.pack(side="left")

        self._dialog_footer_accent(dialog)
        row = tk.Frame(dialog, bg=UI_BG)
        row.pack(fill="x", padx=18, pady=(0, 16))
        self._button(row, "Close", dialog.destroy, True).pack(side="right")


def main() -> None:
    if "--self-test" in sys.argv:
        assert hex_to_fc_cn("88984717") == (34968, 18199)
        assert fc_cn_to_hex(34968, 18199) == "88984717"
        assert clean_candidate_line("Active Christopher Benson, 88984765")["extracted"] == "88984765"
        assert clean_candidate_line("Excel cell 88984765.0")["extracted"] == "88984765"
        assert clean_id_lines_from_text("deadbeef", numeric_only=True)["lines"] == []
        assert clean_id_lines_from_text("deadbeef")["lines"] == ["DEADBEEF"]
        assert extract_eight_digit_id(88984765.0) == "88984765"
        assert extract_eight_digit_id("Badge 8898-4765") == "88984765"
        cleaned_clipboard = clean_id_lines_from_text("Candidate Name\tColleague #\nChris Test\t88984765.0\nJordan Test\t8898-4130")
        assert cleaned_clipboard["lines"] == ["88984765", "88984130"]
        assert import_result_queue_lines({"lines": ["Chris Test, 88984765"], "found_rows": 1}) == ["88984765"]
        results, invalid = convert_lines("88984717\nBAD-LINE\n8898-4765", "TEST")
        assert len(results) == 2
        assert len(invalid) == 1
        unconverted, unconvert_invalid = unconvert_lines("34968,18199\nBAD-LINE", "TEST")
        assert len(unconverted) == 1
        assert len(unconvert_invalid) == 1
        assert unconverted[0].hex_value == "88984717"
        imported = extract_name_id_lines_from_tables(
            [[["Candidate Name", "Colleague #"], ["Christopher Benson", "88984765"], ["Supervisor", "88123456"]]],
            "sample.xlsx",
        )
        assert imported["lines"] == ["Christopher Benson, 88984765"]
        excel_imported = extract_name_id_lines_from_tables(
            [[["Candidate Name", "Colleague #"], ["Christopher Benson", 88984765.0], ["Jordan Smith", "ID 8898-4130"]]],
            "sample.xlsx",
        )
        assert excel_imported["lines"] == ["Christopher Benson, 88984765", "Jordan Smith, 88984130"]
        return
    app = ConverterApp()
    app.mainloop()


if __name__ == "__main__":
    main()
